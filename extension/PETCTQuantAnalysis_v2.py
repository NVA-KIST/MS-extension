"""
PETCTQuantAnalysis_v2.py
========================
3D Slicer Scripted Module — PET/CT Quantitative Analysis  (Version 2)

CHANGES FROM v1:
  - Segmentations are auto-discovered from the Segments folder (no fixed organ list).
  - After detection, a table shows every .nii.gz found across all patients,
    how many scans have each file, and lets the user tick/untick and rename
    each one for the Excel header.
  - runBatch iterates over the user-selected segmentation files dynamically.
  - Excel columns are named after the user-supplied display labels.
  - Full verbose logging throughout (same depth as v1).

DIRECTORY STRUCTURE (same as v1):
    root/
      CT/       {SubjectID}_{YYYY-MM-DD}_CT/   *.dcm
      PET/      {SubjectID}_{YYYY-MM-DD}_PET/  *.dcm
      Segments/ {SubjectID}_{YYYY-MM-DD}_Seg/
                  *.nii.gz  ← any number of segmentation files
"""

import os
import re
import logging
import vtk
import qt
import ctk
import slicer
from slicer.ScriptedLoadableModule import *
import openpyxl
from datetime import datetime
from pathlib import Path
from collections import defaultdict

LOG = logging.getLogger("PETCTQuant")

# Files that are pre-unchecked by default (internal / non-segmentation files).
# Users can still enable them manually.
_DEFAULT_SKIP = {
    "combined_seg", "body_trunc", "body_extremities",
    "body", "skin", "ct",
}


# ── MODULE REGISTRATION ────────────────────────────────────────────────────

class PETCTQuantAnalysis_v2(ScriptedLoadableModule):
    def __init__(self, parent):
        ScriptedLoadableModule.__init__(self, parent)
        parent.title       = "PET/CT Quantitative Analysis v2"
        parent.categories  = ["Quantification"]
        parent.dependencies = ["QuantitativeIndicesCLI"]
        parent.contributors = ["Research Lab"]
        parent.helpText = (
            "Batch PET/CT quantification. "
            "Automatically discovers all segmentation files and lets you "
            "rename them before writing to Excel."
        )


# ── WIDGET (UI) ────────────────────────────────────────────────────────────

class PETCTQuantAnalysis_v2Widget(ScriptedLoadableModuleWidget):

    def setup(self):
        ScriptedLoadableModuleWidget.setup(self)
        self.logic = PETCTQuantAnalysis_v2Logic()
        self._scans = []          # list of scan dicts from detectScans
        self._seg_rows = []       # list of (stem, cb_widget, name_edit)
        self._preview_nodes = []  # nodes loaded for the current preview

        # ── 1. DATA INPUT ──────────────────────────────────────────────
        col1 = ctk.ctkCollapsibleButton()
        col1.text = "1. Data Input"
        self.layout.addWidget(col1)
        lay1 = qt.QFormLayout(col1)

        self.rootEdit = ctk.ctkPathLineEdit()
        self.rootEdit.filters = ctk.ctkPathLineEdit.Dirs
        lay1.addRow("Root folder:", self.rootEdit)

        self.detectBtn = qt.QPushButton("Detect patients + segmentations")
        lay1.addRow("", self.detectBtn)

        self.scansLabel = qt.QLabel("No folder selected")
        lay1.addRow("Scans found:", self.scansLabel)

        # ── 2. SEGMENTATIONS FOUND ────────────────────────────────────
        col2 = ctk.ctkCollapsibleButton()
        col2.text = "2. Segmentations Found"
        self.layout.addWidget(col2)
        lay2 = qt.QVBoxLayout(col2)

        helpLabel = qt.QLabel(
            "Tick the files you want to analyse.\n"
            "Edit 'Excel label' to rename the column header "
            "(e.g. VF instead of visceral_fat)."
        )
        helpLabel.setWordWrap(True)
        lay2.addWidget(helpLabel)

        # Table: checkbox | filename | coverage | excel label
        self.segTable = qt.QTableWidget(0, 4)
        self.segTable.setHorizontalHeaderLabels(
            ["Include", "File (stem)", "Patients", "Excel label"]
        )
        self.segTable.horizontalHeader().setStretchLastSection(True)
        self.segTable.setSelectionMode(qt.QAbstractItemView.NoSelection)
        self.segTable.setFixedHeight(180)
        lay2.addWidget(self.segTable)

        # ── 3. COMPUTATION ────────────────────────────────────────────
        col3 = ctk.ctkCollapsibleButton()
        col3.text = "3. Computation"
        self.layout.addWidget(col3)
        lay3 = qt.QFormLayout(col3)

        suvWidget = qt.QWidget()
        suvBox = qt.QHBoxLayout(suvWidget)
        self.suvBwRadio  = qt.QRadioButton("SUVbw (body weight)")
        self.suvLbmRadio = qt.QRadioButton("SUVlbm (lean body mass)")
        self.suvBwRadio.setChecked(True)
        for r in (self.suvBwRadio, self.suvLbmRadio):
            suvBox.addWidget(r)
        lay3.addRow("SUV type:", suvWidget)

        mWidget = qt.QWidget()
        mBox = qt.QHBoxLayout(mWidget)
        self.cbMean   = qt.QCheckBox("SUVmean")
        self.cbMax    = qt.QCheckBox("SUVmax")
        self.cbPeak   = qt.QCheckBox("SUVpeak")
        self.cbTLG    = qt.QCheckBox("TLG")
        self.cbVolume = qt.QCheckBox("Volume")
        for cb in (self.cbMean, self.cbMax, self.cbPeak, self.cbTLG, self.cbVolume):
            cb.setChecked(True)
            mBox.addWidget(cb)
        lay3.addRow("Metrics:", mWidget)

        self.cbRegistration = qt.QCheckBox("CT-to-PET rigid registration")
        self.cbRegistration.setChecked(False)
        lay3.addRow("", self.cbRegistration)

        self.dilationSpinBox = qt.QDoubleSpinBox()
        self.dilationSpinBox.setRange(0.0, 20.0)
        self.dilationSpinBox.setSingleStep(1.0)
        self.dilationSpinBox.setValue(0.0)
        self.dilationSpinBox.setSuffix(" mm")
        lay3.addRow("Dilation margin (mm):", self.dilationSpinBox)

        # ── 4. EXPORT ─────────────────────────────────────────────────
        col4 = ctk.ctkCollapsibleButton()
        col4.text = "4. Export"
        self.layout.addWidget(col4)
        lay4 = qt.QFormLayout(col4)

        self.outputEdit = ctk.ctkPathLineEdit()
        self.outputEdit.filters   = ctk.ctkPathLineEdit.Files
        self.outputEdit.nameFilters = ["Excel files (*.xlsx)"]
        lay4.addRow("Output Excel:", self.outputEdit)

        self.appendCb = qt.QCheckBox("Append to existing file")
        self.appendCb.setChecked(True)
        lay4.addRow("", self.appendCb)

        # ── 5. PREVIEW ────────────────────────────────────────────────
        col5 = ctk.ctkCollapsibleButton()
        col5.text = "5. Preview — visual QC"
        self.layout.addWidget(col5)
        lay5 = qt.QFormLayout(col5)

        self.previewCombo = qt.QComboBox()
        self.previewCombo.setToolTip("Run 'Detect' first to populate this list")
        lay5.addRow("Patient:", self.previewCombo)

        previewBtnWidget = qt.QWidget()
        previewBtnBox    = qt.QHBoxLayout(previewBtnWidget)
        previewBtnBox.setContentsMargins(0, 0, 0, 0)
        self.loadPreviewBtn  = qt.QPushButton("Load & Preview")
        self.clearPreviewBtn = qt.QPushButton("Clear scene")
        previewBtnBox.addWidget(self.loadPreviewBtn)
        previewBtnBox.addWidget(self.clearPreviewBtn)
        lay5.addRow("", previewBtnWidget)

        self.previewStatusLabel = qt.QLabel("Select a patient and click Load & Preview.")
        self.previewStatusLabel.setWordWrap(True)
        lay5.addRow("Status:", self.previewStatusLabel)

        # ── RUN CONTROLS ──────────────────────────────────────────────
        self.startBtn = qt.QPushButton("START — run all patients")
        self.startBtn.setStyleSheet(
            "QPushButton{background:#2196F3;color:white;font-weight:bold;"
            "padding:8px;border-radius:4px}"
            "QPushButton:hover{background:#1976D2}"
            "QPushButton:disabled{background:#BDBDBD}"
        )
        self.layout.addWidget(self.startBtn)

        self.cancelBtn = qt.QPushButton("Cancel")
        self.cancelBtn.setEnabled(False)
        self.layout.addWidget(self.cancelBtn)

        self.progressBar = qt.QProgressBar()
        self.progressBar.setRange(0, 100)
        self.layout.addWidget(self.progressBar)

        self.statusLabel = qt.QLabel("Ready")
        self.statusLabel.setWordWrap(True)
        self.layout.addWidget(self.statusLabel)

        self.layout.addStretch(1)

        # connections
        self.detectBtn.connect('clicked(bool)', self.onDetect)
        self.startBtn.connect('clicked(bool)', self.onStart)
        self.cancelBtn.connect('clicked(bool)', self.onCancel)
        self.loadPreviewBtn.connect('clicked(bool)', self.onPreview)
        self.clearPreviewBtn.connect('clicked(bool)', self.onClearPreview)
        self._cancel = False

    # ── Detection ─────────────────────────────────────────────────────────

    def onDetect(self):
        root = self.rootEdit.currentPath
        if not root or not os.path.isdir(root):
            self.scansLabel.setText("Invalid folder")
            return

        self._scans = self.logic.detectScans(root)
        n_subj = len(set(s["subject_id"] for s in self._scans))
        self.scansLabel.setText(
            f"{len(self._scans)} scan(s) across {n_subj} subject(s)"
        )

        seg_info = self.logic.detectSegmentations(root, self._scans)
        self._buildSegTable(seg_info)

        self.previewCombo.clear()
        for s in self._scans:
            self.previewCombo.addItem(f"{s['subject_id']}  |  {s['scan_date']}")

    def _buildSegTable(self, seg_info):
        """
        seg_info: dict  stem → {"count": int, "total": int}
        Fills the table with one row per discovered file stem.
        """
        self._seg_rows.clear()
        self.segTable.setRowCount(0)
        total = seg_info.get("__total__", 1)

        for stem, info in sorted(seg_info.items()):
            if stem == "__total__":
                continue

            row_idx = self.segTable.rowCount
            self.segTable.insertRow(row_idx)

            # Checkbox
            cb = qt.QCheckBox()
            cb.setChecked(stem not in _DEFAULT_SKIP)
            cb_widget = qt.QWidget()
            cb_layout = qt.QHBoxLayout(cb_widget)
            cb_layout.addWidget(cb)
            cb_layout.setAlignment(qt.Qt.AlignCenter)
            cb_layout.setContentsMargins(0, 0, 0, 0)
            self.segTable.setCellWidget(row_idx, 0, cb_widget)

            # Filename stem (read-only)
            item_name = qt.QTableWidgetItem(stem)
            item_name.setFlags(qt.Qt.ItemIsEnabled)
            self.segTable.setItem(row_idx, 1, item_name)

            # Coverage
            coverage = f"{info['count']}/{info['total']}"
            item_cov = qt.QTableWidgetItem(coverage)
            item_cov.setFlags(qt.Qt.ItemIsEnabled)
            self.segTable.setItem(row_idx, 2, item_cov)

            # Editable display name
            name_edit = qt.QLineEdit(stem)
            self.segTable.setCellWidget(row_idx, 3, name_edit)

            self._seg_rows.append((stem, cb, name_edit))

        self.segTable.resizeColumnsToContents()

    def _getSegNameMap(self):
        """
        Returns {stem: display_name} for all checked rows.
        display_name defaults to stem if the edit is blank.
        """
        name_map = {}
        for stem, cb, name_edit in self._seg_rows:
            if cb.isChecked():
                label = name_edit.text.strip() or stem
                name_map[stem] = label
        return name_map

    # ── Run ───────────────────────────────────────────────────────────────

    def onStart(self):
        root        = self.rootEdit.currentPath
        output_file = self.outputEdit.currentPath

        if not root or not os.path.isdir(root):
            slicer.util.errorDisplay("Please select a valid root folder.")
            return
        if not output_file:
            slicer.util.errorDisplay("Please select an output Excel file.")
            return
        if not self._scans:
            slicer.util.errorDisplay("Click 'Detect' first.")
            return

        seg_name_map = self._getSegNameMap()
        if not seg_name_map:
            slicer.util.errorDisplay(
                "No segmentations selected. "
                "Run Detect and tick at least one segmentation file."
            )
            return

        metrics = {
            "mean":   self.cbMean.isChecked(),
            "max":    self.cbMax.isChecked(),
            "peak":   self.cbPeak.isChecked(),
            "tlg":    self.cbTLG.isChecked(),
            "volume": self.cbVolume.isChecked(),
        }
        suv_type = "bw" if self.suvBwRadio.isChecked() else "lbm"

        self.startBtn.setEnabled(False)
        self.cancelBtn.setEnabled(True)
        self._cancel = False

        try:
            self.logic.runBatch(
                root_folder     = root,
                scans           = self._scans,
                seg_name_map    = seg_name_map,
                output_file     = output_file,
                metrics         = metrics,
                suv_type        = suv_type,
                append          = self.appendCb.isChecked(),
                progress_cb     = lambda v: (self.progressBar.setValue(int(v)),
                                             slicer.app.processEvents()),
                status_cb       = lambda m: (self.statusLabel.setText(m),
                                             slicer.app.processEvents()),
                cancel_check    = lambda: self._cancel,
                do_registration = self.cbRegistration.isChecked(),
                dilation_mm     = self.dilationSpinBox.value,
            )
            slicer.util.infoDisplay(f"Done!\nResults saved to:\n{output_file}")
        except Exception as e:
            slicer.util.errorDisplay(f"Error: {e}")
            LOG.error(f"[PETCTQuant-v2] top-level error: {e}", exc_info=True)
        finally:
            self.startBtn.setEnabled(True)
            self.cancelBtn.setEnabled(False)

    def onCancel(self):
        self._cancel = True
        self.statusLabel.setText("Cancelling after current scan …")

    # ── Preview ───────────────────────────────────────────────────────────

    def onClearPreview(self):
        for node in self._preview_nodes:
            try:
                slicer.mrmlScene.RemoveNode(node)
            except Exception:
                pass
        self._preview_nodes.clear()
        self.previewStatusLabel.setText("Scene cleared.")

    def onPreview(self):
        idx = self.previewCombo.currentIndex
        if idx < 0 or not self._scans:
            slicer.util.errorDisplay(
                "Run 'Detect patients' first, then select a patient from the dropdown."
            )
            return

        seg_name_map = self._getSegNameMap()
        if not seg_name_map:
            slicer.util.errorDisplay(
                "Tick at least one segmentation in section 2 before previewing."
            )
            return

        scan     = self._scans[idx]
        seg_path = scan.get("seg_path") or ""

        self.onClearPreview()

        # ── Load PET ──────────────────────────────────────────────────
        self.previewStatusLabel.setText("Loading PET DICOM…")
        slicer.app.processEvents()
        try:
            pet_node = self.logic._loadDicomSeries(
                scan["pet_path"], scan["subject_id"]
            )
        except Exception as e:
            slicer.util.errorDisplay(f"Failed to load PET: {e}")
            self.previewStatusLabel.setText("PET load failed.")
            return

        self._preview_nodes.append(pet_node)

        # Make PET the background in all slice viewers
        slicer.util.setSliceViewerLayers(background=pet_node)
        for name in slicer.app.layoutManager().sliceViewNames():
            slicer.app.layoutManager().sliceWidget(name).sliceLogic().FitSliceToAll()

        # ── Fiducial node for hotspots ─────────────────────────────────
        fid_node = slicer.mrmlScene.AddNewNodeByClass(
            'vtkMRMLMarkupsFiducialNode', 'PETCTQuant_Hotspots'
        )
        self._preview_nodes.append(fid_node)
        disp = fid_node.GetDisplayNode()
        if disp:
            disp.SetTextScale(3.5)
            disp.SetGlyphScale(3.0)
            disp.SetColor(1.0, 1.0, 0.0)   # yellow

        # ── Process each selected segmentation ────────────────────────
        loaded_segs = 0
        for stem, display_name in seg_name_map.items():
            seg_file = os.path.join(seg_path, f"{stem}.nii.gz") if seg_path else None
            if not seg_file or not os.path.isfile(seg_file):
                continue

            self.previewStatusLabel.setText(f"Finding hotspot: {display_name}…")
            slicer.app.processEvents()

            try:
                seg_node = self.logic._loadSegmentation(seg_file)
                self._preview_nodes.append(seg_node)

                segmentation = seg_node.GetSegmentation()
                segment_id   = segmentation.GetNthSegmentID(0)

                label_node = slicer.mrmlScene.AddNewNodeByClass(
                    'vtkMRMLLabelMapVolumeNode', f'prev_lbl_{stem}'
                )
                seg_ids = vtk.vtkStringArray()
                seg_ids.InsertNextValue(segment_id)
                slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                    seg_node, seg_ids, label_node, pet_node,
                    slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
                )

                hotspots = self.logic._findHottestVoxels(pet_node, label_node, top_n=1)
                slicer.mrmlScene.RemoveNode(label_node)

                if hotspots:
                    h   = hotspots[0]
                    lbl = f"{display_name}: SUV {h['suv']:.2f}"
                    try:
                        pt = fid_node.AddControlPoint(
                            vtk.vtkVector3d(h["ras_x"], h["ras_y"], h["ras_z"])
                        )
                        fid_node.SetNthControlPointLabel(pt, lbl)
                    except AttributeError:
                        pt = fid_node.AddFiducial(h["ras_x"], h["ras_y"], h["ras_z"])
                        fid_node.SetNthFiducialLabel(pt, lbl)

                loaded_segs += 1

            except Exception as e:
                LOG.warning(f"[PETCTQuant-v2] Preview — {display_name} failed: {e}")

        try:
            n_pts = fid_node.GetNumberOfControlPoints()
        except AttributeError:
            n_pts = fid_node.GetNumberOfFiducials()

        self.previewStatusLabel.setText(
            f"Ready — PET + {loaded_segs} seg(s), {n_pts} hotspot fiducial(s) placed."
        )


# ── LOGIC ──────────────────────────────────────────────────────────────────

class PETCTQuantAnalysis_v2Logic(ScriptedLoadableModuleLogic):

    # ── Logging helper ────────────────────────────────────────────────────

    def _log(self, level, msg, status_cb=None):
        """
        Unified logger: writes to the module logger (visible in Slicer's
        application log via Help → Report a Bug) AND prints to the Python
        console so you see it live without opening the log file.

        level: 'info' | 'warning' | 'error'
        """
        tagged = f"[PETCTQuant-v2] {msg}"
        if level == "error":
            LOG.error(tagged)
        elif level == "warning":
            LOG.warning(tagged)
        else:
            LOG.info(tagged)
        print(tagged)
        if status_cb:
            status_cb(msg)

    def _parseFolderName(self, name):
        m = re.match(r'^(.+)_(\d{4}-\d{2}-\d{2})_(CT|PET|Seg)$', name)
        return (m.group(1), m.group(2), m.group(3)) if m else None

    # ── Discovery ─────────────────────────────────────────────────────────

    def detectScans(self, root_folder):
        """Walk root/PET/ and build the list of scan dicts."""
        self._log("info", f"detectScans: scanning root folder: {root_folder}")

        pet_root = os.path.join(root_folder, "PET")
        ct_root  = os.path.join(root_folder, "CT")
        seg_root = os.path.join(root_folder, "Segments")

        if not os.path.isdir(pet_root):
            self._log("error", f"detectScans: PET/ folder not found in {root_folder}")
            raise ValueError(f"No PET/ folder found in {root_folder}")

        self._log("info", f"detectScans: PET/ found — {pet_root}")

        scans = []
        for folder in sorted(os.listdir(pet_root)):
            fpath = os.path.join(pet_root, folder)
            if not os.path.isdir(fpath):
                continue

            parsed = self._parseFolderName(folder)
            if not parsed:
                self._log("warning",
                    f"detectScans: skipping unrecognised folder name: {folder}")
                continue

            subj, date, _ = parsed
            ct_p  = os.path.join(ct_root,  f"{subj}_{date}_CT")
            seg_p = os.path.join(seg_root, f"{subj}_{date}_Seg")

            ct_found  = os.path.isdir(ct_p)
            seg_found = os.path.isdir(seg_p)

            self._log("info",
                f"detectScans: found {subj} {date} | "
                f"CT: {'OK' if ct_found else 'MISSING'}  "
                f"Seg: {'OK' if seg_found else 'MISSING'}")

            scans.append({
                "subject_id": subj,
                "scan_date":  date,
                "pet_path":   fpath,
                "ct_path":    ct_p  if ct_found  else None,
                "seg_path":   seg_p if seg_found else None,
            })

        self._log("info",
            f"detectScans: complete — {len(scans)} scan(s) found "
            f"across {len(set(s['subject_id'] for s in scans))} subject(s)")
        return scans

    def detectSegmentations(self, root_folder, scans):
        """
        Walk every known Seg folder and collect all .nii.gz stems.
        Returns dict:
            stem → {"count": n_scans_with_file, "total": total_scans}
            "__total__" → total scan count (convenience key)
        """
        self._log("info",
            f"detectSegmentations: scanning {len(scans)} seg folder(s)...")

        stem_counts = defaultdict(int)
        total = len(scans)
        missing_seg = 0

        for scan in scans:
            seg_path = scan.get("seg_path")
            if not seg_path or not os.path.isdir(seg_path):
                missing_seg += 1
                self._log("warning",
                    f"detectSegmentations: no Seg folder for "
                    f"{scan['subject_id']} {scan['scan_date']} — skipping")
                continue

            files_found = [f for f in os.listdir(seg_path) if f.endswith(".nii.gz")]
            self._log("info",
                f"detectSegmentations: {scan['subject_id']} {scan['scan_date']} "
                f"— {len(files_found)} .nii.gz file(s): "
                f"{[f[:-len('.nii.gz')] for f in files_found]}")

            for fname in files_found:
                stem = fname[:-len(".nii.gz")]
                stem_counts[stem] += 1

        result = {stem: {"count": cnt, "total": total}
                  for stem, cnt in stem_counts.items()}
        result["__total__"] = total

        self._log("info",
            f"detectSegmentations: complete — "
            f"{len(result) - 1} unique stem(s) found  "
            f"({missing_seg} scan(s) had no Seg folder)")
        for stem, info in sorted(result.items()):
            if stem == "__total__":
                continue
            self._log("info",
                f"  stem '{stem}': present in {info['count']}/{info['total']} scans")

        return result

    # ── Batch ─────────────────────────────────────────────────────────────

    def runBatch(self, root_folder, scans, seg_name_map,
                 output_file, metrics, suv_type, append,
                 progress_cb, status_cb, cancel_check,
                 do_registration=False, dilation_mm=0.0):
        """
        seg_name_map: {stem: display_name}  — only these files are processed.
        """
        total    = len(scans)
        all_rows = []

        self._log("info",
            f"runBatch: starting  scans={total}  "
            f"segs={list(seg_name_map.keys())}  "
            f"suv_type={suv_type}  append={append}",
            status_cb)

        if total == 0:
            self._log("error", "runBatch: no valid scans found — aborting", status_cb)
            raise ValueError("No valid scans found in root folder.")

        for i, scan in enumerate(scans):
            if cancel_check():
                self._log("warning", "runBatch: cancelled by user", status_cb)
                break

            subj     = scan["subject_id"]
            date     = scan["scan_date"]
            pet_path = scan["pet_path"]
            seg_path = scan["seg_path"]

            self._log("info",
                f"── Patient {i+1}/{total}: {subj}  {date} ──────",
                status_cb)
            progress_cb(i / total * 100)

            # ── Step 1: Load DICOM PET ─────────────────────────────────
            self._log("info",
                f"  Step 1/4 — loading PET DICOM from: {pet_path}", status_cb)
            try:
                pet_node = self._loadDicomSeries(pet_path, subj)
                self._log("info",
                    f"  Step 1/4 — PET DICOM loaded OK  "
                    f"node={pet_node.GetName()}  id={pet_node.GetID()}")
            except Exception as e:
                self._log("error",
                    f"  Step 1/4 FAILED — could not load PET DICOM: {e}", status_cb)
                for stem in seg_name_map:
                    all_rows.append(
                        self._errorRow(subj, date, seg_name_map[stem], "missing_pet"))
                continue

            patient_id = self._getPatientIdFromNode(pet_node)
            self._log("info", f"  patient_id extracted: {patient_id}")

            # ── Step 1b: Convert Bq/mL → SUVbw if needed ──────────────
            self._log("info", "  Step 1b — applying SUV conversion...", status_cb)
            try:
                self._applySuvConversion(pet_node, pet_path)
            except Exception as e:
                self._log("error",
                    f"  Step 1b FAILED — SUV conversion error: {e}  "
                    f"Results will be in raw DICOM units (Bq/mL), not SUV",
                    status_cb)

            # ── Step 1c: CT-to-PET rigid registration (optional) ──────────────
            ct_node             = None
            ct_to_pet_transform = None
            if do_registration:
                ct_path = scan.get("ct_path")
                if ct_path:
                    self._log("info",
                        f"  Step 1c — loading CT DICOM for registration: {ct_path}",
                        status_cb)
                    try:
                        ct_node = self._loadDicomSeries(ct_path, subj)
                        ct_node.SetName(f"CT_{subj}")
                        self._log("info",
                            f"  Step 1c — CT loaded OK  node={ct_node.GetName()}")
                        ct_to_pet_transform = self._registerCtToPet(ct_node, pet_node)
                    except Exception as e:
                        self._log("error",
                            f"  Step 1c FAILED — registration skipped: {e}", status_cb)
                        if ct_node:
                            slicer.mrmlScene.RemoveNode(ct_node)
                        ct_node             = None
                        ct_to_pet_transform = None
                else:
                    self._log("warning",
                        "  Step 1c — do_registration=True but no CT folder found "
                        "— skipping", status_cb)

            # ── Step 2: (TotalSegmentator skipped in v2 — segs assumed present)
            self._log("info", "  Step 2/4 — skipped (using existing segments)")

            # ── Step 3–4: For each selected segmentation ───────────────
            self._log("info",
                f"  Step 3–4/4 — computing metrics for: "
                f"{list(seg_name_map.values())}")

            for stem, display_name in seg_name_map.items():
                seg_file = os.path.join(seg_path or "", f"{stem}.nii.gz") \
                           if seg_path else None

                self._log("info",
                    f"    [{display_name}] looking for: {seg_file}")

                if not seg_file or not os.path.isfile(seg_file):
                    self._log("warning",
                        f"    [{display_name}] MISSING segment file — recording error row",
                        status_cb)
                    all_rows.append(
                        self._errorRow(subj, date, display_name,
                                       "missing_seg", patient_id))
                    continue

                self._log("info",
                    f"    [{display_name}] segment file found — loading...",
                    status_cb)

                try:
                    seg_node = self._loadSegmentation(seg_file)
                    self._log("info",
                        f"    [{display_name}] segmentation loaded OK  "
                        f"node={seg_node.GetName()}")

                    if ct_to_pet_transform is not None:
                        self._log("info",
                            f"    [{display_name}] applying CT-to-PET transform...",
                            status_cb)
                        self._applyTransformToSeg(seg_node, ct_to_pet_transform)

                    if dilation_mm > 0:
                        self._log("info",
                            f"    [{display_name}] dilating mask by {dilation_mm} mm...",
                            status_cb)
                        self._dilateMask(seg_node, dilation_mm)

                    self._log("info",
                        f"    [{display_name}] running pet-indic CLI...", status_cb)
                    results = self._runPetIndic(pet_node, seg_node, metrics, suv_type)
                    slicer.mrmlScene.RemoveNode(seg_node)
                    self._log("info",
                        f"    [{display_name}] seg node removed from scene")

                    row = {
                        "subject_id":  subj,
                        "patient_id":  patient_id,
                        "scan_date":   date,
                        "segment":     display_name,
                        "source_file": f"{stem}.nii.gz",
                        "status":      "done",
                    }
                    row.update(results)
                    all_rows.append(row)

                    self._log("info",
                        f"    [{display_name}] DONE  "
                        f"SUVmean={results.get('suv_mean','?')}  "
                        f"SUVmax={results.get('suv_max','?')}  "
                        f"SUVpeak={results.get('suv_peak','?')}  "
                        f"TLG={results.get('tlg','?')}  "
                        f"vol={results.get('volume_mL','?')} mL")

                except Exception as e:
                    self._log("error",
                        f"    [{display_name}] FAILED — {e}", status_cb)
                    all_rows.append(
                        self._errorRow(subj, date, display_name,
                                       f"error: {str(e)[:80]}", patient_id))

            if ct_to_pet_transform is not None:
                slicer.mrmlScene.RemoveNode(ct_to_pet_transform)
                self._log("info", "  CT-to-PET transform node removed from scene")
            if ct_node is not None:
                slicer.mrmlScene.RemoveNode(ct_node)
                self._log("info", "  CT node removed from scene")
            slicer.mrmlScene.RemoveNode(pet_node)
            self._log("info",
                f"  PET node removed from scene — memory freed for next patient")
            progress_cb((i + 1) / total * 100)

        done_count  = sum(1 for r in all_rows if r.get("status") == "done")
        error_count = len(all_rows) - done_count
        self._log("info",
            f"runBatch: all patients processed — "
            f"{done_count} OK  {error_count} errors  "
            f"saving to {output_file}",
            status_cb)
        self._saveExcel(all_rows, output_file, append)
        self._log("info",
            f"runBatch: COMPLETE — {len(all_rows)} rows written to {output_file}",
            status_cb)

    # ── Slicer helpers ────────────────────────────────────────────────────

    def _loadDicomSeries(self, dicom_folder, subject_id):
        """
        Load a DICOM series from folder into Slicer.
        Always purges any stale DB entry for this folder first so that
        every run behaves like a clean-DB run (avoids ITK/GDCM re-open errors).
        """
        from DICOMLib import DICOMUtils

        dicom_files = []
        for root, _, files in os.walk(dicom_folder):
            for f in files:
                dicom_files.append(os.path.join(root, f))

        self._log("info",
            f"    _loadDicomSeries: {len(dicom_files)} file(s) found in {dicom_folder}")

        if not dicom_files:
            raise ValueError(f"No DICOM files in {dicom_folder}")

        db = slicer.dicomDatabase
        if not db.isOpen:
            default_db_path = os.path.join(slicer.app.temporaryPath, "CtkDicomDatabase")
            db.openDatabase(default_db_path)
            self._log("info", f"runBatch: opened DICOM DB at {default_db_path}", status_cb)
        
        
        def _all_series_uids():
            uids = set()
            for patient in db.patients():
                for study in db.studiesForPatient(patient):
                    for series in db.seriesForStudy(study):
                        uids.add(series)
            return uids

        # Purge stale DB entries for this folder
        folder_variants = set([
            os.path.normpath(dicom_folder).lower(),
            os.path.normpath(os.path.realpath(dicom_folder)).lower(),
        ])
        purged = []
        for uid in list(_all_series_uids()):
            files_in_db = db.filesForSeries(uid)
            if not files_in_db:
                continue
            file_norm = os.path.normpath(files_in_db[0]).lower()
            if any(file_norm.startswith(v) for v in folder_variants):
                try:
                    db.removeSeries(uid)
                    purged.append(uid)
                except Exception as e:
                    self._log("warning",
                        f"    _loadDicomSeries: could not remove UID={uid}: {e}")

        if purged:
            self._log("info",
                f"    _loadDicomSeries: purged {len(purged)} stale DB entry/entries: {purged}")
        else:
            self._log("info",
                "    _loadDicomSeries: no existing DB entries for this folder — clean import")

        # Fresh import
        uids_before = _all_series_uids()
        self._log("info", "    _loadDicomSeries: importing into DICOM database...")
        DICOMUtils.importDicom(dicom_folder)

        uids_after = _all_series_uids()
        new_uids   = list(uids_after - uids_before)
        self._log("info",
            f"    _loadDicomSeries: import done — "
            f"{len(new_uids)} new series UID(s): {new_uids}")

        if not new_uids:
            raise ValueError(
                f"importDicom produced no new series UIDs from {dicom_folder}. "
                f"Folder may be empty, contain no valid DICOM files, "
                f"or all files failed to parse."
            )

        # Load into scene
        self._log("info",
            f"    _loadDicomSeries: loading series into scene: {new_uids}...")
        loaded_node_ids = DICOMUtils.loadSeriesByUID(new_uids)
        self._log("info",
            f"    _loadDicomSeries: loadSeriesByUID returned "
            f"{len(loaded_node_ids)} node id(s): {loaded_node_ids}")

        if not loaded_node_ids:
            raise ValueError(
                f"loadSeriesByUID returned no nodes for UIDs {new_uids}. "
                f"Check the VTK/ITK errors above for the root cause."
            )

        node = slicer.mrmlScene.GetNodeByID(loaded_node_ids[0])
        if not node:
            raise ValueError(
                f"Node '{loaded_node_ids[0]}' not found in scene — "
                "DICOM may have loaded as a non-volume type (check series modality)"
            )

        node.SetName(f"PET_{subject_id}")
        self._log("info",
            f"    _loadDicomSeries: volume node ready — "
            f"name={node.GetName()}  id={node.GetID()}")
        return node

    def _getPatientIdFromNode(self, volume_node):
        """Read patient ID from DICOM tag 0010,0020. Falls back to 'UNKNOWN'."""
        self._log("info", "    _getPatientIdFromNode: reading DICOM tag 0010,0020...")
        try:
            sh_node = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(
                slicer.mrmlScene)
            item_id = sh_node.GetItemByDataNode(volume_node)
            uid     = sh_node.GetItemUID(item_id, "DICOM")
            self._log("info", f"    _getPatientIdFromNode: subject hierarchy UID={uid}")
            db    = slicer.dicomDatabase
            files = db.filesForSeries(uid)
            self._log("info",
                f"    _getPatientIdFromNode: "
                f"{len(files)} DICOM file(s) in DB for this series")
            if files:
                import pydicom
                ds         = pydicom.dcmread(files[0], stop_before_pixels=True)
                patient_id = str(ds[0x0010, 0x0020].value).strip()
                self._log("info",
                    f"    _getPatientIdFromNode: extracted patient_id={patient_id}")
                return patient_id
            else:
                self._log("warning",
                    "    _getPatientIdFromNode: no files in DICOM DB for series "
                    "— falling back to UNKNOWN")
        except Exception as e:
            self._log("warning",
                f"    _getPatientIdFromNode: exception reading DICOM tag: {e} "
                "— falling back to UNKNOWN")
        return "UNKNOWN"

    def _loadSegmentation(self, seg_file):
        """Load a NIfTI segmentation file into Slicer as a segmentation node."""
        self._log("info", f"    _loadSegmentation: loading {seg_file}")
        node = slicer.util.loadSegmentation(seg_file)
        if not node:
            raise ValueError(f"Failed to load segmentation: {seg_file}")
        n_segs = node.GetSegmentation().GetNumberOfSegments()
        self._log("info",
            f"    _loadSegmentation: OK — node={node.GetName()}  segments={n_segs}")
        return node

    def _registerCtToPet(self, ct_node, pet_node):
        """
        BRAINSFit rigid (6 DOF) registration: CT (moving) → PET (fixed).
        Metric: Mattes Mutual Information, sampling 2%.
        Returns vtkMRMLLinearTransformNode with the CT-to-PET transform.
        """
        self._log("info", "    _registerCtToPet: creating output transform node...")
        transform_node = slicer.mrmlScene.AddNewNodeByClass(
            'vtkMRMLLinearTransformNode', 'CT_to_PET_rigid'
        )

        params = {
            'fixedVolume':             pet_node.GetID(),
            'movingVolume':            ct_node.GetID(),
            'linearTransform':         transform_node.GetID(),
            'useRigid':                True,
            'costMetric':              'MMI',
            'samplingPercentage':      0.02,
            'initializeTransformMode': 'useMomentsAlign',
        }

        self._log("info", f"    _registerCtToPet: BRAINSFit params — {params}")
        self._log("info",
            "    _registerCtToPet: running BRAINSFit (wait_for_completion=True)...")

        cli_node = slicer.cli.run(
            slicer.modules.brainsfit, None, params, wait_for_completion=True
        )
        status = cli_node.GetStatusString()
        self._log("info", f"    _registerCtToPet: BRAINSFit finished — status={status}")
        slicer.mrmlScene.RemoveNode(cli_node)

        matrix = vtk.vtkMatrix4x4()
        transform_node.GetMatrixTransformToParent(matrix)
        mat_rows = []
        for r in range(4):
            row = [f"{matrix.GetElement(r, c):+.4f}" for c in range(4)]
            mat_rows.append("[" + "  ".join(row) + "]")
        self._log("info",
            "    _registerCtToPet: CT→PET transform matrix:\n"
            + "\n".join(f"      {r}" for r in mat_rows))

        return transform_node

    def _applyTransformToSeg(self, seg_node, transform_node):
        """Harden transform_node onto seg_node (bakes the transform in-place)."""
        self._log("info",
            f"    _applyTransformToSeg: setting transform {transform_node.GetID()} "
            f"on '{seg_node.GetName()}' and hardening...")
        seg_node.SetAndObserveTransformNodeID(transform_node.GetID())
        slicer.vtkSlicerTransformLogic().hardenTransform(seg_node)
        self._log("info", "    _applyTransformToSeg: transform hardened successfully")

    def _dilateMask(self, seg_node, margin_mm):
        """
        Dilate the first segment in seg_node by margin_mm mm using the Segment
        Editor Margin effect.  Logs voxel count before and after via
        slicer.modules.segmentations.logic().
        """
        import numpy as np

        if margin_mm <= 0:
            return

        segmentation = seg_node.GetSegmentation()
        segment_id   = segmentation.GetNthSegmentID(0)

        def _count_voxels():
            tmp = slicer.mrmlScene.AddNewNodeByClass(
                'vtkMRMLLabelMapVolumeNode', 'tmp_dilate_count'
            )
            slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(
                seg_node, tmp
            )
            count = int(np.count_nonzero(slicer.util.arrayFromVolume(tmp) > 0))
            slicer.mrmlScene.RemoveNode(tmp)
            return count

        before_count = _count_voxels()
        self._log("info",
            f"    _dilateMask: voxels before dilation = {before_count}")

        seg_editor_widget = slicer.qMRMLSegmentEditorWidget()
        seg_editor_widget.setMRMLScene(slicer.mrmlScene)
        seg_editor_node = slicer.mrmlScene.AddNewNodeByClass(
            'vtkMRMLSegmentEditorNode', 'SegEditor_dilate'
        )
        seg_editor_widget.setMRMLSegmentEditorNode(seg_editor_node)
        seg_editor_widget.setSegmentationNode(seg_node)
        seg_editor_widget.setCurrentSegmentID(segment_id)

        seg_editor_widget.setActiveEffectByName("Margin")
        effect = seg_editor_widget.activeEffect()
        if effect is None:
            slicer.mrmlScene.RemoveNode(seg_editor_node)
            raise RuntimeError(
                "Margin effect not available — ensure SegmentEditorMarginEffect "
                "extension is loaded"
            )

        effect.setParameter("MarginSizeMm", str(margin_mm))
        effect.self().onApply()

        seg_editor_widget.setActiveEffectByName(None)
        slicer.mrmlScene.RemoveNode(seg_editor_node)

        after_count = _count_voxels()
        self._log("info",
            f"    _dilateMask: dilation by {margin_mm} mm complete — "
            f"voxels: {before_count} → {after_count}  "
            f"(Δ = {after_count - before_count:+d})")

    def _findHottestVoxels(self, pet_node, label_node, top_n=10):
        """
        Find the top_n voxels with highest SUV inside the label mask.

        Returns list of dicts (length ≤ top_n), sorted descending by SUV:
            {"suv": float, "ras_x": float, "ras_y": float, "ras_z": float}

        arrayFromVolume returns shape (Z, Y, X) = (K, J, I).
        GetIJKToRASMatrix maps column vector [I, J, K, 1]^T → RAS.
        """
        import numpy as np

        pet_arr   = slicer.util.arrayFromVolume(pet_node)    # (Z, Y, X)
        label_arr = slicer.util.arrayFromVolume(label_node)  # (Z, Y, X)

        mask_indices = np.argwhere(label_arr > 0)  # shape (N, 3): [z, y, x]
        n_mask = len(mask_indices)
        self._log("info",
            f"    _findHottestVoxels: {n_mask} label voxel(s) found  top_n={top_n}")

        if n_mask == 0:
            self._log("warning",
                "    _findHottestVoxels: mask is empty — returning []")
            return []

        pet_vals = pet_arr[
            mask_indices[:, 0],
            mask_indices[:, 1],
            mask_indices[:, 2],
        ]

        order   = np.argsort(pet_vals)[::-1]
        top_idx = order[:top_n]

        ijk_to_ras = vtk.vtkMatrix4x4()
        pet_node.GetIJKToRASMatrix(ijk_to_ras)

        hotspots = []
        for idx in top_idx:
            z, y, x = mask_indices[idx]
            suv_val = float(pet_vals[idx])
            ras_h   = ijk_to_ras.MultiplyPoint([float(x), float(y), float(z), 1.0])
            hotspots.append({
                "suv":   suv_val,
                "ras_x": float(ras_h[0]),
                "ras_y": float(ras_h[1]),
                "ras_z": float(ras_h[2]),
            })

        self._log("info",
            f"    _findHottestVoxels: top-{len(hotspots)} hotspots computed  "
            f"max SUV={hotspots[0]['suv']:.4f}")
        return hotspots

    def _runPetIndic(self, pet_node, seg_node, metrics, suv_type):
        """
        Run QuantitativeIndicesTool CLI on pet_node + seg_node.
        Returns dict of metric results keyed by: suv_mean, suv_max, suv_peak,
        tlg, volume_mL (only keys where CLI returned a numeric value).
        """
        segmentation = seg_node.GetSegmentation()
        n_segs = segmentation.GetNumberOfSegments()
        self._log("info", f"    _runPetIndic: segmentation has {n_segs} segment(s)")

        if n_segs == 0:
            raise ValueError("Segmentation has no segments")

        segment_id = segmentation.GetNthSegmentID(0)
        self._log("info", f"    _runPetIndic: using segment_id={segment_id}")

        self._log("info", "    _runPetIndic: checking QuantitativeIndicesCLI module...")
        qi_module = slicer.modules.quantitativeindicescli
        self._log("info", f"    _runPetIndic: module OK — {qi_module}")

        # Export segment to temporary label map aligned to PET geometry
        self._log("info", "    _runPetIndic: exporting segment to label map...")
        label_node = slicer.mrmlScene.AddNewNodeByClass(
            'vtkMRMLLabelMapVolumeNode', 'temp_label'
        )
        segment_ids = vtk.vtkStringArray()
        segment_ids.InsertNextValue(segment_id)
        slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
            seg_node, segment_ids, label_node, pet_node,
            slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
        )
        self._log("info",
            f"    _runPetIndic: label map created — id={label_node.GetID()}")

        try:
            parameters = {
                'Grayscale_Image': pet_node.GetID(),
                'Label_Image':     label_node.GetID(),
                'Label_Value':     '1',
            }
            if metrics.get("mean"):   parameters['Mean']   = 'true'
            if metrics.get("max"):    parameters['Max']    = 'true'
            if metrics.get("peak"):   parameters['Peak']   = 'true'
            if metrics.get("tlg"):    parameters['TLG']    = 'true'
            if metrics.get("volume"): parameters['Volume'] = 'true'

            self._log("info",
                f"    _runPetIndic: CLI parameters — {parameters}")
            self._log("info",
                "    _runPetIndic: running CLI (wait_for_completion=True)...")

            cli_node = slicer.cli.run(
                qi_module, None, parameters, wait_for_completion=True
            )

            cli_status = cli_node.GetStatusString()
            self._log("info", f"    _runPetIndic: CLI finished — status={cli_status}")

            # Parse results
            name_map = {
                'Mean':   'suv_mean',
                'Max':    'suv_max',
                'Peak':   'suv_peak',
                'TLG':    'tlg',
                'Volume': 'volume_mL',
            }
            results  = {}
            import math
            n_params = cli_node.GetNumberOfParametersInGroup(3)
            self._log("info",
                f"    _runPetIndic: parsing {n_params} output parameter(s) from group 3...")

            for i in range(n_params):
                raw  = cli_node.GetParameterDefault(3, i)
                name = cli_node.GetParameterName(3, i)
                self._log("info",
                    f"    _runPetIndic:   raw param [{name}] = {raw!r}")
                if raw == '--':
                    continue
                name_clean = name.replace('_s', '').replace('_', ' ').strip()
                if name_clean in name_map:
                    try:
                        val = float(raw)
                        if not math.isnan(val):
                            results[name_map[name_clean]] = val
                    except ValueError:
                        self._log("warning",
                            f"    _runPetIndic: could not convert "
                            f"{raw!r} to float for param '{name}'")

            slicer.mrmlScene.RemoveNode(cli_node)
            self._log("info", f"    _runPetIndic: results parsed — {results}")

            if not results:
                self._log("warning",
                    "    _runPetIndic: results dict is EMPTY — "
                    "CLI may have run but returned no numeric values. "
                    "Check that PET units are correct and label map is non-empty.")

            if "volume_mL" in results:
                self._log("info",
                    f"    _runPetIndic: volume={results['volume_mL']:.3f} mL  "
                    f"(measured in PET voxel space — ~1-2% smaller than CT-resolution "
                    f"Segment Stats due to resampling, this is expected)")

            self._log("info", "    _runPetIndic: finding hottest voxels...")
            try:
                hotspots = self._findHottestVoxels(pet_node, label_node, top_n=10)
                if hotspots:
                    h = hotspots[0]
                    results["hotspot_suv"]   = h["suv"]
                    results["hotspot_ras_x"] = h["ras_x"]
                    results["hotspot_ras_y"] = h["ras_y"]
                    results["hotspot_ras_z"] = h["ras_z"]
                    self._log("info",
                        f"    _runPetIndic: top-1 hotspot — "
                        f"SUV={h['suv']:.4f}  "
                        f"RAS=({h['ras_x']:.1f}, {h['ras_y']:.1f}, {h['ras_z']:.1f})")
                else:
                    self._log("warning",
                        "    _runPetIndic: no label voxels found for hotspot analysis")
            except Exception as e:
                self._log("warning",
                    f"    _runPetIndic: hotspot finding failed: {e}")

            return results

        finally:
            slicer.mrmlScene.RemoveNode(label_node)
            self._log("info", "    _runPetIndic: temp label map removed")

    def _applySuvConversion(self, pet_node, dicom_folder):
        """
        Convert the PET volume from Bq/mL → SUVbw in-place.
        SUVbw = (Bq/mL × weight_g) / decay_corrected_dose_Bq
        If DICOM Units tag is already 'SUV', conversion is skipped.
        """
        import pydicom
        import numpy as np

        # Find first readable DICOM file
        first_dcm = None
        for root, _, files in os.walk(dicom_folder):
            for f in sorted(files):
                fp = os.path.join(root, f)
                try:
                    pydicom.dcmread(fp, stop_before_pixels=True)
                    first_dcm = fp
                    break
                except Exception:
                    continue
            if first_dcm:
                break

        if not first_dcm:
            raise ValueError(f"No readable DICOM file found in {dicom_folder}")

        ds    = pydicom.dcmread(first_dcm, stop_before_pixels=True)
        units = str(getattr(ds, "Units", "BQML")).strip().upper()
        self._log("info", f"    _applySuvConversion: DICOM Units tag = '{units}'")

        if units == "SUV":
            self._log("info",
                "    _applySuvConversion: already SUV — skipping conversion")
            return

        if units not in ("BQML", "BQ/ML", ""):
            self._log("warning",
                f"    _applySuvConversion: unexpected units '{units}' — "
                f"attempting conversion anyway (assuming Bq/mL)")

        # Patient weight
        try:
            weight_kg = float(str(getattr(ds, "PatientWeight", None)))
            if __import__("math").isnan(weight_kg):
                raise ValueError("PatientWeight is NaN")
        except (TypeError, ValueError) as e:
            raise ValueError(
                f"PatientWeight missing or unreadable in DICOM header: {e}")
        weight_g = weight_kg * 1000
        self._log("info",
            f"    _applySuvConversion: PatientWeight = {weight_kg} kg")

        # Injected dose
        try:
            radio_seq = ds[0x0054, 0x0016][0]
        except (KeyError, IndexError):
            raise ValueError(
                "RadiopharmaceuticalInformationSequence (0054,0016) missing — "
                "cannot compute SUV")

        try:
            dose_bq = float(str(radio_seq[0x0018, 0x1074].value))
        except (KeyError, ValueError):
            raise ValueError(
                "RadionuclideTotalDose (0018,1074) missing from "
                "RadiopharmaceuticalInformationSequence")

        # Injection time
        inj_dt = getattr(radio_seq, "RadiopharmaceuticalStartDateTime", None)
        inj_t  = getattr(radio_seq, "RadiopharmaceuticalStartTime", None)
        if inj_dt:
            inj_str = str(inj_dt).split(".")[0]
        elif inj_t:
            inj_str = str(inj_t).split(".")[0]
        else:
            raise ValueError(
                "No injection time found in RadiopharmaceuticalInformationSequence "
                "(tried RadiopharmaceuticalStartDateTime and RadiopharmaceuticalStartTime)")

        # Radionuclide half-life (standard tag, then GE private fallback, then F-18 default)
        try:
            half_life_s = float(str(radio_seq[0x0018, 0x1075].value))
            self._log("info",
                f"    _applySuvConversion: half-life from standard tag = {half_life_s}s")
        except (KeyError, ValueError):
            try:
                half_life_s = float(str(ds[0x0009, 0x103F].value))
                self._log("info",
                    f"    _applySuvConversion: half-life from GE private tag = {half_life_s}s")
            except (KeyError, ValueError):
                half_life_s = 6586.2  # F-18 default
                self._log("warning",
                    f"    _applySuvConversion: half-life not in DICOM — "
                    f"using F-18 default = {half_life_s}s")

        # Acquisition time and decay correction type
        acq_t_raw  = str(getattr(ds, "AcquisitionTime", "") or
                         getattr(ds, "SeriesTime", "")).split(".")[0].strip()
        decay_corr = str(getattr(ds, "DecayCorrection", "START")).strip().upper()

        self._log("info",
            f"    _applySuvConversion: dose={dose_bq:.0f} Bq  "
            f"inj={inj_str}  acq={acq_t_raw}  "
            f"half_life={half_life_s}s  decay_correction={decay_corr}")

        def _hhmmss_to_s(t):
            t = t.strip()[-6:]
            try:
                return int(t[0:2]) * 3600 + int(t[2:4]) * 60 + int(t[4:6])
            except (ValueError, IndexError):
                return 0

        if decay_corr == "ADMIN":
            decay_corrected_dose_bq = dose_bq
            self._log("info",
                "    _applySuvConversion: ADMIN correction — "
                "using injected dose directly (no further decay calc)")
        else:
            inj_s = _hhmmss_to_s(inj_str)
            acq_s = _hhmmss_to_s(acq_t_raw)
            if acq_s < inj_s:
                acq_s += 86400          # midnight crossing
            decay_time_s = acq_s - inj_s
            decay_corrected_dose_bq = dose_bq * (2.0 ** (-decay_time_s / half_life_s))
            self._log("info",
                f"    _applySuvConversion: Δt={decay_time_s}s  "
                f"decay_corrected_dose={decay_corrected_dose_bq:.0f} Bq  "
                f"(={decay_corrected_dose_bq/1e6:.3f} MBq)")

        if decay_corrected_dose_bq <= 0:
            raise ValueError(
                f"decay_corrected_dose={decay_corrected_dose_bq} — "
                "check injection time and dose values in DICOM header")

        suv_factor = weight_g / decay_corrected_dose_bq
        self._log("info",
            f"    _applySuvConversion: SUV factor = {suv_factor:.6f}  "
            f"(weight={weight_g}g / decay_dose={decay_corrected_dose_bq:.0f} Bq)")

        arr = slicer.util.arrayFromVolume(pet_node)
        max_before = float(arr.max())
        arr_suv    = (arr * suv_factor).astype(np.float32)
        slicer.util.updateVolumeFromArray(pet_node, arr_suv)
        max_after  = float(arr_suv.max())

        self._log("info",
            f"    _applySuvConversion: DONE — "
            f"max: {max_before:.1f} Bq/mL → {max_after:.4f} SUV")

    # ── Excel ─────────────────────────────────────────────────────────────

    def _saveExcel(self, rows, output_file, append):
        """
        Data sheet  — one row per scan × segment.
          Columns: subject_id, patient_id, scan_date, segment, source_file,
                   volume_mL, suv_mean, suv_max, suv_peak, tlg, status, computed_at

        Summary sheet — pivoted: one row per scan, columns = {display_label}_{metric}.
          e.g. visceral_fat_volume_mL, spleen_suv_mean …
          Rebuilt from scratch on every save so it always reflects the full Data sheet.
        """
        DATA_COLS   = ["subject_id", "patient_id", "scan_date",
                       "segment", "source_file",
                       "volume_mL", "suv_mean", "suv_max", "suv_peak", "tlg",
                       "hotspot_suv", "hotspot_ras_x", "hotspot_ras_y", "hotspot_ras_z",
                       "status", "computed_at"]
        METRIC_COLS = ["volume_mL", "suv_mean", "suv_max", "suv_peak", "tlg",
                       "hotspot_suv", "hotspot_ras_x", "hotspot_ras_y", "hotspot_ras_z"]

        self._log("info",
            f"_saveExcel: {len(rows)} row(s) to write  append={append}  "
            f"file={output_file}")

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        for row in rows:
            row["computed_at"] = now_str

        # Load or create workbook
        if append and os.path.isfile(output_file):
            self._log("info", "_saveExcel: opening existing workbook (append mode)")
            wb = openpyxl.load_workbook(output_file)
            for stale in ("Sheet", "Sheet1", "Results"):
                if stale in wb.sheetnames:
                    del wb[stale]
                    self._log("info", f"_saveExcel: removed stale sheet '{stale}'")
        else:
            self._log("info", "_saveExcel: creating new workbook")
            wb = openpyxl.Workbook()

        # ── Data sheet ────────────────────────────────────────────────
        if "Data" in wb.sheetnames:
            ws = wb["Data"]
        else:
            if wb.active and wb.active.title in ("Sheet", "Results"):
                ws = wb.active
                ws.title = "Data"
            else:
                ws = wb.create_sheet("Data", 0)
            for ci, col in enumerate(DATA_COLS, 1):
                ws.cell(row=1, column=ci, value=col)

        for row in rows:
            ws.append([row.get(c, "") for c in DATA_COLS])

        self._log("info",
            f"_saveExcel: Data sheet — {len(rows)} new row(s) appended  "
            f"total rows now={ws.max_row - 1}")

        # ── Summary sheet — rebuild from full Data sheet ───────────────
        header = [ws.cell(row=1, column=c).value
                  for c in range(1, ws.max_column + 1)]
        all_data = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value
                    for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in vals):
                all_data.append(dict(zip(header, vals)))

        seen_segs = list(dict.fromkeys(
            r.get("segment", "") for r in all_data if r.get("status") == "done"
        ))

        pivot = {}
        for r in all_data:
            if r.get("status") != "done":
                continue
            key = (r.get("subject_id", ""), r.get("patient_id", ""),
                   r.get("scan_date", ""))
            pivot.setdefault(key, {})
            seg = r.get("segment", "")
            for m in METRIC_COLS:
                col_name = f"{seg}_{m}"
                val = r.get(m)
                if val not in (None, ""):
                    pivot[key][col_name] = val

        id_cols      = ["subject_id", "patient_id", "scan_date"]
        metric_order = [f"{seg}_{m}" for seg in seen_segs for m in METRIC_COLS]
        sum_cols     = id_cols + metric_order

        self._log("info",
            f"_saveExcel: rebuilding Summary sheet from {len(all_data)} total Data row(s)  "
            f"pivot keys={len(pivot)}  segment combos={len(seen_segs)}")

        if "Summary" in wb.sheetnames:
            del wb["Summary"]
            self._log("info", "_saveExcel: old Summary sheet removed")
        ws_s = wb.create_sheet("Summary")

        for ci, col in enumerate(sum_cols, 1):
            ws_s.cell(row=1, column=ci, value=col)

        for ri, (subj, pid, date) in enumerate(sorted(pivot), start=2):
            ws_s.cell(ri, 1, subj)
            ws_s.cell(ri, 2, pid)
            ws_s.cell(ri, 3, date)
            data = pivot[(subj, pid, date)]
            for ci, col in enumerate(metric_order, start=4):
                ws_s.cell(ri, ci, data.get(col, ""))

        wb.save(output_file)
        self._log("info",
            f"_saveExcel: workbook saved — sheets={wb.sheetnames}  path={output_file}")

    def _errorRow(self, subject_id, scan_date, segment, status, patient_id=""):
        return {
            "subject_id": subject_id,
            "patient_id": patient_id,
            "scan_date":  scan_date,
            "segment":    segment,
            "status":     status,
        }


# ── TEST ───────────────────────────────────────────────────────────────────

class PETCTQuantAnalysis_v2Test(ScriptedLoadableModuleTest):
    def setUp(self):
        slicer.mrmlScene.Clear(0)

    def runTest(self):
        logic = PETCTQuantAnalysis_v2Logic()
        result = logic._parseFolderName("MSP0001_2025-07-09_PET")
        assert result == ("MSP0001", "2025-07-09", "PET"), f"parse failed: {result}"
        assert logic._parseFolderName("bad_folder") is None
        self.delayDisplay("v2 parse test passed")
