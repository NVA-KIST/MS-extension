# Modified metabolic-radiomics version of PETBiomarkerStudio.
import os
import re
import math
import tempfile
import subprocess
import numpy as np
from scipy import ndimage

import qt
import ctk
import vtk
import slicer
from slicer.i18n import tr as _
from slicer.i18n import translate
from slicer.ScriptedLoadableModule import *

# extension_new/lib on sys.path (slicer_modules/PETBiomarkerStudio -> ../..)
import sys as _sys
_EXT_NEW_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if _EXT_NEW_ROOT not in _sys.path:
    _sys.path.insert(0, _EXT_NEW_ROOT)

from lib.processing.ablation import (
    apply_urinary_cleanup as _lib_apply_urinary_cleanup,
    build_pet_urinary_mask as _lib_build_pet_urinary_mask,
    clip_binary_mask_by_ras_z as _lib_clip_binary_mask_by_ras_z,
    exclude_dilated_structure as _lib_exclude_dilated_structure,
    flag_qc_outliers as _lib_flag_qc_outliers,
    physical_dilation_structure as _lib_physical_dilation_structure,
    qc_suv_stats as _lib_qc_suv_stats,
)
from lib.quantification import radiomics as _lib_radiomics
from lib.quantification.biomarker_batch import (
    AUXILIARY_STEM_TOKENS as _LIB_AUXILIARY_STEM_TOKENS,
    batch_error_row as _lib_batch_error_row,
    computation_signature as _lib_computation_signature,
    cross_roi_derived_by_subject as _lib_cross_roi_derived_by_subject,
    default_excel_label as _lib_default_excel_label,
    existing_batch_keys as _lib_existing_batch_keys,
    find_batch_segment_file as _lib_find_batch_segment_file,
    is_auxiliary_segment_stem as _lib_is_auxiliary_segment_stem,
    parse_batch_base_name as _lib_parse_batch_base_name,
    parse_quantitative_indices_results as _lib_parse_qi_results,
    safe_asymmetry as _lib_safe_asymmetry,
    save_batch_rows_to_excel as _lib_save_batch_rows_to_excel,
    save_qc_rows_to_excel as _lib_save_qc_rows_to_excel,
    scan_batch_dataset as _lib_scan_batch_dataset,
    segment_role as _lib_segment_role,
)


# ==================================================================================
# PETBiomarkerStudio
#
# A single-subject PET biomarker workbench for Metabolic Syndrome research.
# It unifies four previously separate tools:
#   1. PETAblationStudyTool   -> ROI ablation (spine clip / kidney / urinary) + metrics
#   2. UreterPostProcess      -> robust PET-derived urinary (ureter) exclusion mask
#   3. PETCTQuantAnalysis_v3  -> SUV metrics + radiomics + Data/Summary Excel export
#   4. HotspotFromSegmentation-> SUVmax hotspot localisation + scene fiducial
#
# Workflow narrative:
#   Extract SUV from low-uptake target organs (Visceral Fat, Psoas, Spleen) ->
#   contamination from neighbouring hot structures (bladder/ureter/kidney) shows up
#   as an anomalously high SUVmax -> the QC/Hotspot panel flags those outlier
#   segments, shows WHERE the hot voxel sits in the scene, and lets you re-run after
#   ablation to confirm the spike was removed.
# ==================================================================================
class PETBiomarkerStudio(ScriptedLoadableModule):
    """Uses ScriptedLoadableModule base class, available at:
    https://github.com/Slicer/Slicer/blob/main/Base/Python/slicer/ScriptedLoadableModule.py
    """

    def __init__(self, parent):
        ScriptedLoadableModule.__init__(self, parent)
        self.parent.title = _("PET Biomarker Studio")
        self.parent.categories = ["Metabolic Syndrome Toolkit"] # [translate("qSlicerAbstractCoreModule", "Quantification")]
        self.parent.dependencies = ["QuantitativeIndicesCLI"]
        self.parent.contributors = ["Sumin Cho (KIST)", "Ishita Singh Faujdar (KIST)"]
        self.parent.helpText = _("""
            Single-subject PET biomarker workbench for metabolic-syndrome research.
            Combines ROI ablation (spine clip / kidney exclusion / urinary cleanup),
            SUV metric + radiomics extraction, an Excel exporter, and a QC / hotspot
            panel that flags outlier segments and localises the contaminating voxel
            in the scene.

            Pure algorithms live in extension_new/lib
            (processing/ablation.py, quantification/radiomics.py,
            quantification/biomarker_batch.py); this module is the Slicer UI + adapter.
            """)
        self.parent.acknowledgementText = _("KIST NVA")

# --------------------------------------------------
# Widget (UI)
# --------------------------------------------------
class PETBiomarkerStudioWidget(ScriptedLoadableModuleWidget):
    def setup(self):
        ScriptedLoadableModuleWidget.setup(self)

        self.logic = PETBiomarkerStudioLogic()

        # --------------------------------------------------
        # Top-level tabs: Interactive (single subject) | Batch (cohort)
        # Both tabs share the SAME PETBiomarkerStudioLogic instance, so the SUV /
        # ablation / radiomics maths cannot diverge between single and bulk runs.
        # --------------------------------------------------
        self.mainTabs = qt.QTabWidget()
        self.layout.addWidget(self.mainTabs)

        interactiveTab = qt.QWidget()
        self.interactiveLayout = qt.QVBoxLayout(interactiveTab)
        self.mainTabs.addTab(interactiveTab, "Interactive (single subject)")

        batchTab = qt.QWidget()
        self.batchLayout = qt.QVBoxLayout(batchTab)
        self.mainTabs.addTab(batchTab, "Batch (cohort quantification)")

        visualizationTab = qt.QWidget()
        self.visualizationLayout = qt.QVBoxLayout(visualizationTab)
        self.visualizationTabIndex = self.mainTabs.addTab(
            visualizationTab, "Visualization"
        )

        # --------------------------------------------------
        # UI Components (Interactive tab)
        # --------------------------------------------------

        self.statuslabel = qt.QLabel("PET Biomarker Studio — interactive workbench.")
        self.statuslabel.setWordWrap(True)
        self.statuslabel.setTextInteractionFlags(qt.Qt.TextSelectableByMouse)
        self.interactiveLayout.addWidget(self.statuslabel)

        
        # -------------------- 1. Inputs Collapsible -------------------- 
        inputCollapsible = ctk.ctkCollapsibleButton()
        inputCollapsible.text = "1. Inputs"
        self.interactiveLayout.addWidget(inputCollapsible)
        inputsLayout = qt.QFormLayout(inputCollapsible)

    
        # PET Selector
        petSelectorBox = qt.QHBoxLayout()

        self.petSelector = slicer.qMRMLNodeComboBox()
        self.petSelector.nodeTypes = ["vtkMRMLScalarVolumeNode"]
        self.petSelector.addEnabled = False
        self.petSelector.removeEnabled = False
        self.petSelector.noneEnabled = True
        self.petSelector.showHidden = False
        self.petSelector.setMRMLScene(slicer.mrmlScene)
        self.petSelector.setToolTip("Select a PET volume from the current Slicer scene.")

        self.loadPETButton = qt.QPushButton("Load File")
        self.loadPETButton.setToolTip("Load PET volume from a NIfTI file (.nii/.nii.gz).")

        self.loadPETDicomButton = qt.QPushButton("Load DICOM")
        self.loadPETDicomButton.setToolTip(
            "Import a PET DICOM series folder (one folder of .dcm slices). "
            "An SUVbw-converted volume is preferred when available."
        )

        petSelectorBox.addWidget(self.petSelector)
        petSelectorBox.addWidget(self.loadPETButton)
        petSelectorBox.addWidget(self.loadPETDicomButton)

        inputsLayout.addRow("PET:", petSelectorBox)

        # CT Selector
        ctSelectorBox = qt.QHBoxLayout()

        self.ctSelector = slicer.qMRMLNodeComboBox()
        self.ctSelector.nodeTypes = ["vtkMRMLScalarVolumeNode"]
        self.ctSelector.addEnabled = False
        self.ctSelector.removeEnabled = False
        self.ctSelector.noneEnabled = True
        self.ctSelector.showHidden = False
        self.ctSelector.setMRMLScene(slicer.mrmlScene)
        self.ctSelector.setToolTip("Select a CT volume from the current Slicer scene.")

        self.loadCTButton = qt.QPushButton("Load File")
        self.loadCTButton.setToolTip("Load CT volume from a NIfTI file (.nii/.nii.gz).")

        self.loadCTDicomButton = qt.QPushButton("Load DICOM")
        self.loadCTDicomButton.setToolTip(
            "Import a CT DICOM series folder (one folder of .dcm slices)."
        )

        ctSelectorBox.addWidget(self.ctSelector)
        ctSelectorBox.addWidget(self.loadCTButton)
        ctSelectorBox.addWidget(self.loadCTDicomButton)

        inputsLayout.addRow("CT:", ctSelectorBox)

        # Segmentation(ROI) Selector
        segmentationSelectorBox = qt.QHBoxLayout()

        self.segmentationSelector = slicer.qMRMLNodeComboBox()
        self.segmentationSelector.nodeTypes = ["vtkMRMLSegmentationNode"]
        self.segmentationSelector.addEnabled = False
        self.segmentationSelector.removeEnabled = False
        self.segmentationSelector.noneEnabled = True
        self.segmentationSelector.showHidden = False
        self.segmentationSelector.setMRMLScene(slicer.mrmlScene)
        self.segmentationSelector.setToolTip("Select a segmentation(ROI) from the current Slicer scene.")

        self.loadSegmentationButton = qt.QPushButton("Load File")
        self.loadSegmentationButton.setToolTip("Load segmentation from the file.")

        self.generateSegmentationButton = qt.QPushButton("Generate")
        self.generateSegmentationButton.setToolTip("Generate segmentation from CT using TotalSegmentator.")

        segmentationSelectorBox.addWidget(self.segmentationSelector)
        segmentationSelectorBox.addWidget(self.loadSegmentationButton)
        segmentationSelectorBox.addWidget(self.generateSegmentationButton)

        inputsLayout.addRow("Segmentation(ROI):", segmentationSelectorBox)

        # -------------------- 2. Preprocessing Options Collapsible -------------------- 
        preprocessingCollapsible = ctk.ctkCollapsibleButton()
        preprocessingCollapsible.text = "2. ROI Preprocessing Options"
        self.interactiveLayout.addWidget(preprocessingCollapsible)
        preprocessingLayout = qt.QFormLayout(preprocessingCollapsible)

        # ----- Spine Slicing Options -----
        self.useSpineSlicingCheckBox = qt.QCheckBox("Apply Spine Level Slicing")
        self.useSpineSlicingCheckBox.setChecked(False)
        preprocessingLayout.addRow(self.useSpineSlicingCheckBox)

        vertebraeInputWidget = qt.QWidget()
        vertebraeInputLayout = qt.QHBoxLayout(vertebraeInputWidget)
        vertebraeInputLayout.setContentsMargins(0, 0, 0, 0)

        self.vertebraeSelector = slicer.qMRMLNodeComboBox()
        self.vertebraeSelector.nodeTypes = ["vtkMRMLSegmentationNode"]
        self.vertebraeSelector.addEnabled = False
        self.vertebraeSelector.removeEnabled = False
        self.vertebraeSelector.noneEnabled = True
        self.vertebraeSelector.showHidden = False
        self.vertebraeSelector.setMRMLScene(slicer.mrmlScene)
        self.vertebraeSelector.setToolTip("Select a segmentation node containing L1-L5 vertebrae.")

        self.loadVertebraeButton = qt.QPushButton("Load File")
        self.loadVertebraeButton.setToolTip("Load vertebrae segmentation from the file.")

        self.generateVertebraeButton = qt.QPushButton("Generate")
        self.generateVertebraeButton.setToolTip("Generate L1-L5 vertebrae masks from CT using TotalSegmentator.")

        vertebraeInputLayout.addWidget(self.vertebraeSelector)
        vertebraeInputLayout.addWidget(self.loadVertebraeButton)
        vertebraeInputLayout.addWidget(self.generateVertebraeButton)

        preprocessingLayout.addRow("Vertebrae:", vertebraeInputWidget)

        spineRangeWidget = qt.QWidget()
        spineRangeLayout = qt.QHBoxLayout(spineRangeWidget)
        spineRangeLayout.setContentsMargins(0, 0, 0, 0)

        self.spineStartSelector =  qt.QComboBox()
        self.spineEndSelector = qt.QComboBox()

        for level in ["L1", "L2", "L3", "L4", "L5"]:
            self.spineStartSelector.addItem(level)
            self.spineEndSelector.addItem(level)
        
        self.spineStartSelector.setCurrentIndex(0)
        self.spineEndSelector.setCurrentIndex(4)

        spineRangeLayout.addWidget(qt.QLabel("Start:"))
        spineRangeLayout.addWidget(self.spineStartSelector)
        spineRangeLayout.addWidget(qt.QLabel("End:"))
        spineRangeLayout.addWidget(self.spineEndSelector)

        preprocessingLayout.addRow("Spine Range: ", spineRangeWidget)

        self.computeSpineBoundsButton = qt.QPushButton("Compute Spine Bounds")
        self.computeSpineBoundsButton.setToolTip("Compute RAS(Right-Anterior-Superior) Z-axis bounds from selected vertebrae levels.")

        preprocessingLayout.addRow("", self.computeSpineBoundsButton)

        self.applySpineSlicingButton = qt.QPushButton("Apply Spine Slicing")
        self.applySpineSlicingButton.setToolTip(
            "Create a processed ROI segmentation clipped to the selected spine range."
        )

        preprocessingLayout.addRow("", self.applySpineSlicingButton)

        # ----- Kidney Exclusion Options -----
        self.useKidneyExclusionCheckBox = qt.QCheckBox("Apply Kidney Exclusion")
        self.useKidneyExclusionCheckBox.setChecked(False)
        preprocessingLayout.addRow(self.useKidneyExclusionCheckBox)

        kidneyInputWidget = qt.QWidget()
        kidneyInputLayout = qt.QHBoxLayout(kidneyInputWidget)
        kidneyInputLayout.setContentsMargins(0, 0, 0, 0)

        self.kidneySelector = slicer.qMRMLNodeComboBox()
        self.kidneySelector.nodeTypes = ["vtkMRMLSegmentationNode"]
        self.kidneySelector.addEnabled = False
        self.kidneySelector.removeEnabled = False
        self.kidneySelector.noneEnabled = True
        self.kidneySelector.showHidden = False
        self.kidneySelector.setMRMLScene(slicer.mrmlScene)
        self.kidneySelector.setToolTip(
            "Select a segmentation node containing kidney_left and kidney_right."
        )

        self.loadKidneyButton = qt.QPushButton("Load File")
        self.loadKidneyButton.setToolTip("Load kidney segmentation from file.")

        self.generateKidneyButton = qt.QPushButton("Generate")
        self.generateKidneyButton.setToolTip(
            "Generate kidney_left and kidney_right masks from CT using TotalSegmentator."
        )

        kidneyInputLayout.addWidget(self.kidneySelector)
        kidneyInputLayout.addWidget(self.loadKidneyButton)
        kidneyInputLayout.addWidget(self.generateKidneyButton)

        preprocessingLayout.addRow("Kidneys:", kidneyInputWidget)

        self.kidneyDilationSpinBox = qt.QDoubleSpinBox()
        self.kidneyDilationSpinBox.setMinimum(0.0)
        self.kidneyDilationSpinBox.setMaximum(50.0)
        self.kidneyDilationSpinBox.setSingleStep(1.0)
        self.kidneyDilationSpinBox.setValue(10.0)
        self.kidneyDilationSpinBox.setSuffix(" mm")
        self.kidneyDilationSpinBox.setToolTip(
            "Dilation radius used before subtracting kidney mask from the ROI."
        )

        preprocessingLayout.addRow("Dilation Radius:", self.kidneyDilationSpinBox)

        self.applyKidneyExclusionButton = qt.QPushButton("Apply Kidney Exclusion")
        self.applyKidneyExclusionButton.setToolTip(
            "Create a processed ROI segmentation by subtracting dilated kidney masks."
        )
        preprocessingLayout.addRow("", self.applyKidneyExclusionButton)

        # ----- Urinary Activity Cleanup Options -----
        self.useUreterCleanupCheckBox = qt.QCheckBox("Apply PET-derived urinary activity cleanup")
        self.useUreterCleanupCheckBox.setChecked(False)
        preprocessingLayout.addRow(self.useUreterCleanupCheckBox)

        ureterInputWidget = qt.QWidget()
        ureterInputLayout = qt.QHBoxLayout(ureterInputWidget)
        ureterInputLayout.setContentsMargins(0, 0, 0, 0)

        self.ureterMaskSelector = slicer.qMRMLNodeComboBox()
        self.ureterMaskSelector.nodeTypes = ["vtkMRMLSegmentationNode"]
        self.ureterMaskSelector.addEnabled = False
        self.ureterMaskSelector.removeEnabled = False
        self.ureterMaskSelector.noneEnabled = True
        self.ureterMaskSelector.showHidden = False
        self.ureterMaskSelector.setMRMLScene(slicer.mrmlScene)
        self.ureterMaskSelector.setToolTip(
            "Select or generate a PET-derived urinary activity exclusion mask."
        )

        self.generateUreterMaskButton = qt.QPushButton("Generate")
        self.generateUreterMaskButton.setToolTip(
            "Generate PET-derived urinary activity exclusion mask from PET/SUV volume."
        )

        ureterInputLayout.addWidget(self.ureterMaskSelector)
        ureterInputLayout.addWidget(self.generateUreterMaskButton)

        preprocessingLayout.addRow("Urinary activity mask:", ureterInputWidget)

        self.ureterSUVThresholdSpinBox = qt.QDoubleSpinBox()
        self.ureterSUVThresholdSpinBox.setMinimum(0.0)
        self.ureterSUVThresholdSpinBox.setMaximum(100.0)
        self.ureterSUVThresholdSpinBox.setSingleStep(0.5)
        self.ureterSUVThresholdSpinBox.setValue(4.0)
        self.ureterSUVThresholdSpinBox.setToolTip(
            "SUV threshold used to detect urinary hot regions."
        )

        preprocessingLayout.addRow("Urinary SUV threshold:", self.ureterSUVThresholdSpinBox)

        self.ureterDilationSpinBox = qt.QDoubleSpinBox()
        self.ureterDilationSpinBox.setMinimum(0.0)
        self.ureterDilationSpinBox.setMaximum(50.0)
        self.ureterDilationSpinBox.setSingleStep(1.0)
        self.ureterDilationSpinBox.setValue(5.0)
        self.ureterDilationSpinBox.setSuffix(" mm")
        self.ureterDilationSpinBox.setToolTip(
            "Dilation radius applied to the detected urinary activity mask."
        )

        preprocessingLayout.addRow("Urinary mask dilation:", self.ureterDilationSpinBox)

        self.ureterCleanSUVThresholdSpinBox = qt.QDoubleSpinBox()
        self.ureterCleanSUVThresholdSpinBox.setMinimum(0.0)
        self.ureterCleanSUVThresholdSpinBox.setMaximum(100.0)
        self.ureterCleanSUVThresholdSpinBox.setSingleStep(0.5)
        self.ureterCleanSUVThresholdSpinBox.setValue(2.0)
        self.ureterCleanSUVThresholdSpinBox.setToolTip(
            "Only ROI voxels overlapping the urinary mask and exceeding this SUV threshold will be removed."
        )
        preprocessingLayout.addRow("ROI cleanup SUV threshold:", self.ureterCleanSUVThresholdSpinBox)

        self.applyUreterCleanupButton = qt.QPushButton("Apply Urinary Activity Cleanup")
        self.applyUreterCleanupButton.setToolTip(
            "Create a processed ROI by removing PET-hot voxels overlapping the urinary activity mask."
        )
        preprocessingLayout.addRow("", self.applyUreterCleanupButton)


        

        # -------------------- 3. Metrics Calculation Options -------------------- 
        metricsCalculationCollapsible = ctk.ctkCollapsibleButton()
        metricsCalculationCollapsible.text = "3. Metrics Calculation Options"
        self.interactiveLayout.addWidget(metricsCalculationCollapsible)
        metricsLayout = qt.QFormLayout(metricsCalculationCollapsible)

        metricsWidget = qt.QWidget()
        metricsBox = qt.QHBoxLayout(metricsWidget)
        metricsBox.setContentsMargins(0, 0, 0, 0)

        self.suvMeanCheckBox = qt.QCheckBox("SUVmean")
        self.suvMaxCheckBox = qt.QCheckBox("SUVmax")
        self.suvPeakCheckBox = qt.QCheckBox("SUVpeak")
        self.tlgCheckBox = qt.QCheckBox("TLG")
        self.volumeCheckBox = qt.QCheckBox("Volume")

        for cb in (self.suvMeanCheckBox, self.suvMaxCheckBox, self.suvPeakCheckBox, self.tlgCheckBox, self.volumeCheckBox):
            metricsBox.addWidget(cb)
        
        metricsLayout.addRow("Metrics: ", metricsWidget)

        # The current study starts from eight pre-selected, interpretable
        # radiomic features. Each feature can be enabled independently so the
        # extracted columns exactly match the user's current analysis plan.
        selectedRadiomicsWidget = qt.QWidget()
        selectedRadiomicsLayout = qt.QGridLayout(selectedRadiomicsWidget)
        selectedRadiomicsLayout.setContentsMargins(0, 0, 0, 0)

        self.radP10CheckBox = qt.QCheckBox("P10")
        self.radP90CheckBox = qt.QCheckBox("P90")
        self.radEntropyCheckBox = qt.QCheckBox("Entropy")
        self.radSkewnessCheckBox = qt.QCheckBox("Skewness")
        self.radContrastCheckBox = qt.QCheckBox("GLCM Contrast")
        self.radSahgleCheckBox = qt.QCheckBox("SAHGLE")
        self.radLalgleCheckBox = qt.QCheckBox("LALGLE")
        self.radZoneEntropyCheckBox = qt.QCheckBox("ZoneEntropy")

        selectedRadiomics = (
            (
                self.radP10CheckBox,
                "10th Percentile: representative low-end Standardized Uptake Value (SUV) within the Region of Interest (ROI).",
            ),
            (
                self.radP90CheckBox,
                "90th Percentile: representative high-end Standardized Uptake Value (SUV), less sensitive than SUVmax to one extreme voxel.",
            ),
            (
                self.radEntropyCheckBox,
                "First-Order Entropy: uncertainty and diversity of Standardized Uptake Value (SUV) intensities within the Region of Interest (ROI).",
            ),
            (
                self.radSkewnessCheckBox,
                "First-Order Skewness: asymmetry of the Standardized Uptake Value (SUV) distribution, including a tail toward high uptake.",
            ),
            (
                self.radContrastCheckBox,
                "Gray Level Co-occurrence Matrix (GLCM) Contrast: magnitude of local intensity differences between neighboring voxels; higher values indicate greater local heterogeneity.",
            ),
            (
                self.radSahgleCheckBox,
                "Gray Level Size Zone Matrix (GLSZM) Small Area High Gray Level Emphasis (SAHGLE): prevalence of small high-uptake zones.",
            ),
            (
                self.radLalgleCheckBox,
                "Gray Level Size Zone Matrix (GLSZM) Large Area Low Gray Level Emphasis (LALGLE): prevalence of large low-uptake zones.",
            ),
            (
                self.radZoneEntropyCheckBox,
                "Gray Level Size Zone Matrix (GLSZM) Zone Entropy: diversity of connected-zone sizes and intensity levels.",
            ),
        )

        # One feature per row keeps the expanded English names readable.
        for row, (checkBox, description) in enumerate(selectedRadiomics):
            checkBox.setChecked(True)
            checkBox.setToolTip(description)

            descriptionLabel = qt.QLabel(description)
            descriptionLabel.setWordWrap(True)
            descriptionLabel.setStyleSheet("color:#666;")
            descriptionLabel.setToolTip(description)

            selectedRadiomicsLayout.addWidget(checkBox, row, 0)
            selectedRadiomicsLayout.addWidget(descriptionLabel, row, 1)

        selectedRadiomicsLayout.setColumnStretch(1, 1)

        metricsLayout.addRow("Selected radiomics:", selectedRadiomicsWidget)

        self.radBinWidthSpinBox = qt.QDoubleSpinBox()
        self.radBinWidthSpinBox.setRange(0.001, 10.0)
        self.radBinWidthSpinBox.setDecimals(3)
        self.radBinWidthSpinBox.setSingleStep(0.05)
        self.radBinWidthSpinBox.setValue(0.25)
        self.radBinWidthSpinBox.setToolTip(
            "Fixed SUV bin width used for texture discretisation. Keep this "
            "constant across a cohort; test 0.05, 0.10 and 0.25 separately."
        )
        metricsLayout.addRow("Radiomics bin width:", self.radBinWidthSpinBox)

        resampleWidget = qt.QWidget()
        resampleLayout = qt.QHBoxLayout(resampleWidget)
        resampleLayout.setContentsMargins(0, 0, 0, 0)
        self.radResampleCheckBox = qt.QCheckBox("Isotropic resampling")
        self.radResampleCheckBox.setChecked(False)
        self.radResampleSpacingSpinBox = qt.QDoubleSpinBox()
        self.radResampleSpacingSpinBox.setRange(0.5, 20.0)
        self.radResampleSpacingSpinBox.setDecimals(2)
        self.radResampleSpacingSpinBox.setSingleStep(0.5)
        self.radResampleSpacingSpinBox.setValue(4.0)
        self.radResampleSpacingSpinBox.setSuffix(" mm")
        self.radResampleSpacingSpinBox.setEnabled(False)
        self.radResampleCheckBox.toggled.connect(
            self.radResampleSpacingSpinBox.setEnabled
        )
        resampleLayout.addWidget(self.radResampleCheckBox)
        resampleLayout.addWidget(self.radResampleSpacingSpinBox)
        metricsLayout.addRow("Texture geometry:", resampleWidget)

        self.outputExcelEdit = ctk.ctkPathLineEdit()
        self.outputExcelEdit.filters = ctk.ctkPathLineEdit.Files
        self.outputExcelEdit.nameFilters = ["Excel files (*.xlsx)"]
        metricsLayout.addRow("Output Excel: ", self.outputExcelEdit)

        self.appendExcelCheckBox = qt.QCheckBox("Append to existing file")
        self.appendExcelCheckBox.setChecked(True)
        metricsLayout.addRow("", self.appendExcelCheckBox)

        self.calculateMetricsButton = qt.QPushButton("Calculate Metrics")
        metricsLayout.addRow("", self.calculateMetricsButton)

        # -------------------- 4. QC: Hotspot & Outlier Analysis --------------------
        qcCollapsible = ctk.ctkCollapsibleButton()
        qcCollapsible.text = "4. QC: Hotspot & Outlier Analysis"
        self.interactiveLayout.addWidget(qcCollapsible)
        qcLayout = qt.QVBoxLayout(qcCollapsible)

        qcHelp = qt.QLabel(
            "Scans every segment (or every tracked ROI variant) for SUV metrics and "
            "the location of its hottest voxel. Segments whose SUVmax spikes relative "
            "to the cohort (a classic contamination signature) are flagged. "
            "Double-click a row to jump the scene to that hotspot and drop a marker."
        )
        qcHelp.setWordWrap(True)
        qcHelp.setStyleSheet("color:#555; font-style:italic;")
        qcLayout.addWidget(qcHelp)

        # ----- Source of rows -----
        qcSourceForm = qt.QFormLayout()
        qcLayout.addLayout(qcSourceForm)

        qcSourceBox = qt.QHBoxLayout()
        self.qcScanSegmentsButton = qt.QPushButton("Scan segments of selected ROI")
        self.qcScanSegmentsButton.setToolTip(
            "Add one QC row per segment in the currently selected Segmentation(ROI)."
        )
        self.qcAddCurrentButton = qt.QPushButton("+ Track current ROI as variant")
        self.qcAddCurrentButton.setToolTip(
            "Add the whole currently selected ROI as one row, so raw vs ablated "
            "variants can be compared side by side."
        )
        qcSourceBox.addWidget(self.qcScanSegmentsButton)
        qcSourceBox.addWidget(self.qcAddCurrentButton)
        qcSourceForm.addRow("Add rows:", qcSourceBox)

        # ----- Outlier sensitivity -----
        self.qcMadKSpinBox = qt.QDoubleSpinBox()
        self.qcMadKSpinBox.setMinimum(1.0)
        self.qcMadKSpinBox.setMaximum(10.0)
        self.qcMadKSpinBox.setSingleStep(0.5)
        self.qcMadKSpinBox.setValue(3.5)
        self.qcMadKSpinBox.setToolTip(
            "Robust outlier threshold. A segment is flagged when its SUVmax exceeds "
            "median + k * MAD across the analysed segments (k = this value)."
        )
        qcSourceForm.addRow("Outlier k (MAD):", self.qcMadKSpinBox)

        self.qcRatioSpinBox = qt.QDoubleSpinBox()
        self.qcRatioSpinBox.setMinimum(1.0)
        self.qcRatioSpinBox.setMaximum(50.0)
        self.qcRatioSpinBox.setSingleStep(0.5)
        self.qcRatioSpinBox.setValue(4.0)
        self.qcRatioSpinBox.setToolTip(
            "Also flag a segment when SUVmax / SUVmean exceeds this ratio. "
            "A focal hot voxel (contamination) inflates max while mean stays low."
        )
        qcSourceForm.addRow("Outlier max/mean ratio:", self.qcRatioSpinBox)

        # ----- QC table -----
        self.qcTable = qt.QTableWidget(0, 8)
        self.qcTable.setHorizontalHeaderLabels([
            "ROI node", "Segment", "SUVmean", "SUVmax",
            "SUVpeak", "max/mean", "ΔSUVmax% (vs base)", "Flag",
        ])
        self.qcTable.horizontalHeader().setStretchLastSection(True)
        self.qcTable.setSelectionBehavior(qt.QAbstractItemView.SelectRows)
        self.qcTable.setSelectionMode(qt.QAbstractItemView.SingleSelection)
        self.qcTable.setEditTriggers(qt.QAbstractItemView.NoEditTriggers)
        self.qcTable.setFixedHeight(200)
        qcLayout.addWidget(self.qcTable)

        # ----- Table action buttons -----
        qcButtonRow1 = qt.QHBoxLayout()
        self.qcComputeButton = qt.QPushButton("Compute QC table")
        self.qcComputeButton.setToolTip(
            "Fill SUV metrics + hotspot for every row and run outlier detection."
        )
        self.qcSetBaselineButton = qt.QPushButton("Set selected row as baseline")
        self.qcSetBaselineButton.setToolTip(
            "Use the selected row's SUVmax as the baseline for the ΔSUVmax% column "
            "(e.g. set the raw ROI as baseline, then read how much each ablation "
            "step dropped the max)."
        )
        qcButtonRow1.addWidget(self.qcComputeButton)
        qcButtonRow1.addWidget(self.qcSetBaselineButton)
        qcLayout.addLayout(qcButtonRow1)

        qcButtonRow2 = qt.QHBoxLayout()
        self.qcJumpButton = qt.QPushButton("Jump to hotspot (selected row)")
        self.qcJumpButton.setToolTip(
            "Centre the slice views on the selected segment's hottest voxel and "
            "drop a fiducial there (red if flagged, green otherwise)."
        )
        self.qcPlaceAllButton = qt.QPushButton("Place all hotspots in scene")
        self.qcPlaceAllButton.setToolTip(
            "Drop a fiducial at every row's hotspot. Flagged outliers are red."
        )
        qcButtonRow2.addWidget(self.qcJumpButton)
        qcButtonRow2.addWidget(self.qcPlaceAllButton)
        qcLayout.addLayout(qcButtonRow2)

        qcButtonRow3 = qt.QHBoxLayout()
        self.qcClearButton = qt.QPushButton("Clear table")
        self.qcExportQcButton = qt.QPushButton("Export QC table to Excel")
        self.qcExportQcButton.setToolTip(
            "Write the QC table to the 'Output Excel' file (a 'QC' sheet)."
        )
        qcButtonRow3.addWidget(self.qcClearButton)
        qcButtonRow3.addWidget(self.qcExportQcButton)
        qcLayout.addLayout(qcButtonRow3)

        self.qcSummaryLabel = qt.QLabel("")
        self.qcSummaryLabel.setWordWrap(True)
        qcLayout.addWidget(self.qcSummaryLabel)

        # QC state: list of row dicts + baseline SUVmax
        self._qcRows = []
        self._qcBaselineSuvMax = None

        self.interactiveLayout.addStretch(1)

        # Build the Batch and Visualization tabs.
        self.setupBatchTab(self.batchLayout)

        self._visualizationRecords = []
        self._visualizationMetricKeys = []
        self._visualizationPlotNodeIds = []
        self._visualizationCurrentChartNodeId = None
        self.setupVisualizationTab(self.visualizationLayout)

        # --------------------------------------------------
        # UI Connections
        # --------------------------------------------------
        self.useSpineSlicingCheckBox.toggled.connect(self.onSpineSlicingOptionsChanged)

        self.loadPETButton.connect("clicked(bool)", self.onLoadPETClicked)
        self.loadCTButton.connect("clicked(bool)", self.onLoadCTClicked)
        self.loadPETDicomButton.connect("clicked(bool)", self.onLoadPETDicomClicked)
        self.loadCTDicomButton.connect("clicked(bool)", self.onLoadCTDicomClicked)
        self.loadSegmentationButton.connect("clicked(bool)", self.onLoadSegmentationClicked)
        self.loadVertebraeButton.connect("clicked(bool)", self.onLoadVertebraeClicked)

        self.spineStartSelector.currentIndexChanged.connect(self.onSpineLevelSelectionChanged)
        self.spineEndSelector.currentIndexChanged.connect(self.onSpineLevelSelectionChanged)

        self.generateVertebraeButton.connect("clicked(bool)", self.onGenerateVertebraeClicked)

        self.computeSpineBoundsButton.connect("clicked(bool)", self.onComputeSpineBoundsClicked)
        self.applySpineSlicingButton.connect("clicked(bool)", self.onApplySpineSlicingClicked)

        self.useKidneyExclusionCheckBox.toggled.connect(self.onKidneyExclusionOptionsChanged)
        self.loadKidneyButton.connect("clicked(bool)", self.onLoadKidneyClicked)
        self.generateKidneyButton.connect("clicked(bool)", self.onGenerateKidneyClicked)
        self.applyKidneyExclusionButton.connect("clicked(bool)", self.onApplyKidneyExclusionClicked)

        self.useUreterCleanupCheckBox.toggled.connect(self.onUreterCleanupOptionsChanged)
        self.generateUreterMaskButton.connect("clicked(bool)",self.onGenerateUreterMaskClicked)
        self.applyUreterCleanupButton.connect("clicked(bool)", self.onApplyUreterCleanupClicked)

        self.calculateMetricsButton.connect("clicked(bool)", self.onCalculateMetricsClicked)

        # ----- QC / Hotspot connections -----
        self.qcScanSegmentsButton.connect("clicked(bool)", self.onQcScanSegmentsClicked)
        self.qcAddCurrentButton.connect("clicked(bool)", self.onQcAddCurrentClicked)
        self.qcComputeButton.connect("clicked(bool)", self.onQcComputeClicked)
        self.qcSetBaselineButton.connect("clicked(bool)", self.onQcSetBaselineClicked)
        self.qcJumpButton.connect("clicked(bool)", self.onQcJumpClicked)
        self.qcPlaceAllButton.connect("clicked(bool)", self.onQcPlaceAllClicked)
        self.qcClearButton.connect("clicked(bool)", self.onQcClearClicked)
        self.qcExportQcButton.connect("clicked(bool)", self.onQcExportClicked)
        self.qcTable.connect("cellDoubleClicked(int,int)", self.onQcRowDoubleClicked)

        
        # --------------------------------------------------
        # Initialization
        # --------------------------------------------------
        self.onSpineSlicingOptionsChanged()
        self.onKidneyExclusionOptionsChanged()
        self.onUreterCleanupOptionsChanged()

    # --------------------------------------------------
    # Callback Functions
    # --------------------------------------------------    
    def onSpineSlicingOptionsChanged(self, checked=False):
        useSpineSlicing = self.useSpineSlicingCheckBox.isChecked()

        self.vertebraeSelector.setEnabled(useSpineSlicing)
        self.loadVertebraeButton.setEnabled(useSpineSlicing)
        self.generateVertebraeButton.setEnabled(useSpineSlicing)

        self.spineStartSelector.setEnabled(useSpineSlicing)
        self.spineEndSelector.setEnabled(useSpineSlicing)

        self.computeSpineBoundsButton.setEnabled(useSpineSlicing)

        self.applySpineSlicingButton.setEnabled(useSpineSlicing)

    def onLoadPETClicked(self, checked=False):
        self.loadVolumeIntoSelector(
            title="Load PET Volume",
            selector=self.petSelector,
            roleName="PET"
        )
    
    def onLoadCTClicked(self, checked=False):
        self.loadVolumeIntoSelector(
            title="Load CT Volume",
            selector=self.ctSelector,
            roleName="CT"
        )

    def onLoadPETDicomClicked(self, checked=False):
        self.loadDicomIntoSelector(
            title="Select PET DICOM folder",
            selector=self.petSelector,
            roleName="PET",
            modalityPreference=("PT", "NM"),
            preferSUV=True,
        )

    def onLoadCTDicomClicked(self, checked=False):
        self.loadDicomIntoSelector(
            title="Select CT DICOM folder",
            selector=self.ctSelector,
            roleName="CT",
            modalityPreference=("CT",),
            preferSUV=False,
        )
    
    def onLoadSegmentationClicked(self, checked=False):
        self.loadSegmentationIntoSelector(
            title="Load Segmentation",
            selector=self.segmentationSelector,
            roleName="Segmentation"
        )
    
    # ----- Vertebrae related functions -----
    def onLoadVertebraeClicked(self, checked=False):
        self.loadSegmentationIntoSelector(
            title="Load Vertebrae Segmentation",
            selector=self.vertebraeSelector,
            roleName="Vertebrae"
        )
    
    def onSpineLevelSelectionChanged(self, checked=False):
        levels = self.collectSelectedSpineLevels()
        print(levels)

    def onGenerateVertebraeClicked(self, checked=False):
        print("onGenerateVertebraeClicked")
        try:
            ctNode = self.ctSelector.currentNode()

            vertebraeNode, generationInfo = self.logic.generateVertebraeFromCTNode(ctNode=ctNode)

            self.vertebraeSelector.setCurrentNode(vertebraeNode)

            message = (
                "Vertebrae segmentation generated.\n"
                f"CT: {ctNode.GetName()}\n"
                f"Vertebrae node: {vertebraeNode.GetName()}\n"
                f"Generated stems: {generationInfo['selectedStems']}\n"
                f"Output folder: {generationInfo['outputDir']}\n"
                f"Output files: {generationInfo['outputFiles']}\n"
            )

            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def onComputeSpineBoundsClicked(self, checked=False):
        try:
            vertebraeNode = self.vertebraeSelector.currentNode()
            selectedLevels = self.collectSelectedSpineLevels()

            if not selectedLevels:
                return
            
            boundsInfo = self.logic.computeSpineRangeZBounds(vertebraeNode, selectedLevels)
            
            message = (
                "Spine bounds computed.\n"
                f"Vertebrae node: {vertebraeNode.GetName()}\n"
                f"Selected levels: {boundsInfo['selectedLevels']}\n"
                f"Matched segments: {boundsInfo['segmentNames']}\n"
                f"Z inferior: {boundsInfo['zInferior']:.3f} mm\n"
                f"Z superior: {boundsInfo['zSuperior']:.3f} mm\n"
                f"Voxel count: {boundsInfo['voxelCount']}\n"
            )
                
            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def onApplySpineSlicingClicked(self, checked=False):
        try:
            roiNode = self.segmentationSelector.currentNode()
            vertebraeNode = self.vertebraeSelector.currentNode()
            referenceNode = self.ctSelector.currentNode()
            selectedLevels = self.collectSelectedSpineLevels()

            if not selectedLevels:
                return

            processedNode, slicingInfo = self.logic.applySpineRangeSlicing(
                roiNode,
                vertebraeNode,
                selectedLevels,
                referenceNode,
            )

            # For immediate testing, select the processed ROI.
            # The original ROI remains in the scene.
            self.segmentationSelector.setCurrentNode(processedNode)

            message = (
                "Spine slicing completed.\n"
                f"Original ROI: {roiNode.GetName()}\n"
                f"Processed ROI: {processedNode.GetName()}\n"
                f"Selected levels: {slicingInfo['selectedLevels']}\n"
                f"Matched segments: {slicingInfo['segmentNames']}\n"
                f"Z inferior: {slicingInfo['zInferior']:.3f} mm\n"
                f"Z superior: {slicingInfo['zSuperior']:.3f} mm\n"
                f"Original voxel count: {slicingInfo['originalVoxelCount']}\n"
                f"Remaining voxel count: {slicingInfo['remainingVoxelCount']}\n"
                f"Removed voxel count: {slicingInfo['removedVoxelCount']}\n"
            )

            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))
    
    # ----- Kidney exclusion related functions -----
    def onKidneyExclusionOptionsChanged(self, checked=False):
        useKidneyExclusion = self.useKidneyExclusionCheckBox.isChecked()

        self.kidneySelector.setEnabled(useKidneyExclusion)
        self.loadKidneyButton.setEnabled(useKidneyExclusion)
        self.generateKidneyButton.setEnabled(useKidneyExclusion)
        self.kidneyDilationSpinBox.setEnabled(useKidneyExclusion)
        self.applyKidneyExclusionButton.setEnabled(useKidneyExclusion)

    def onApplyKidneyExclusionClicked(self, checked=False):
        try:
            roiNode = self.segmentationSelector.currentNode()
            kidneyNode = self.kidneySelector.currentNode()
            referenceNode = self.ctSelector.currentNode()
            dilationRadiusMm = self.kidneyDilationSpinBox.value

            processedNode, exclusionInfo = self.logic.applyKidneyExclusion(
                roiNode=roiNode,
                kidneyNode=kidneyNode,
                referenceVolumeNode=referenceNode,
                dilationRadiusMm=dilationRadiusMm,
            )

            # 테스트 편의를 위해 processed ROI를 현재 ROI selector에 자동 선택
            self.segmentationSelector.setCurrentNode(processedNode)

            message = (
                "Kidney exclusion completed.\n"
                f"Original ROI: {roiNode.GetName()}\n"
                f"Processed ROI: {processedNode.GetName()}\n"
                f"Kidney node: {kidneyNode.GetName()}\n"
                f"Kidney segments: {exclusionInfo['kidneySegmentNames']}\n"
                f"Dilation radius: {exclusionInfo['dilationRadiusMm']:.3f} mm\n"
                f"Original voxel count: {exclusionInfo['originalVoxelCount']}\n"
                f"Remaining voxel count: {exclusionInfo['remainingVoxelCount']}\n"
                f"Removed voxel count: {exclusionInfo['removedVoxelCount']}\n"
            )

            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))
    
    def onLoadKidneyClicked(self, checked=False):
        self.loadSegmentationIntoSelector(
            title="Load Kidney Segmentation",
            selector=self.kidneySelector,
            roleName="Kidneys",
        )

    def onGenerateKidneyClicked(self, checked=False):
        try:
            ctNode = self.ctSelector.currentNode()

            kidneyNode, generationInfo = self.logic.generateKidneysFromCTNode(ctNode=ctNode)

            self.kidneySelector.setCurrentNode(kidneyNode)

            message = (
                "Kidney segmentation generated.\n"
                f"CT: {ctNode.GetName()}\n"
                f"Kidney node: {kidneyNode.GetName()}\n"
                f"Generated stems: {generationInfo['selectedStems']}\n"
                f"Output folder: {generationInfo['outputDir']}\n"
                f"Output files: {generationInfo['outputFiles']}\n"
            )

            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    
    # ----- Ureter cleanup related functions -----
    def onUreterCleanupOptionsChanged(self, checked=False):
        useUreterCleanup = self.useUreterCleanupCheckBox.isChecked()

        self.ureterMaskSelector.setEnabled(useUreterCleanup)
        self.generateUreterMaskButton.setEnabled(useUreterCleanup)
        self.ureterSUVThresholdSpinBox.setEnabled(useUreterCleanup)
        self.ureterDilationSpinBox.setEnabled(useUreterCleanup)

        self.ureterCleanSUVThresholdSpinBox.setEnabled(useUreterCleanup)
        self.applyUreterCleanupButton.setEnabled(useUreterCleanup)

    def onGenerateUreterMaskClicked(self, checked=False):
        try:
            petNode = self.petSelector.currentNode()
            vertebraeNode = self.vertebraeSelector.currentNode()
            selectedLevels = self.collectSelectedSpineLevels()

            if not selectedLevels:
                return

            suvThreshold = self.ureterSUVThresholdSpinBox.value
            dilationRadiusMm = self.ureterDilationSpinBox.value

            ureterMaskNode, maskInfo = self.logic.generatePETDerivedUreterMask(
                petNode=petNode,
                vertebraeNode=vertebraeNode,
                selectedLevels=selectedLevels,
                suvThreshold=suvThreshold,
                dilationRadiusMm=dilationRadiusMm,
            )

            self.ureterMaskSelector.setCurrentNode(ureterMaskNode)

            message = (
                "PET-derived urinary activity mask generated.\n"
                f"PET: {petNode.GetName()}\n"
                f"Mask node: {ureterMaskNode.GetName()}\n"
                f"Selected spine levels: {maskInfo['selectedLevels']}\n"
                f"Z range used: {maskInfo['zInferiorUsed']:.3f} to {maskInfo['zSuperiorUsed']:.3f} mm\n"
                f"SUV threshold: {maskInfo['suvThreshold']:.3f}\n"
                f"Dilation radius: {maskInfo['dilationRadiusMm']:.3f} mm\n"
                f"Initial hot voxel count: {maskInfo['initialHotVoxelCount']}\n"
                f"After Z clipping: {maskInfo['zClippedHotVoxelCount']}\n"
                f"After bladder removal: {maskInfo['afterBladderRemovalVoxelCount']}\n"
                f"Final mask voxel count: {maskInfo['finalMaskVoxelCount']}\n"
            )

            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def onApplyUreterCleanupClicked(self, checked=False):
        try:
            roiNode = self.segmentationSelector.currentNode()
            petNode = self.petSelector.currentNode()
            ureterMaskNode = self.ureterMaskSelector.currentNode()
            cleanSUVThreshold = self.ureterCleanSUVThresholdSpinBox.value

            processedNode, cleanupInfo = self.logic.applyUrinaryActivityCleanup(
                roiNode=roiNode,
                petNode=petNode,
                urinaryMaskNode=ureterMaskNode,
                cleanSUVThreshold=cleanSUVThreshold,
            )

            # 테스트 편의를 위해 processed ROI를 현재 ROI selector에 자동 선택
            self.segmentationSelector.setCurrentNode(processedNode)

            message = (
                "Urinary activity cleanup completed.\n"
                f"Original ROI: {roiNode.GetName()}\n"
                f"Processed ROI: {processedNode.GetName()}\n"
                f"Urinary mask: {ureterMaskNode.GetName()}\n"
                f"Cleanup SUV threshold: {cleanupInfo['cleanSUVThreshold']:.3f}\n"
                f"Original voxel count: {cleanupInfo['originalVoxelCount']}\n"
                f"Overlap voxel count: {cleanupInfo['overlapVoxelCount']}\n"
                f"Removed voxel count: {cleanupInfo['removedVoxelCount']}\n"
                f"Remaining voxel count: {cleanupInfo['remainingVoxelCount']}\n"
            )

            print("[PETBiomarkerExtractor]")
            print(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))
    
    # ----- Metrics calculation related functions -----
    def onCalculateMetricsClicked(self, checked=False):
        try:
            petNode = self.petSelector.currentNode()
            roiNode = self.segmentationSelector.currentNode()
            metricsOptions = self.collectMetricOptions()
            radiomicsOptions = self.collectRadiomicsOptions()

            results = self.logic.runPetIndic(
                petNode=petNode,
                roiNode=roiNode,
                metricsOptions=metricsOptions,
            )

            radiomicsStatus = "not_run"

            if self.logic.isRadiomicsEnabled(radiomicsOptions):
                try:
                    radResults = self.logic.runRadiomics(
                        petNode=petNode,
                        roiNode=roiNode,
                        radiomicsOptions=radiomicsOptions,
                    )
                    results.update(radResults)
                    radiomicsStatus = "done"
                except Exception as radError:
                    radiomicsStatus = f"radiomics_error: {str(radError)[:120]}"
                    print("[PETBiomarkerExtractor] Radiomics failed:")
                    print(radError)

            processingLabel = self.inferProcessingLabel(roiNode)

            row = self.makeMetricResultRow(
                petNode=petNode,
                roiNode=roiNode,
                results=results,
                processingLabel=processingLabel
            )

            row["radiomics_status"] = radiomicsStatus

            outputFile = self.outputExcelEdit.currentPath

            savedPath = None
            if outputFile:
                savedPath = self.logic.saveMetricRowsToExcel(
                    rows=[row],
                    outputFile=outputFile,
                    append=self.appendExcelCheckBox.isChecked()
                )
            
            message = self.formatPetMetricResults(
                petNode=petNode,
                roiNode=roiNode,
                results=results
            )

            if radiomicsStatus == "done":
                radCount = len([k for k in results.keys() if k.startswith("rad_")])
                message += f"\nRadiomics features: {radCount}"

            if savedPath:
                message += f"\nSaved to: {savedPath}"

                if self.visualizationAutoShowCheckBox.isChecked():
                    self.showVisualizationFromExcel(
                        savedPath,
                        preferredSegment=roiNode.GetName(),
                        preferredMetric="suv_max",
                    )

            print("[PETBiomarkerExtractor]")
            print(message)
            slicer.util.infoDisplay(message)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    # --------------------------------------------------
    # Helper Functions
    # --------------------------------------------------
    def loadVolumeIntoSelector(self, title, selector, roleName):
        filePath = qt.QFileDialog.getOpenFileName(
            slicer.util.mainWindow(),
            title,
            "",
            "Volume files (*.nii; *.nii.gz; *.dcm; *.dicom)",
        )

        if not filePath:
            return
        
        try:
            node = self.logic.loadVolumeNodeFromFile(filePath)
            selector.setCurrentNode(node)

            message = f"{roleName} loaded: {node.GetName()}"
            print("[PETBiomarkerExtractor]", message)
        except Exception as e:
            slicer.util.errorDisplay(str(e))

    def loadDicomIntoSelector(self, title, selector, roleName,
                              modalityPreference=None, preferSUV=False):
        dicomDir = qt.QFileDialog.getExistingDirectory(
            slicer.util.mainWindow(),
            title,
            "",
        )

        if not dicomDir:
            return

        try:
            node = self.logic.loadDicomDirAsVolumeNode(
                dicomDir=dicomDir,
                modalityPreference=modalityPreference,
                preferSUV=preferSUV,
                roleName=roleName,
            )
            selector.setCurrentNode(node)

            message = f"{roleName} loaded from DICOM: {node.GetName()}"
            print("[PETBiomarkerStudio]", message)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def loadSegmentationIntoSelector(self, title, selector, roleName):
        filePath = qt.QFileDialog.getOpenFileName(
            slicer.util.mainWindow(),
            title,
            "",
            "Segmentation files (*.nrrd; *.seg.nrrd; *.nii; *.nii.gz )",
        )

        if not filePath:
            return
        
        try:
            node = self.logic.loadSegmentationNodeFromFile(filePath)
            selector.setCurrentNode(node)

            message = f"{roleName} loaded: {node.GetName()}"
            print("[PETBiomarkerExtractor]", message)
        except Exception as e:
            slicer.util.errorDisplay(str(e))
    
    def collectMetricOptions(self):
        return {
            "mean": self.suvMeanCheckBox.isChecked(),
            "max": self.suvMaxCheckBox.isChecked(),
            "peak": self.suvPeakCheckBox.isChecked(),
            "tlg": self.tlgCheckBox.isChecked(),
            "volume": self.volumeCheckBox.isChecked(),
        }

    def formatPetMetricResults(self, petNode, roiNode, results):
        lines = [
            "PET metrics completed.",
            f"PET: {petNode.GetName()}",
            f"ROI: {roiNode.GetName()}",
        ]

        displayNames = {
            "suv_mean": "SUVmean",
            "suv_max": "SUVmax",
            "suv_peak": "SUVpeak",
            "tlg": "TLG",
            "volume_mL": "Volume (mL)",

        }

        for key in displayNames:
            if key in results:
                lines.append(f"{displayNames[key]}: {results[key]:.6g}")
        
        return "\n".join(lines)

    def collectSelectedSpineLevels(self):
        levels = ["L1", "L2", "L3", "L4", "L5"]

        start = self.spineStartSelector.currentText
        end = self.spineEndSelector.currentText

        startIndex = self.spineStartSelector.currentIndex
        endIndex = self.spineEndSelector.currentIndex

        if startIndex > endIndex:
            slicer.util.errorDisplay(
                f"Invalid spine range: {start} to {end}."
                "Start level must be above or equal to end level."
            )
            return []

        return levels[startIndex:endIndex+1]

    def makeMetricResultRow(self, petNode, roiNode, results, processingLabel="raw"):
        import datetime

        row = {
            "subject_id": "",
            "patient_id": "",
            "scan_date": "",
            "pet_node": petNode.GetName(),
            "roi_node": roiNode.GetName(),
            "segment": roiNode.GetName(),
            "processing": processingLabel,
            "status": "done",
            "computed_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        row.update(results)
        return row

    def collectRadiomicsOptions(self):
        selectedFeatureKeys = []
        featureCheckBoxes = (
            ("p10", self.radP10CheckBox),
            ("p90", self.radP90CheckBox),
            ("entropy", self.radEntropyCheckBox),
            ("skewness", self.radSkewnessCheckBox),
            ("contrast", self.radContrastCheckBox),
            ("sahgle", self.radSahgleCheckBox),
            ("lalgle", self.radLalgleCheckBox),
            ("zone_entropy", self.radZoneEntropyCheckBox),
        )
        for featureKey, checkBox in featureCheckBoxes:
            if checkBox.isChecked():
                selectedFeatureKeys.append(featureKey)

        return {
            "selected_feature_keys": selectedFeatureKeys,
            "derived": False,
            "bin_width": float(self.radBinWidthSpinBox.value),
            "resample_isotropic": self.radResampleCheckBox.isChecked(),
            "resampled_spacing_mm": float(self.radResampleSpacingSpinBox.value),
            # Retained as false for compatibility with the existing logic/API.
            "firstorder": False,
            "shape": False,
            "glcm": False,
            "glrlm": False,
            "glszm": False,
            "gldm": False,
            "ngtdm": False,
        }

    def inferProcessingLabel(self, roiNode):
        name = roiNode.GetName().lower()

        labels = []

        if "_spine_" in name:
            labels.append("spine_sliced")

        if "kidney_excluded" in name:
            labels.append("kidney_excluded")
        if "urinary_cleaned" in name:
            labels.append("urinary_cleaned")

        if not labels:
            return "raw"

        return "+".join(labels)

    # --------------------------------------------------
    # QC / Hotspot callbacks
    # --------------------------------------------------
    def onQcScanSegmentsClicked(self, checked=False):
        roiNode = self.segmentationSelector.currentNode()
        if roiNode is None:
            slicer.util.errorDisplay("Select a Segmentation(ROI) first.")
            return
        try:
            segmentation = roiNode.GetSegmentation()
            n = segmentation.GetNumberOfSegments()
            if n == 0:
                slicer.util.errorDisplay(f"Segmentation has no segments: {roiNode.GetName()}")
                return
            added = 0
            for i in range(n):
                segId = segmentation.GetNthSegmentID(i)
                segName = segmentation.GetNthSegment(i).GetName()
                self._qcRows.append(self._makeQcRow(roiNode, segId, segName))
                added += 1
            self._refreshQcTable()
            self.qcSummaryLabel.text = f"Added {added} segment row(s) from {roiNode.GetName()}."
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def onQcAddCurrentClicked(self, checked=False):
        roiNode = self.segmentationSelector.currentNode()
        if roiNode is None:
            slicer.util.errorDisplay("Select a Segmentation(ROI) first.")
            return
        segmentation = roiNode.GetSegmentation()
        if segmentation.GetNumberOfSegments() == 0:
            slicer.util.errorDisplay(f"Segmentation has no segments: {roiNode.GetName()}")
            return
        # Track the whole ROI as one variant -> use its first segment for the hotspot
        segId = segmentation.GetNthSegmentID(0)
        segName = segmentation.GetNthSegment(0).GetName()
        self._qcRows.append(self._makeQcRow(roiNode, segId, segName, wholeRoi=True))
        self._refreshQcTable()
        self.qcSummaryLabel.text = f"Tracking variant: {roiNode.GetName()} ({segName})."

    def _makeQcRow(self, roiNode, segmentId, segmentName, wholeRoi=False):
        return {
            "roiNodeId": roiNode.GetID(),
            "roiNodeName": roiNode.GetName(),
            "segmentId": segmentId,
            "segmentName": segmentName,
            "wholeRoi": wholeRoi,
            "suv_mean": None,
            "suv_max": None,
            "suv_peak": None,
            "ratio": None,
            "deltaMaxPct": None,
            "rasHotspot": None,
            "flag": "",
        }

    def onQcComputeClicked(self, checked=False):
        petNode = self.petSelector.currentNode()
        if petNode is None:
            slicer.util.errorDisplay("Select a PET volume first.")
            return
        if not self._qcRows:
            slicer.util.errorDisplay(
                "No QC rows. Use 'Scan segments' or 'Track current ROI' first."
            )
            return
        try:
            self._qcRows = self.logic.computeQcRows(petNode, self._qcRows)

            madK = self.qcMadKSpinBox.value
            ratioThresh = self.qcRatioSpinBox.value
            self._qcRows = self.logic.flagQcOutliers(
                self._qcRows, madK=madK, ratioThresh=ratioThresh
            )

            # Delta vs baseline
            for row in self._qcRows:
                if self._qcBaselineSuvMax and row["suv_max"] is not None:
                    base = self._qcBaselineSuvMax
                    if base != 0:
                        row["deltaMaxPct"] = (row["suv_max"] - base) / base * 100.0
                else:
                    row["deltaMaxPct"] = None

            self._refreshQcTable()
            nFlag = sum(1 for r in self._qcRows if r["flag"])
            self.qcSummaryLabel.text = (
                f"Computed {len(self._qcRows)} row(s). "
                f"{nFlag} flagged as SUVmax outlier(s)."
            )
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def onQcSetBaselineClicked(self, checked=False):
        row = self._currentQcRow()
        if row is None:
            slicer.util.errorDisplay("Select a row first.")
            return
        if row["suv_max"] is None:
            slicer.util.errorDisplay("Compute the QC table before setting a baseline.")
            return
        self._qcBaselineSuvMax = row["suv_max"]
        for r in self._qcRows:
            if r["suv_max"] is not None and self._qcBaselineSuvMax:
                r["deltaMaxPct"] = (
                    (r["suv_max"] - self._qcBaselineSuvMax) / self._qcBaselineSuvMax * 100.0
                )
        self._refreshQcTable()
        self.qcSummaryLabel.text = (
            f"Baseline SUVmax = {self._qcBaselineSuvMax:.3f} "
            f"({row['roiNodeName']} / {row['segmentName']})."
        )

    def onQcJumpClicked(self, checked=False):
        row = self._currentQcRow()
        if row is None:
            slicer.util.errorDisplay("Select a row first.")
            return
        self._jumpToQcRow(row, placeFiducial=True)

    def onQcRowDoubleClicked(self, rowIdx, colIdx):
        if 0 <= rowIdx < len(self._qcRows):
            self._jumpToQcRow(self._qcRows[rowIdx], placeFiducial=True)

    def _jumpToQcRow(self, row, placeFiducial=True):
        if row.get("rasHotspot") is None:
            slicer.util.errorDisplay("Compute the QC table first (no hotspot location yet).")
            return
        isOutlier = bool(row["flag"])
        self.logic.jumpToHotspot(
            ras=row["rasHotspot"],
            label=f"{row['segmentName']} SUVmax={row['suv_max']:.1f}"
                  if row["suv_max"] is not None else row["segmentName"],
            placeFiducial=placeFiducial,
            isOutlier=isOutlier,
        )

    def onQcPlaceAllClicked(self, checked=False):
        placed = 0
        for row in self._qcRows:
            if row.get("rasHotspot") is None:
                continue
            self.logic.jumpToHotspot(
                ras=row["rasHotspot"],
                label=f"{row['segmentName']} SUVmax={row['suv_max']:.1f}"
                      if row["suv_max"] is not None else row["segmentName"],
                placeFiducial=True,
                isOutlier=bool(row["flag"]),
                jump=False,
            )
            placed += 1
        self.qcSummaryLabel.text = f"Placed {placed} hotspot fiducial(s) in the scene."

    def onQcClearClicked(self, checked=False):
        self._qcRows = []
        self._qcBaselineSuvMax = None
        self._refreshQcTable()
        self.qcSummaryLabel.text = "QC table cleared."

    def onQcExportClicked(self, checked=False):
        outputFile = self.outputExcelEdit.currentPath
        if not outputFile:
            slicer.util.errorDisplay("Set an 'Output Excel' path in section 3 first.")
            return
        if not self._qcRows:
            slicer.util.errorDisplay("No QC rows to export.")
            return
        try:
            savedPath = self.logic.saveQcRowsToExcel(
                self._qcRows,
                outputFile,
                baselineSuvMax=self._qcBaselineSuvMax,
            )
            self.qcSummaryLabel.text = f"QC table saved to: {savedPath}"
            slicer.util.infoDisplay(f"QC table saved to:\n{savedPath}")
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            slicer.util.errorDisplay(str(e))

    def _currentQcRow(self):
        rowIdx = self.qcTable.currentRow()
        if 0 <= rowIdx < len(self._qcRows):
            return self._qcRows[rowIdx]
        return None

    def _refreshQcTable(self):
        def fmt(v):
            return "" if v is None else f"{v:.3f}"

        self.qcTable.setRowCount(len(self._qcRows))
        for r, row in enumerate(self._qcRows):
            ras = row.get("rasHotspot")
            values = [
                row["roiNodeName"],
                row["segmentName"],
                fmt(row["suv_mean"]),
                fmt(row["suv_max"]),
                fmt(row["suv_peak"]),
                fmt(row["ratio"]),
                "" if row["deltaMaxPct"] is None else f"{row['deltaMaxPct']:+.1f}%",
                row["flag"],
            ]
            for c, val in enumerate(values):
                item = qt.QTableWidgetItem(str(val))
                if row["flag"]:
                    item.setBackground(qt.QColor(255, 224, 224))  # light red
                self.qcTable.setItem(r, c, item)
        self.qcTable.resizeColumnsToContents()


    # ==================================================================
    # Cohort visualization
    #
    # Quantification produces one row per subject and segment. A histogram or
    # ranked-value plot is only meaningful after these rows have accumulated.
    # Therefore the visualization reads the exported Data sheet and keeps the
    # selected segment fixed while comparing one metric across subjects.
    # ==================================================================
    def setupVisualizationTab(self, layout):
        helpLabel = qt.QLabel(
            "Visualize one quantified metric across subjects. The selected "
            "segment is held constant, the three largest values are highlighted "
            "in red, and cohort statistics are shown beside the Slicer plot view. "
            "Batch figure export can also save the same figure type for "
            "multiple metrics of one segment at once."
        )
        helpLabel.setWordWrap(True)
        helpLabel.setStyleSheet("color:#555; font-style:italic;")
        layout.addWidget(helpLabel)

        sourceCollapsible = ctk.ctkCollapsibleButton()
        sourceCollapsible.text = "1. Data source"
        layout.addWidget(sourceCollapsible)
        sourceForm = qt.QFormLayout(sourceCollapsible)

        sourceWidget = qt.QWidget()
        sourceRow = qt.QHBoxLayout(sourceWidget)
        sourceRow.setContentsMargins(0, 0, 0, 0)

        self.visualizationExcelEdit = ctk.ctkPathLineEdit()
        self.visualizationExcelEdit.filters = ctk.ctkPathLineEdit.Files
        self.visualizationExcelEdit.nameFilters = ["Excel files (*.xlsx)"]
        self.visualizationExcelEdit.setToolTip(
            "Excel file generated by PET Biomarker Studio. The Data sheet is used."
        )

        self.visualizationRefreshButton = qt.QPushButton("Load / Refresh")
        self.visualizationRefreshButton.setToolTip(
            "Reload the Data sheet and rebuild the metric and segment selections."
        )

        sourceRow.addWidget(self.visualizationExcelEdit)
        sourceRow.addWidget(self.visualizationRefreshButton)
        sourceForm.addRow("Quantification Excel:", sourceWidget)

        self.visualizationAutoShowCheckBox = qt.QCheckBox(
            "Automatically show the figure after quantification"
        )
        self.visualizationAutoShowCheckBox.setChecked(True)
        sourceForm.addRow("", self.visualizationAutoShowCheckBox)

        controlsCollapsible = ctk.ctkCollapsibleButton()
        controlsCollapsible.text = "2. Figure selection"
        layout.addWidget(controlsCollapsible)
        controlsForm = qt.QFormLayout(controlsCollapsible)

        self.visualizationSegmentComboBox = qt.QComboBox()
        self.visualizationSegmentComboBox.setToolTip(
            "Select one anatomical segment. Mixing different organs in one "
            "distribution would make the cohort statistics difficult to interpret."
        )
        controlsForm.addRow("Segment:", self.visualizationSegmentComboBox)

        self.visualizationMetricComboBox = qt.QComboBox()
        self.visualizationMetricComboBox.setToolTip(
            "Select the SUV or radiomic metric to display."
        )
        controlsForm.addRow("Metric:", self.visualizationMetricComboBox)

        self.visualizationBatchMetricList = qt.QListWidget()
        self.visualizationBatchMetricList.setSelectionMode(
            qt.QAbstractItemView.MultiSelection
        )
        self.visualizationBatchMetricList.setToolTip(
            "Select multiple metrics when you want to generate and save the "
            "same figure type for one segment at once. The current Segment "
            "and Figure selections are reused for every selected metric."
        )
        self.visualizationBatchMetricList.setMinimumHeight(130)
        self.visualizationBatchMetricList.setMaximumHeight(160)

        batchMetricWidget = qt.QWidget()
        batchMetricLayout = qt.QVBoxLayout(batchMetricWidget)
        batchMetricLayout.setContentsMargins(0, 0, 0, 0)
        batchMetricLayout.addWidget(self.visualizationBatchMetricList)

        batchMetricButtonRow = qt.QHBoxLayout()
        self.visualizationSelectAllMetricsButton = qt.QPushButton("Select All")
        self.visualizationClearMetricsButton = qt.QPushButton("Clear")
        batchMetricButtonRow.addWidget(self.visualizationSelectAllMetricsButton)
        batchMetricButtonRow.addWidget(self.visualizationClearMetricsButton)
        batchMetricButtonRow.addStretch(1)
        batchMetricLayout.addLayout(batchMetricButtonRow)

        controlsForm.addRow("Batch metrics:", batchMetricWidget)

        self.visualizationPlotTypeComboBox = qt.QComboBox()
        self.visualizationPlotTypeComboBox.addItems([
            "Histogram",
            "Values by patient",
            "Ranked values",
        ])
        self.visualizationPlotTypeComboBox.setToolTip(
            "Histogram shows the cohort distribution. Values by patient plots "
            "one bar per patient in subject-ID order. Ranked values show every "
            "patient from the largest value to the smallest value."
        )
        controlsForm.addRow("Figure:", self.visualizationPlotTypeComboBox)

        self.visualizationShowButton = qt.QPushButton("Show Figure")
        controlsForm.addRow("", self.visualizationShowButton)

        saveWidget = qt.QWidget()
        saveRow = qt.QHBoxLayout(saveWidget)
        saveRow.setContentsMargins(0, 0, 0, 0)

        self.visualizationFigurePathEdit = ctk.ctkPathLineEdit()
        self.visualizationFigurePathEdit.filters = ctk.ctkPathLineEdit.Files
        self.visualizationFigurePathEdit.nameFilters = [
            "Figure files (*.png *.svg)",
            "PNG image (*.png)",
            "SVG vector image (*.svg)",
        ]
        self.visualizationFigurePathEdit.setToolTip(
            "Save the currently displayed plot as a PNG image or an SVG vector "
            "graphic. The default folder is Downloads."
        )

        self.visualizationSaveFigureButton = qt.QPushButton("Save Figure")
        self.visualizationSaveFigureButton.setToolTip(
            "Save the plot currently shown in the Slicer plot view."
        )

        saveRow.addWidget(self.visualizationFigurePathEdit)
        saveRow.addWidget(self.visualizationSaveFigureButton)
        controlsForm.addRow("Figure output:", saveWidget)

        self.visualizationSaveSelectedFiguresButton = qt.QPushButton(
            "Generate + Save Selected Metric Figures"
        )
        self.visualizationSaveSelectedFiguresButton.setToolTip(
            "Create a separate figure for every metric selected in 'Batch "
            "metrics' and save all figure files to the chosen output folder. "
            "Because the Slicer plot view can show only one chart at a time, "
            "batch generation saves the figures as separate image files."
        )
        controlsForm.addRow("", self.visualizationSaveSelectedFiguresButton)

        statisticsCollapsible = ctk.ctkCollapsibleButton()
        statisticsCollapsible.text = "3. Cohort statistics"
        layout.addWidget(statisticsCollapsible)
        statisticsLayout = qt.QVBoxLayout(statisticsCollapsible)

        statisticsWidget = qt.QWidget()
        statisticsGrid = qt.QGridLayout(statisticsWidget)
        statisticsGrid.setContentsMargins(0, 0, 0, 0)

        statisticDefinitions = (
            ("N", "Number of valid subject-level values."),
            ("Mean", "Arithmetic mean across subjects."),
            ("Standard deviation", "Sample standard deviation across subjects."),
            ("Median", "50th percentile across subjects."),
            ("Interquartile range", "75th percentile minus 25th percentile."),
            ("Tukey outliers", "Values outside Q1 - 1.5 × IQR or Q3 + 1.5 × IQR."),
        )
        self.visualizationStatisticLabels = {}
        for row, (name, tooltip) in enumerate(statisticDefinitions):
            nameLabel = qt.QLabel(name + ":")
            nameLabel.setToolTip(tooltip)
            valueLabel = qt.QLabel("—")
            valueLabel.setTextInteractionFlags(qt.Qt.TextSelectableByMouse)
            valueLabel.setToolTip(tooltip)
            statisticsGrid.addWidget(nameLabel, row, 0)
            statisticsGrid.addWidget(valueLabel, row, 1)
            self.visualizationStatisticLabels[name] = valueLabel

        statisticsGrid.setColumnStretch(1, 1)
        statisticsLayout.addWidget(statisticsWidget)

        top3Label = qt.QLabel("Top 3 largest values")
        top3Label.setStyleSheet("font-weight:bold;")
        statisticsLayout.addWidget(top3Label)

        self.visualizationTop3Table = qt.QTableWidget(0, 4)
        self.visualizationTop3Table.setHorizontalHeaderLabels([
            "Rank", "Subject", "Scan date", "Value"
        ])
        self.visualizationTop3Table.setEditTriggers(
            qt.QAbstractItemView.NoEditTriggers
        )
        self.visualizationTop3Table.setSelectionBehavior(
            qt.QAbstractItemView.SelectRows
        )
        self.visualizationTop3Table.horizontalHeader().setStretchLastSection(True)
        self.visualizationTop3Table.setFixedHeight(125)
        statisticsLayout.addWidget(self.visualizationTop3Table)

        self.visualizationPatientTableLabel = qt.QLabel(
            "Patient values shown in the figure"
        )
        self.visualizationPatientTableLabel.setStyleSheet("font-weight:bold;")
        statisticsLayout.addWidget(self.visualizationPatientTableLabel)

        self.visualizationPatientTable = qt.QTableWidget(0, 6)
        self.visualizationPatientTable.setHorizontalHeaderLabels([
            "Plot index", "Subject", "Patient ID", "Scan date", "Value", "Top 3"
        ])
        self.visualizationPatientTable.setEditTriggers(
            qt.QAbstractItemView.NoEditTriggers
        )
        self.visualizationPatientTable.setSelectionBehavior(
            qt.QAbstractItemView.SelectRows
        )
        self.visualizationPatientTable.horizontalHeader().setStretchLastSection(True)
        self.visualizationPatientTable.setFixedHeight(190)
        statisticsLayout.addWidget(self.visualizationPatientTable)
        self.visualizationPatientTableLabel.setVisible(False)
        self.visualizationPatientTable.setVisible(False)

        self.visualizationStatusLabel = qt.QLabel("No quantification file loaded.")
        self.visualizationStatusLabel.setWordWrap(True)
        self.visualizationStatusLabel.setTextInteractionFlags(
            qt.Qt.TextSelectableByMouse
        )
        layout.addWidget(self.visualizationStatusLabel)
        layout.addStretch(1)

        self.visualizationRefreshButton.connect(
            "clicked(bool)", self.onVisualizationRefreshClicked
        )
        self.visualizationShowButton.connect(
            "clicked(bool)", self.onVisualizationShowClicked
        )
        self.visualizationSaveFigureButton.connect(
            "clicked(bool)", self.onVisualizationSaveFigureClicked
        )
        self.visualizationSaveSelectedFiguresButton.connect(
            "clicked(bool)", self.onVisualizationSaveSelectedFiguresClicked
        )
        self.visualizationSelectAllMetricsButton.connect(
            "clicked(bool)", self.onVisualizationSelectAllMetricsClicked
        )
        self.visualizationClearMetricsButton.connect(
            "clicked(bool)", self.onVisualizationClearMetricsClicked
        )
        self.visualizationSegmentComboBox.connect(
            "currentIndexChanged(int)", self.onVisualizationSegmentChanged
        )
        self.visualizationMetricComboBox.connect(
            "currentIndexChanged(int)", self.onVisualizationSelectionChanged
        )
        self.visualizationPlotTypeComboBox.connect(
            "currentIndexChanged(int)", self.onVisualizationSelectionChanged
        )

    def onVisualizationRefreshClicked(self, checked=False):
        path = self.visualizationExcelEdit.currentPath
        if not path:
            self.visualizationStatusLabel.text = "Select a quantification Excel file."
            return
        try:
            self.loadVisualizationData(path)
            self.updateVisualizationFigure()
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.visualizationStatusLabel.text = f"Visualization error: {e}"
            slicer.util.errorDisplay(str(e))

    def onVisualizationShowClicked(self, checked=False):
        try:
            self.updateVisualizationFigure()
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.visualizationStatusLabel.text = f"Visualization error: {e}"
            slicer.util.errorDisplay(str(e))

    def onVisualizationSelectAllMetricsClicked(self, checked=False):
        for index in range(self.visualizationBatchMetricList.count):
            item = self.visualizationBatchMetricList.item(index)
            item.setSelected(True)

    def onVisualizationClearMetricsClicked(self, checked=False):
        self.visualizationBatchMetricList.clearSelection()

    def onVisualizationSaveSelectedFiguresClicked(self, checked=False):
        try:
            segment = self.visualizationSegmentComboBox.currentText
            if not segment:
                raise ValueError("Select a segment first.")

            selectedMetricKeys = self.selectedVisualizationBatchMetricKeys()
            if not selectedMetricKeys:
                raise ValueError(
                    "Select at least one item in 'Batch metrics' first."
                )

            plotType = self.visualizationPlotTypeComboBox.currentText
            if not plotType:
                raise ValueError(
                    "Select a figure type before batch figure export."
                )

            basePath = str(self.visualizationFigurePathEdit.currentPath or "").strip()
            outputDirectory, extension = self.resolveVisualizationOutputSettings(
                basePath
            )

            savedPaths = []
            skippedMetrics = []
            lastRenderedMetricKey = None

            for metricKey in selectedMetricKeys:
                observations = self.logic.metricObservations(
                    self._visualizationRecords, segment, metricKey
                )
                if not observations:
                    skippedMetrics.append(self.logic.metricDisplayName(metricKey))
                    continue

                statistics = self.logic.computeMetricStatistics(observations)
                self.createSlicerMetricPlot(
                    observations=observations,
                    statistics=statistics,
                    segment=segment,
                    metricKey=metricKey,
                    plotType=plotType,
                )
                outputPath = self.suggestVisualizationFigurePath(
                    extension=extension,
                    segment=segment,
                    metricKey=metricKey,
                    plotType=plotType,
                    directory=outputDirectory,
                )
                savedPaths.append(
                    self.saveCurrentVisualizationFigure(outputPath)
                )
                lastRenderedMetricKey = metricKey

            if not savedPaths:
                raise RuntimeError(
                    "No figures were saved. None of the selected metrics had "
                    "valid values for the chosen segment."
                )

            if lastRenderedMetricKey and lastRenderedMetricKey in self._visualizationMetricKeys:
                self.visualizationMetricComboBox.setCurrentIndex(
                    self._visualizationMetricKeys.index(lastRenderedMetricKey)
                )
                self.updateVisualizationSuggestedFigurePath()

            message = (
                f"Saved {len(savedPaths)} figure(s) to: {outputDirectory}"
            )
            if skippedMetrics:
                message += (
                    " | Skipped: " + ", ".join(skippedMetrics)
                )
            self.visualizationStatusLabel.text = message
            slicer.util.infoDisplay(
                "Saved figure files\n" + "\n".join(savedPaths)
            )
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.visualizationStatusLabel.text = (
                f"Batch figure export error: {e}"
            )
            slicer.util.errorDisplay(str(e))

    def onVisualizationSaveFigureClicked(self, checked=False):
        try:
            outputPath = self.visualizationFigurePathEdit.currentPath
            if not outputPath:
                outputPath = self.suggestVisualizationFigurePath()

            outputPath = str(outputPath)
            root, extension = os.path.splitext(outputPath)
            extension = extension.lower()
            if not extension:
                outputPath += ".png"
                extension = ".png"
            if extension not in (".png", ".svg"):
                raise ValueError(
                    "Figure output must use a .png or .svg filename extension."
                )

            savedPath = self.saveCurrentVisualizationFigure(outputPath)
            try:
                self.visualizationFigurePathEdit.currentPath = savedPath
            except Exception:
                self.visualizationFigurePathEdit.setCurrentPath(savedPath)

            self.visualizationStatusLabel.text = (
                f"Figure saved to: {savedPath}"
            )
            slicer.util.infoDisplay(f"Figure saved to:\n{savedPath}")
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.visualizationStatusLabel.text = f"Figure save error: {e}"
            slicer.util.errorDisplay(str(e))

    def _safeFigureFilenamePart(self, value):
        value = str(value or "").strip()
        value = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
        value = value.strip("._-")
        return value or "figure"

    def defaultVisualizationFigureDirectory(self):
        """Return the user's Downloads directory with a cross-platform fallback."""
        outputDirectory = ""
        try:
            outputDirectory = qt.QStandardPaths.writableLocation(
                qt.QStandardPaths.DownloadLocation
            )
        except Exception:
            outputDirectory = ""

        if not outputDirectory:
            outputDirectory = os.path.join(os.path.expanduser("~"), "Downloads")
        return os.path.abspath(str(outputDirectory))

    def suggestVisualizationFigurePath(self, extension=".png", segment=None,
                                      metricKey=None, plotType=None,
                                      directory=None):
        if not extension.startswith("."):
            extension = "." + extension

        segment = segment or self.visualizationSegmentComboBox.currentText or "segment"
        metricKey = metricKey or self.currentVisualizationMetricKey() or "metric"
        plotType = plotType or self.visualizationPlotTypeComboBox.currentText or "figure"
        directory = directory or self.defaultVisualizationFigureDirectory()

        filename = "_".join((
            self._safeFigureFilenamePart(segment),
            self._safeFigureFilenamePart(metricKey),
            self._safeFigureFilenamePart(plotType),
        )) + extension
        return os.path.join(os.path.abspath(str(directory)), filename)

    def resolveVisualizationOutputSettings(self, currentPath=""):
        currentPath = str(currentPath or "").strip()
        if currentPath:
            outputDirectory = os.path.dirname(currentPath) or self.defaultVisualizationFigureDirectory()
            extension = os.path.splitext(currentPath)[1].lower() or ".png"
        else:
            outputDirectory = self.defaultVisualizationFigureDirectory()
            extension = ".png"

        if extension not in (".png", ".svg"):
            raise ValueError(
                "Figure output must use a .png or .svg filename extension."
            )
        return os.path.abspath(outputDirectory), extension

    def selectedVisualizationBatchMetricKeys(self):
        selected = []
        for item in self.visualizationBatchMetricList.selectedItems():
            data = item.data(qt.Qt.UserRole)
            if data:
                selected.append(str(data))
        return selected

    def updateVisualizationSuggestedFigurePath(self):
        suggestedPath = self.suggestVisualizationFigurePath(extension=".png")
        try:
            self.visualizationFigurePathEdit.currentPath = suggestedPath
        except Exception:
            self.visualizationFigurePathEdit.setCurrentPath(suggestedPath)

    def saveCurrentVisualizationFigure(self, outputPath):
        chartNode = None
        if self._visualizationCurrentChartNodeId:
            chartNode = slicer.mrmlScene.GetNodeByID(
                self._visualizationCurrentChartNodeId
            )
        if chartNode is None:
            raise RuntimeError(
                "No figure is currently available. Click 'Show Figure' first."
            )

        outputPath = os.path.abspath(str(outputPath))
        outputDirectory = os.path.dirname(outputPath)
        if outputDirectory:
            os.makedirs(outputDirectory, exist_ok=True)

        # Display the selected chart before exporting its plot view.
        slicer.modules.plots.logic().ShowChartInLayout(chartNode)
        slicer.app.processEvents()

        layoutManager = slicer.app.layoutManager()
        plotWidget = layoutManager.plotWidget(0)
        if plotWidget is None:
            raise RuntimeError("The Slicer plot view is not available.")
        plotView = plotWidget.plotView()
        if plotView is None:
            raise RuntimeError("The current plot view could not be accessed.")

        slicer.app.processEvents()
        extension = os.path.splitext(outputPath)[1].lower()
        if extension == ".svg":
            # qMRMLPlotView provides native vector export.
            plotView.saveAsSVG(outputPath)
        elif extension == ".png":
            try:
                import ScreenCapture
                ScreenCapture.ScreenCaptureLogic().captureImageFromView(
                    plotView, outputPath
                )
            except Exception:
                # Fallback when ScreenCapture cannot capture qMRMLPlotView.
                try:
                    pixmap = plotView.grab()
                except Exception:
                    pixmap = qt.QPixmap.grabWidget(plotView)
                if not pixmap.save(outputPath):
                    raise RuntimeError("Qt failed to save the plot image.")
        else:
            raise ValueError("Only PNG and SVG figure export are supported.")

        if not os.path.isfile(outputPath) or os.path.getsize(outputPath) == 0:
            raise RuntimeError(
                f"Slicer did not create a valid figure file: {outputPath}"
            )
        return outputPath

    def onVisualizationSegmentChanged(self, index=-1):
        self.populateVisualizationMetricComboBox()
        if self._visualizationRecords:
            self.updateVisualizationFigure(silent=True)

    def onVisualizationSelectionChanged(self, index=-1):
        if self._visualizationRecords:
            self.updateVisualizationFigure(silent=True)

    def showVisualizationFromExcel(self, excelPath, preferredSegment=None,
                                   preferredMetric=None):
        if not excelPath or not os.path.isfile(excelPath):
            return

        try:
            self.visualizationExcelEdit.currentPath = excelPath
        except Exception:
            self.visualizationExcelEdit.setCurrentPath(excelPath)

        self.loadVisualizationData(
            excelPath,
            preferredSegment=preferredSegment,
            preferredMetric=preferredMetric,
        )
        self.mainTabs.setCurrentIndex(self.visualizationTabIndex)
        self.updateVisualizationFigure()

    def loadVisualizationData(self, excelPath, preferredSegment=None,
                              preferredMetric=None):
        records = self.logic.loadMetricRowsFromExcel(excelPath)
        self._visualizationRecords = records

        previousSegment = preferredSegment or self.visualizationSegmentComboBox.currentText
        self.visualizationSegmentComboBox.blockSignals(True)
        self.visualizationSegmentComboBox.clear()

        segments = self.logic.visualizationSegments(records)
        for segment in segments:
            self.visualizationSegmentComboBox.addItem(segment)

        if previousSegment in segments:
            self.visualizationSegmentComboBox.setCurrentIndex(
                segments.index(previousSegment)
            )
        elif segments:
            self.visualizationSegmentComboBox.setCurrentIndex(0)

        self.visualizationSegmentComboBox.blockSignals(False)
        self.populateVisualizationMetricComboBox(preferredMetric=preferredMetric)

        self.visualizationStatusLabel.text = (
            f"Loaded {len(records)} completed row(s) from: {excelPath}"
        )

    def populateVisualizationMetricComboBox(self, preferredMetric=None):
        segment = self.visualizationSegmentComboBox.currentText
        previousMetric = preferredMetric or self.currentVisualizationMetricKey()
        metricKeys = self.logic.visualizationMetricKeys(
            self._visualizationRecords, segment
        )
        self._visualizationMetricKeys = metricKeys

        self.visualizationMetricComboBox.blockSignals(True)
        self.visualizationMetricComboBox.clear()
        for metricKey in metricKeys:
            self.visualizationMetricComboBox.addItem(
                self.logic.metricDisplayName(metricKey)
            )

        if previousMetric in metricKeys:
            self.visualizationMetricComboBox.setCurrentIndex(
                metricKeys.index(previousMetric)
            )
        elif "suv_max" in metricKeys:
            self.visualizationMetricComboBox.setCurrentIndex(
                metricKeys.index("suv_max")
            )
        elif metricKeys:
            self.visualizationMetricComboBox.setCurrentIndex(0)
        self.visualizationMetricComboBox.blockSignals(False)

        self.populateVisualizationBatchMetricList(preferredMetric=previousMetric)

    def populateVisualizationBatchMetricList(self, preferredMetric=None):
        self.visualizationBatchMetricList.clear()
        for metricKey in self._visualizationMetricKeys:
            item = qt.QListWidgetItem(self.logic.metricDisplayName(metricKey))
            item.setData(qt.Qt.UserRole, metricKey)
            self.visualizationBatchMetricList.addItem(item)
            if preferredMetric and str(metricKey) == str(preferredMetric):
                item.setSelected(True)

    def currentVisualizationMetricKey(self):
        index = self.visualizationMetricComboBox.currentIndex
        if 0 <= index < len(self._visualizationMetricKeys):
            return self._visualizationMetricKeys[index]
        return ""

    def updateVisualizationFigure(self, silent=False):
        segment = self.visualizationSegmentComboBox.currentText
        metricKey = self.currentVisualizationMetricKey()
        if not segment or not metricKey:
            if not silent:
                self.visualizationStatusLabel.text = (
                    "No segment or numeric metric is available for plotting."
                )
            return

        observations = self.logic.metricObservations(
            self._visualizationRecords, segment, metricKey
        )
        if not observations:
            if not silent:
                self.visualizationStatusLabel.text = (
                    f"No valid values for {segment} / {metricKey}."
                )
            return

        statistics = self.logic.computeMetricStatistics(observations)
        self.updateVisualizationStatistics(statistics)

        plotType = self.visualizationPlotTypeComboBox.currentText
        self.updateVisualizationPatientTable(
            observations=observations,
            statistics=statistics,
            plotType=plotType,
        )
        self.createSlicerMetricPlot(
            observations=observations,
            statistics=statistics,
            segment=segment,
            metricKey=metricKey,
            plotType=plotType,
        )
        self.visualizationStatusLabel.text = (
            f"Showing {plotType.lower()} for {segment} / "
            f"{self.logic.metricDisplayName(metricKey)} "
            f"({statistics['n']} valid subjects)."
        )

    def updateVisualizationStatistics(self, statistics):
        def fmt(value):
            if value is None or not np.isfinite(value):
                return "—"
            return f"{value:.6g}"

        self.visualizationStatisticLabels["N"].text = str(statistics["n"])
        self.visualizationStatisticLabels["Mean"].text = fmt(statistics["mean"])
        self.visualizationStatisticLabels["Standard deviation"].text = fmt(
            statistics["std"]
        )
        self.visualizationStatisticLabels["Median"].text = fmt(
            statistics["median"]
        )
        self.visualizationStatisticLabels["Interquartile range"].text = fmt(
            statistics["iqr"]
        )
        self.visualizationStatisticLabels["Tukey outliers"].text = (
            f"{statistics['outlier_count']} "
            f"({fmt(statistics['lower_fence'])} to "
            f"{fmt(statistics['upper_fence'])})"
        )

        top3 = statistics["top3"]
        self.visualizationTop3Table.setRowCount(len(top3))
        for row, observation in enumerate(top3):
            values = (
                row + 1,
                observation.get("subject_id", ""),
                observation.get("scan_date", ""),
                f"{observation['value']:.6g}",
            )
            for column, value in enumerate(values):
                item = qt.QTableWidgetItem(str(value))
                item.setBackground(qt.QColor(255, 224, 224))
                self.visualizationTop3Table.setItem(row, column, item)
        self.visualizationTop3Table.resizeColumnsToContents()
        self.visualizationTop3Table.horizontalHeader().setStretchLastSection(True)

    def _patientObservationSortKey(self, observation):
        subject = str(observation.get("subject_id", "") or "").strip()
        patient = str(observation.get("patient_id", "") or "").strip()
        scanDate = str(observation.get("scan_date", "") or "").strip()
        return (subject.lower(), patient.lower(), scanDate.lower())

    def updateVisualizationPatientTable(self, observations, statistics, plotType):
        showPatientTable = plotType == "Values by patient"
        self.visualizationPatientTableLabel.setVisible(showPatientTable)
        self.visualizationPatientTable.setVisible(showPatientTable)

        if not showPatientTable:
            self.visualizationPatientTable.setRowCount(0)
            return

        ordered = sorted(observations, key=self._patientObservationSortKey)
        top3Ids = {id(observation) for observation in statistics["top3"]}

        self.visualizationPatientTable.setRowCount(len(ordered))
        for row, observation in enumerate(ordered):
            isTop3 = id(observation) in top3Ids
            values = (
                row + 1,
                observation.get("subject_id", ""),
                observation.get("patient_id", ""),
                observation.get("scan_date", ""),
                f"{observation['value']:.6g}",
                "Yes" if isTop3 else "",
            )
            for column, value in enumerate(values):
                item = qt.QTableWidgetItem(str(value))
                if isTop3:
                    item.setBackground(qt.QColor(255, 224, 224))
                self.visualizationPatientTable.setItem(row, column, item)

        self.visualizationPatientTable.resizeColumnsToContents()
        self.visualizationPatientTable.horizontalHeader().setStretchLastSection(True)

    def _removeVisualizationPlotNodes(self):
        for nodeId in reversed(self._visualizationPlotNodeIds):
            node = slicer.mrmlScene.GetNodeByID(nodeId)
            if node is not None:
                slicer.mrmlScene.RemoveNode(node)
        self._visualizationPlotNodeIds = []
        self._visualizationCurrentChartNodeId = None

    def _addVisualizationSeries(self, name, xValues, yValues, plotType,
                                color, markerStyle=None, markerSize=None,
                                lineStyle=None, lineWidth=None):
        data = np.column_stack((
            np.asarray(xValues, dtype=float),
            np.asarray(yValues, dtype=float),
        ))
        tableNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLTableNode", name + " table"
        )
        slicer.util.updateTableFromArray(tableNode, data, ["X", "Y"])

        seriesNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLPlotSeriesNode", name
        )
        seriesNode.SetAndObserveTableNodeID(tableNode.GetID())
        seriesNode.SetXColumnName("X")
        seriesNode.SetYColumnName("Y")
        seriesNode.SetPlotType(plotType)
        seriesNode.SetColor(*color)

        if markerStyle is not None:
            seriesNode.SetMarkerStyle(markerStyle)
        if markerSize is not None:
            seriesNode.SetMarkerSize(float(markerSize))
        if lineStyle is not None:
            seriesNode.SetLineStyle(lineStyle)
        if lineWidth is not None:
            seriesNode.SetLineWidth(float(lineWidth))

        self._visualizationPlotNodeIds.extend([
            tableNode.GetID(), seriesNode.GetID()
        ])
        return seriesNode

    def _shortVisualizationMetricName(self, metricKey):
        fullName = self.logic.metricDisplayName(metricKey)
        # UI descriptions may include an expanded definition in parentheses.
        # The plot only needs the concise metric name to preserve figure space.
        return str(fullName).split(" (", 1)[0].strip()

    def _shortVisualizationSegmentName(self, segment):
        name = str(segment or "Segment").strip()
        lowerName = name.lower()
        suffixes = (
            "_3ds_mirrored_subtracted_processed",
            "_mirrored_subtracted_processed",
            "_subtracted_processed",
            "_processed",
        )
        for suffix in suffixes:
            if lowerName.endswith(suffix):
                name = name[:-len(suffix)]
                break
        name = re.sub(r"[_\-]+", " ", name)
        name = re.sub(r"\s+", " ", name).strip()
        return name.title() if name else "Segment"

    def createSlicerMetricPlot(self, observations, statistics, segment,
                               metricKey, plotType):
        self._removeVisualizationPlotNodes()

        metricName = self.logic.metricDisplayName(metricKey)
        shortMetricName = self._shortVisualizationMetricName(metricKey)
        shortSegmentName = self._shortVisualizationSegmentName(segment)
        values = np.asarray(
            [observation["value"] for observation in observations], dtype=float
        )
        top3 = statistics["top3"]

        xAxisRange = None

        if plotType == "Values by patient":
            orderedObservations = sorted(
                observations, key=self._patientObservationSortKey
            )
            patientIndices = np.arange(
                1, len(orderedObservations) + 1, dtype=float
            )
            patientValues = np.asarray(
                [item["value"] for item in orderedObservations], dtype=float
            )
            top3Ids = {id(observation) for observation in top3}

            # A single bar series keeps every bar centred on its integer patient
            # index. Top-three values are overlaid as markers instead of a second
            # bar series, because multiple VTK bar series are laid out side by
            # side and otherwise make selected bars look horizontally shifted.
            mainSeries = self._addVisualizationSeries(
                "Patients",
                patientIndices,
                patientValues,
                slicer.vtkMRMLPlotSeriesNode.PlotTypeScatterBar,
                (0.18, 0.48, 0.78),
                markerStyle=slicer.vtkMRMLPlotSeriesNode.MarkerStyleNone,
            )

            topIndices = []
            topValues = []
            for patientIndex, observation in zip(
                    patientIndices, orderedObservations):
                if id(observation) in top3Ids:
                    topIndices.append(patientIndex)
                    topValues.append(observation["value"])

            topSeries = self._addVisualizationSeries(
                "Top 3 patients",
                topIndices,
                topValues,
                slicer.vtkMRMLPlotSeriesNode.PlotTypeScatter,
                (0.90, 0.12, 0.12),
                markerStyle=slicer.vtkMRMLPlotSeriesNode.MarkerStyleDiamond,
                markerSize=14,
                lineStyle=slicer.vtkMRMLPlotSeriesNode.LineStyleNone,
            )
            xTitle = "Patient index (see table)"
            yTitle = self._shortVisualizationMetricName(metricKey)
            xAxisRange = (0.5, len(orderedObservations) + 0.5)

        elif plotType == "Ranked values":
            sortedObservations = sorted(
                observations, key=lambda item: item["value"], reverse=True
            )
            ranks = np.arange(1, len(sortedObservations) + 1, dtype=float)
            sortedValues = np.asarray(
                [item["value"] for item in sortedObservations], dtype=float
            )

            mainSeries = self._addVisualizationSeries(
                "All subjects",
                ranks,
                sortedValues,
                slicer.vtkMRMLPlotSeriesNode.PlotTypeLine,
                (0.18, 0.48, 0.78),
                markerStyle=slicer.vtkMRMLPlotSeriesNode.MarkerStyleCircle,
                markerSize=6,
                lineStyle=slicer.vtkMRMLPlotSeriesNode.LineStyleSolid,
                lineWidth=2,
            )
            topRanks = np.arange(1, len(top3) + 1, dtype=float)
            topValues = np.asarray(
                [item["value"] for item in top3], dtype=float
            )
            topSeries = self._addVisualizationSeries(
                "Top 3",
                topRanks,
                topValues,
                slicer.vtkMRMLPlotSeriesNode.PlotTypeScatter,
                (0.90, 0.12, 0.12),
                markerStyle=slicer.vtkMRMLPlotSeriesNode.MarkerStyleCircle,
                markerSize=13,
                lineStyle=slicer.vtkMRMLPlotSeriesNode.LineStyleNone,
            )
            xTitle = "Rank (largest to smallest)"
            yTitle = shortMetricName
        else:
            binCount = int(min(20, max(5, math.ceil(math.sqrt(len(values))))))
            counts, edges = np.histogram(values, bins=binCount)
            centers = (edges[:-1] + edges[1:]) / 2.0

            mainSeries = self._addVisualizationSeries(
                "Distribution",
                centers,
                counts.astype(float),
                slicer.vtkMRMLPlotSeriesNode.PlotTypeScatterBar,
                (0.18, 0.48, 0.78),
                markerStyle=slicer.vtkMRMLPlotSeriesNode.MarkerStyleNone,
            )

            topX = np.asarray([item["value"] for item in top3], dtype=float)
            topY = []
            markerOffset = max(0.25, float(np.max(counts)) * 0.06)
            for value in topX:
                binIndex = int(np.searchsorted(edges, value, side="right") - 1)
                binIndex = min(max(binIndex, 0), len(counts) - 1)
                topY.append(float(counts[binIndex]) + markerOffset)

            topSeries = self._addVisualizationSeries(
                "Top 3",
                topX,
                np.asarray(topY, dtype=float),
                slicer.vtkMRMLPlotSeriesNode.PlotTypeScatter,
                (0.90, 0.12, 0.12),
                markerStyle=slicer.vtkMRMLPlotSeriesNode.MarkerStyleDiamond,
                markerSize=13,
                lineStyle=slicer.vtkMRMLPlotSeriesNode.LineStyleNone,
            )
            xTitle = shortMetricName
            yTitle = "Subject count"

        chartNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLPlotChartNode", "PET Biomarker cohort figure"
        )
        chartNode.AddAndObservePlotSeriesNodeID(mainSeries.GetID())
        chartNode.AddAndObservePlotSeriesNodeID(topSeries.GetID())
        if plotType == "Values by patient":
            chartTitle = f"{shortMetricName} by patient — {shortSegmentName}"
        elif plotType == "Ranked values":
            chartTitle = f"{shortMetricName} ranked values — {shortSegmentName}"
        else:
            chartTitle = f"{shortMetricName} distribution — {shortSegmentName}"

        chartNode.SetTitle(chartTitle)
        chartNode.SetXAxisTitle(xTitle)
        chartNode.SetYAxisTitle(yTitle)
        chartNode.SetTitleVisibility(True)
        chartNode.SetXAxisTitleVisibility(True)
        chartNode.SetYAxisTitleVisibility(True)
        chartNode.SetLegendVisibility(True)
        chartNode.SetGridVisibility(True)
        chartNode.SetTitleFontSize(16)
        chartNode.SetAxisTitleFontSize(14)
        chartNode.SetAxisLabelFontSize(11)
        chartNode.SetLegendFontSize(12)
        if xAxisRange is not None:
            chartNode.SetXAxisRangeAuto(False)
            chartNode.SetXAxisRange(float(xAxisRange[0]), float(xAxisRange[1]))
        self._visualizationPlotNodeIds.append(chartNode.GetID())
        self._visualizationCurrentChartNodeId = chartNode.GetID()

        slicer.modules.plots.logic().ShowChartInLayout(chartNode)
        self.updateVisualizationSuggestedFigurePath()

    def cleanup(self):
        self._removeVisualizationPlotNodes()

    # ==================================================================
    # Batch tab (cohort quantification)
    #
    # Expected dataset layout (matches dataset_clean):
    #   <root>/PET/<base>_PET/            DICOM series folder per subject
    #   <root>/CT/<base>_CT/              (optional)
    #   <root>/Segments/<base>_Seg/       organ masks (*.nii.gz / *.seg.nrrd),
    #                                     incl. *_processed.nii.gz from ablation
    # where <base> is e.g. "MSP0001_2025-07-09".
    # ==================================================================
    def setupBatchTab(self, batchLayout):
        intro = qt.QLabel(
            "Two stages, materialised to disk so you don't recompute every time:\n"
            "  Stage 1 (Prepare) generates ablated masks once -> *_processed.nii.gz.\n"
            "  Stage 2 (Quantify) reads masks (raw or *_processed) -> metrics Excel.\n"
            "All stages share the Interactive tab's logic, so results never diverge."
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color:#555; font-style:italic;")
        batchLayout.addWidget(intro)

        # ----- Shared dataset section -----
        dataCol = ctk.ctkCollapsibleButton()
        dataCol.text = "Dataset (shared by both stages)"
        batchLayout.addWidget(dataCol)
        dataForm = qt.QFormLayout(dataCol)

        self.batchRootEdit = ctk.ctkPathLineEdit()
        self.batchRootEdit.filters = ctk.ctkPathLineEdit.Dirs
        self.batchRootEdit.setToolTip(
            "Root folder containing PET/, CT/ and Segments/ subfolders."
        )
        dataForm.addRow("Root folder:", self.batchRootEdit)

        self.batchDetectButton = qt.QPushButton("Detect subjects + segment files")
        dataForm.addRow("", self.batchDetectButton)

        self.batchScansLabel = qt.QLabel("No folder scanned.")
        self.batchScansLabel.setWordWrap(True)
        dataForm.addRow("Found:", self.batchScansLabel)

        # ----- Stage sub-tabs -----
        self.batchSubTabs = qt.QTabWidget()
        batchLayout.addWidget(self.batchSubTabs)

        prepareTab = qt.QWidget()
        prepareLayout = qt.QVBoxLayout(prepareTab)
        self.batchSubTabs.addTab(prepareTab, "1. Prepare (masks + ablation)")
        self._setupPrepareSubTab(prepareLayout)

        quantifyTab = qt.QWidget()
        quantifyLayout = qt.QVBoxLayout(quantifyTab)
        self.batchSubTabs.addTab(quantifyTab, "2. Quantify (metrics)")
        self._setupQuantifySubTab(quantifyLayout)

        batchLayout.addStretch(1)

        # shared state + detect connection
        self._batchSegRows = []      # quantify table rows: (includeCheck, stem, labelEdit)
        self._prepOrganRows = []     # prepare table rows:  (includeCheck, stem)
        self._batchCancel = False
        self._prepCancel = False
        self.batchDetectButton.connect("clicked(bool)", self.onBatchDetectClicked)

    # ---------- Prepare sub-tab ----------
    def _setupPrepareSubTab(self, layout):
        help1 = qt.QLabel(
            "Apply ablation to organ masks once and write <stem>_processed.nii.gz "
            "into each subject's _Seg folder. Auxiliary masks (vertebrae, kidney) "
            "are read from the _Seg folder, or generated from CT via "
            "TotalSegmentator when missing (if enabled)."
        )
        help1.setWordWrap(True)
        help1.setStyleSheet("color:#555; font-style:italic;")
        layout.addWidget(help1)

        # Ablation steps
        stepsCol = ctk.ctkCollapsibleButton()
        stepsCol.text = "Ablation steps"
        layout.addWidget(stepsCol)
        stepsForm = qt.QFormLayout(stepsCol)

        self.prepSpineClipCheck = qt.QCheckBox("Spine L1-L5 clip")
        self.prepKidneyExcludeCheck = qt.QCheckBox("Kidney exclusion")
        self.prepUrinaryCleanCheck = qt.QCheckBox("Urinary (ureter) cleanup")
        self.prepSpineClipCheck.setChecked(True)
        self.prepUrinaryCleanCheck.setChecked(True)
        stepsW = qt.QWidget()
        stepsB = qt.QHBoxLayout(stepsW)
        stepsB.setContentsMargins(0, 0, 0, 0)
        for cb in (self.prepSpineClipCheck, self.prepKidneyExcludeCheck,
                   self.prepUrinaryCleanCheck):
            stepsB.addWidget(cb)
        stepsForm.addRow("Steps:", stepsW)

        self.prepKidneyDilationSpin = qt.QDoubleSpinBox()
        self.prepKidneyDilationSpin.setRange(0.0, 50.0)
        self.prepKidneyDilationSpin.setValue(10.0)
        self.prepKidneyDilationSpin.setSuffix(" mm")
        stepsForm.addRow("Kidney dilation:", self.prepKidneyDilationSpin)

        self.prepUreterSuvThreshSpin = qt.QDoubleSpinBox()
        self.prepUreterSuvThreshSpin.setRange(0.0, 100.0)
        self.prepUreterSuvThreshSpin.setSingleStep(0.5)
        self.prepUreterSuvThreshSpin.setValue(4.0)
        stepsForm.addRow("Urinary SUV threshold:", self.prepUreterSuvThreshSpin)

        self.prepUreterDilationSpin = qt.QDoubleSpinBox()
        self.prepUreterDilationSpin.setRange(0.0, 50.0)
        self.prepUreterDilationSpin.setValue(5.0)
        self.prepUreterDilationSpin.setSuffix(" mm")
        stepsForm.addRow("Urinary mask dilation:", self.prepUreterDilationSpin)

        self.prepCleanSuvThreshSpin = qt.QDoubleSpinBox()
        self.prepCleanSuvThreshSpin.setRange(0.0, 100.0)
        self.prepCleanSuvThreshSpin.setSingleStep(0.5)
        self.prepCleanSuvThreshSpin.setValue(2.0)
        stepsForm.addRow("ROI cleanup SUV threshold:", self.prepCleanSuvThreshSpin)

        # Auxiliary mask sources
        auxCol = ctk.ctkCollapsibleButton()
        auxCol.text = "Auxiliary masks (vertebrae / kidney)"
        layout.addWidget(auxCol)
        auxForm = qt.QFormLayout(auxCol)

        self.prepVertebraeFileEdit = qt.QLineEdit("")
        self.prepVertebraeFileEdit.setToolTip(
            "Filename in each _Seg folder holding L1-L5 vertebrae (e.g. a TotalSeg "
            ".seg.nrrd, or vertebrae.nii.gz). Leave blank to rely on generation."
        )
        auxForm.addRow("Vertebrae file:", self.prepVertebraeFileEdit)

        self.prepKidneyFileEdit = qt.QLineEdit("")
        self.prepKidneyFileEdit.setToolTip(
            "Filename in each _Seg folder holding kidneys. Leave blank to rely on "
            "generation."
        )
        auxForm.addRow("Kidney file:", self.prepKidneyFileEdit)

        self.prepGenerateIfMissingCheck = qt.QCheckBox(
            "Generate from CT via TotalSegmentator when the file is missing (slow)"
        )
        auxForm.addRow("", self.prepGenerateIfMissingCheck)

        # Organ files to process
        organCol = ctk.ctkCollapsibleButton()
        organCol.text = "Organ masks to process"
        layout.addWidget(organCol)
        organLayout = qt.QVBoxLayout(organCol)
        organLayout.addWidget(qt.QLabel("Tick the organ masks to ablate:"))

        self.prepOrganTable = qt.QTableWidget(0, 2)
        self.prepOrganTable.setHorizontalHeaderLabels(["Include", "File (stem)"])
        self.prepOrganTable.horizontalHeader().setStretchLastSection(True)
        self.prepOrganTable.setFixedHeight(150)
        organLayout.addWidget(self.prepOrganTable)

        suffixRow = qt.QFormLayout()
        self.prepSuffixEdit = qt.QLineEdit("_processed")
        suffixRow.addRow("Output suffix:", self.prepSuffixEdit)
        self.prepSkipDoneCheck = qt.QCheckBox(
            "Skip organs whose _processed file already exists"
        )
        self.prepSkipDoneCheck.setChecked(True)
        suffixRow.addRow("", self.prepSkipDoneCheck)
        organLayout.addLayout(suffixRow)

        # Run controls
        runRow = qt.QHBoxLayout()
        self.prepRunButton = qt.QPushButton("Run prepare")
        self.prepRunButton.setStyleSheet(
            "QPushButton{background:#1565c0;color:white;font-weight:bold;"
            "padding:8px;border-radius:4px;}"
            "QPushButton:hover{background:#0d47a1;}"
            "QPushButton:disabled{background:#888;}"
        )
        self.prepCancelButton = qt.QPushButton("Cancel")
        self.prepCancelButton.setEnabled(False)
        runRow.addWidget(self.prepRunButton)
        runRow.addWidget(self.prepCancelButton)
        layout.addLayout(runRow)

        self.prepProgressBar = qt.QProgressBar()
        self.prepProgressBar.setValue(0)
        layout.addWidget(self.prepProgressBar)

        self.prepStatusLabel = qt.QLabel("Ready.")
        self.prepStatusLabel.setWordWrap(True)
        layout.addWidget(self.prepStatusLabel)

        layout.addStretch(1)

        self.prepRunButton.connect("clicked(bool)", self.onPrepareRunClicked)
        self.prepCancelButton.connect("clicked(bool)", self.onPrepareCancelClicked)

    # ---------- Quantify sub-tab ----------
    def _setupQuantifySubTab(self, layout):
        segCol = ctk.ctkCollapsibleButton()
        segCol.text = "Segment files to quantify"
        layout.addWidget(segCol)
        segLayout = qt.QVBoxLayout(segCol)

        segHelp = qt.QLabel(
            "Tick the masks to analyse (raw or *_processed). Edit 'Excel label' to "
            "set the column name (e.g. VF). Auxiliary files are unticked by default."
        )
        segHelp.setWordWrap(True)
        segLayout.addWidget(segHelp)

        self.batchSegTable = qt.QTableWidget(0, 3)
        self.batchSegTable.setHorizontalHeaderLabels(
            ["Include", "File (stem)", "Excel label"]
        )
        self.batchSegTable.horizontalHeader().setStretchLastSection(True)
        self.batchSegTable.setFixedHeight(180)
        segLayout.addWidget(self.batchSegTable)

        compCol = ctk.ctkCollapsibleButton()
        compCol.text = "Computation"
        layout.addWidget(compCol)
        compForm = qt.QFormLayout(compCol)

        metricsW = qt.QWidget()
        metricsB = qt.QHBoxLayout(metricsW)
        metricsB.setContentsMargins(0, 0, 0, 0)
        self.batchSuvMeanCheckBox = qt.QCheckBox("SUVmean")
        self.batchSuvMaxCheckBox = qt.QCheckBox("SUVmax")
        self.batchSuvPeakCheckBox = qt.QCheckBox("SUVpeak")
        self.batchTlgCheckBox = qt.QCheckBox("TLG")
        self.batchVolumeCheckBox = qt.QCheckBox("Volume")
        for cb in (self.batchSuvMeanCheckBox, self.batchSuvMaxCheckBox,
                   self.batchSuvPeakCheckBox, self.batchTlgCheckBox,
                   self.batchVolumeCheckBox):
            cb.setChecked(True)
            metricsB.addWidget(cb)
        compForm.addRow("Metrics:", metricsW)

        selectedRadW = qt.QWidget()
        selectedRadGrid = qt.QGridLayout(selectedRadW)
        selectedRadGrid.setContentsMargins(0, 0, 0, 0)

        self.batchRadP10CheckBox = qt.QCheckBox("P10")
        self.batchRadP90CheckBox = qt.QCheckBox("P90")
        self.batchRadEntropyCheckBox = qt.QCheckBox("Entropy")
        self.batchRadSkewnessCheckBox = qt.QCheckBox("Skewness")
        self.batchRadContrastCheckBox = qt.QCheckBox("GLCM Contrast")
        self.batchRadSahgleCheckBox = qt.QCheckBox("SAHGLE")
        self.batchRadLalgleCheckBox = qt.QCheckBox("LALGLE")
        self.batchRadZoneEntropyCheckBox = qt.QCheckBox("ZoneEntropy")

        selectedBatchRadiomics = (
            (
                self.batchRadP10CheckBox,
                "10th Percentile: representative low-end Standardized Uptake Value (SUV) within the Region of Interest (ROI).",
            ),
            (
                self.batchRadP90CheckBox,
                "90th Percentile: representative high-end Standardized Uptake Value (SUV), less sensitive than SUVmax to one extreme voxel.",
            ),
            (
                self.batchRadEntropyCheckBox,
                "First-Order Entropy: uncertainty and diversity of Standardized Uptake Value (SUV) intensities within the Region of Interest (ROI).",
            ),
            (
                self.batchRadSkewnessCheckBox,
                "First-Order Skewness: asymmetry of the Standardized Uptake Value (SUV) distribution, including a tail toward high uptake.",
            ),
            (
                self.batchRadContrastCheckBox,
                "Gray Level Co-occurrence Matrix (GLCM) Contrast: magnitude of local intensity differences between neighboring voxels; higher values indicate greater local heterogeneity.",
            ),
            (
                self.batchRadSahgleCheckBox,
                "Gray Level Size Zone Matrix (GLSZM) Small Area High Gray Level Emphasis (SAHGLE): prevalence of small high-uptake zones.",
            ),
            (
                self.batchRadLalgleCheckBox,
                "Gray Level Size Zone Matrix (GLSZM) Large Area Low Gray Level Emphasis (LALGLE): prevalence of large low-uptake zones.",
            ),
            (
                self.batchRadZoneEntropyCheckBox,
                "Gray Level Size Zone Matrix (GLSZM) Zone Entropy: diversity of connected-zone sizes and intensity levels.",
            ),
        )
        for row, (checkBox, description) in enumerate(selectedBatchRadiomics):
            checkBox.setChecked(True)
            checkBox.setToolTip(description)

            descriptionLabel = qt.QLabel(description)
            descriptionLabel.setWordWrap(True)
            descriptionLabel.setStyleSheet("color:#666;")
            descriptionLabel.setToolTip(description)

            selectedRadGrid.addWidget(checkBox, row, 0)
            selectedRadGrid.addWidget(descriptionLabel, row, 1)

        selectedRadGrid.setColumnStretch(1, 1)

        compForm.addRow("Selected radiomics:", selectedRadW)

        self.batchRadBinWidthSpinBox = qt.QDoubleSpinBox()
        self.batchRadBinWidthSpinBox.setRange(0.001, 10.0)
        self.batchRadBinWidthSpinBox.setDecimals(3)
        self.batchRadBinWidthSpinBox.setSingleStep(0.05)
        self.batchRadBinWidthSpinBox.setValue(0.25)
        self.batchRadBinWidthSpinBox.setToolTip(
            "Use one fixed SUV bin width for the entire cohort."
        )
        compForm.addRow("Radiomics bin width:", self.batchRadBinWidthSpinBox)

        batchResampleW = qt.QWidget()
        batchResampleB = qt.QHBoxLayout(batchResampleW)
        batchResampleB.setContentsMargins(0, 0, 0, 0)
        self.batchRadResampleCheckBox = qt.QCheckBox("Isotropic resampling")
        self.batchRadResampleCheckBox.setChecked(False)
        self.batchRadResampleSpacingSpinBox = qt.QDoubleSpinBox()
        self.batchRadResampleSpacingSpinBox.setRange(0.5, 20.0)
        self.batchRadResampleSpacingSpinBox.setDecimals(2)
        self.batchRadResampleSpacingSpinBox.setSingleStep(0.5)
        self.batchRadResampleSpacingSpinBox.setValue(4.0)
        self.batchRadResampleSpacingSpinBox.setSuffix(" mm")
        self.batchRadResampleSpacingSpinBox.setEnabled(False)
        self.batchRadResampleCheckBox.toggled.connect(
            self.batchRadResampleSpacingSpinBox.setEnabled
        )
        batchResampleB.addWidget(self.batchRadResampleCheckBox)
        batchResampleB.addWidget(self.batchRadResampleSpacingSpinBox)
        compForm.addRow("Texture geometry:", batchResampleW)

        self.batchSkipDoneCheckBox = qt.QCheckBox(
            "Skip subjects already present in the output Excel"
        )
        self.batchSkipDoneCheckBox.setChecked(True)
        compForm.addRow("", self.batchSkipDoneCheckBox)

        exportCol = ctk.ctkCollapsibleButton()
        exportCol.text = "Export"
        layout.addWidget(exportCol)
        exportForm = qt.QFormLayout(exportCol)

        self.batchOutputEdit = ctk.ctkPathLineEdit()
        self.batchOutputEdit.filters = ctk.ctkPathLineEdit.Files
        self.batchOutputEdit.nameFilters = ["Excel files (*.xlsx)"]
        exportForm.addRow("Output Excel:", self.batchOutputEdit)

        self.batchAppendCheckBox = qt.QCheckBox("Append to existing file")
        self.batchAppendCheckBox.setChecked(True)
        exportForm.addRow("", self.batchAppendCheckBox)

        runRow = qt.QHBoxLayout()
        self.batchRunButton = qt.QPushButton("Run quantify")
        self.batchRunButton.setStyleSheet(
            "QPushButton{background:#2e7d32;color:white;font-weight:bold;"
            "padding:8px;border-radius:4px;}"
            "QPushButton:hover{background:#1b5e20;}"
            "QPushButton:disabled{background:#888;}"
        )
        self.batchCancelButton = qt.QPushButton("Cancel")
        self.batchCancelButton.setEnabled(False)
        runRow.addWidget(self.batchRunButton)
        runRow.addWidget(self.batchCancelButton)
        layout.addLayout(runRow)

        self.batchProgressBar = qt.QProgressBar()
        self.batchProgressBar.setValue(0)
        layout.addWidget(self.batchProgressBar)

        self.batchStatusLabel = qt.QLabel("Ready.")
        self.batchStatusLabel.setWordWrap(True)
        layout.addWidget(self.batchStatusLabel)

        layout.addStretch(1)

        self.batchRunButton.connect("clicked(bool)", self.onBatchRunClicked)
        self.batchCancelButton.connect("clicked(bool)", self.onBatchCancelClicked)

    def onBatchDetectClicked(self, checked=False):
        root = self.batchRootEdit.currentPath
        try:
            info = self.logic.scanBatchDataset(root)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.batchScansLabel.text = f"Scan failed: {e}"
            return

        self.batchScansLabel.text = (
            f"{info['subjectCount']} subject(s) · "
            f"{len(info['segmentStems'])} segment file(s) in first subject."
        )
        self._populateBatchSegTable(info["segmentStems"])
        self._populatePrepareOrganTable(info["segmentStems"])

    # ---------- Prepare table + callbacks ----------
    def _populatePrepareOrganTable(self, stems):
        self._prepOrganRows = []
        self.prepOrganTable.setRowCount(len(stems))
        for r, stem in enumerate(stems):
            includeCheck = qt.QCheckBox()
            # default-tick true organ masks, skip aux and already-processed files
            isProcessed = self.logic.defaultExcelLabel(stem).endswith(
                self.prepSuffixEdit.text.strip() or "_processed")
            includeCheck.setChecked(
                not self.logic.isAuxiliarySegmentStem(stem) and not isProcessed)
            cellW = qt.QWidget()
            cellL = qt.QHBoxLayout(cellW)
            cellL.setContentsMargins(0, 0, 0, 0)
            cellL.setAlignment(qt.Qt.AlignCenter)
            cellL.addWidget(includeCheck)
            self.prepOrganTable.setCellWidget(r, 0, cellW)

            stemItem = qt.QTableWidgetItem(stem)
            stemItem.setFlags(qt.Qt.ItemIsEnabled)
            self.prepOrganTable.setItem(r, 1, stemItem)

            self._prepOrganRows.append((includeCheck, stem))
        self.prepOrganTable.resizeColumnsToContents()
        self.prepOrganTable.horizontalHeader().setStretchLastSection(True)

    def _collectPrepareOrganStems(self):
        return [stem for includeCheck, stem in self._prepOrganRows
                if includeCheck.isChecked()]

    def onPrepareCancelClicked(self, checked=False):
        self._prepCancel = True
        self.prepStatusLabel.text = "Cancelling after the current subject…"

    def onPrepareRunClicked(self, checked=False):
        root = self.batchRootEdit.currentPath
        organStems = self._collectPrepareOrganStems()
        steps = {
            "spine": self.prepSpineClipCheck.isChecked(),
            "kidney": self.prepKidneyExcludeCheck.isChecked(),
            "urinary": self.prepUrinaryCleanCheck.isChecked(),
        }
        if not organStems:
            self.prepStatusLabel.text = "Select at least one organ mask."
            return
        if not any(steps.values()):
            self.prepStatusLabel.text = "Select at least one ablation step."
            return

        auxConfig = {
            "vertebraeFile": self.prepVertebraeFileEdit.text.strip(),
            "kidneyFile": self.prepKidneyFileEdit.text.strip(),
            "generateIfMissing": self.prepGenerateIfMissingCheck.isChecked(),
        }
        params = {
            "spineLevels": ["L1", "L2", "L3", "L4", "L5"],
            "kidneyDilationMm": self.prepKidneyDilationSpin.value,
            "ureterSuvThresh": self.prepUreterSuvThreshSpin.value,
            "ureterDilationMm": self.prepUreterDilationSpin.value,
            "cleanSuvThresh": self.prepCleanSuvThreshSpin.value,
        }
        suffix = self.prepSuffixEdit.text.strip() or "_processed"

        self._prepCancel = False
        self.prepRunButton.setEnabled(False)
        self.prepCancelButton.setEnabled(True)
        self.prepProgressBar.setValue(0)
        slicer.app.processEvents()

        def progressCb(done, total, subject):
            self.prepProgressBar.setMaximum(total)
            self.prepProgressBar.setValue(done)
            self.prepStatusLabel.text = f"[{done}/{total}] {subject}"
            slicer.app.processEvents()

        def shouldCancel():
            slicer.app.processEvents()
            return self._prepCancel

        try:
            summary = self.logic.runBatchPrepare(
                root=root,
                organStems=organStems,
                steps=steps,
                auxConfig=auxConfig,
                params=params,
                outputSuffix=suffix,
                skipDone=self.prepSkipDoneCheck.isChecked(),
                progressCb=progressCb,
                shouldCancel=shouldCancel,
            )
            msg = (
                f"Prepare done. Subjects: {summary['processed']}, "
                f"skipped: {summary['skipped']}, errors: {summary['errors']}. "
                f"Re-run Detect to see the new *{suffix} files in the Quantify tab."
            )
            self.prepStatusLabel.text = msg
            slicer.util.infoDisplay(msg)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.prepStatusLabel.text = f"ERROR — see Python console: {e}"
            slicer.util.errorDisplay(str(e))
        finally:
            self.prepRunButton.setEnabled(True)
            self.prepCancelButton.setEnabled(False)

    def _populateBatchSegTable(self, stems):
        self._batchSegRows = []
        self.batchSegTable.setRowCount(len(stems))
        for r, stem in enumerate(stems):
            includeCheck = qt.QCheckBox()
            includeCheck.setChecked(not self.logic.isAuxiliarySegmentStem(stem))
            cellW = qt.QWidget()
            cellL = qt.QHBoxLayout(cellW)
            cellL.setContentsMargins(0, 0, 0, 0)
            cellL.setAlignment(qt.Qt.AlignCenter)
            cellL.addWidget(includeCheck)
            self.batchSegTable.setCellWidget(r, 0, cellW)

            stemItem = qt.QTableWidgetItem(stem)
            stemItem.setFlags(qt.Qt.ItemIsEnabled)
            self.batchSegTable.setItem(r, 1, stemItem)

            labelEdit = qt.QLineEdit(self.logic.defaultExcelLabel(stem))
            self.batchSegTable.setCellWidget(r, 2, labelEdit)

            self._batchSegRows.append((includeCheck, stem, labelEdit))
        self.batchSegTable.resizeColumnsToContents()
        self.batchSegTable.horizontalHeader().setStretchLastSection(True)

    def _collectBatchSegmentSelections(self):
        selections = []
        for includeCheck, stem, labelEdit in self._batchSegRows:
            if includeCheck.isChecked():
                label = labelEdit.text.strip() or self.logic.defaultExcelLabel(stem)
                selections.append((stem, label))
        return selections

    def _collectBatchMetricOptions(self):
        return {
            "mean": self.batchSuvMeanCheckBox.isChecked(),
            "max": self.batchSuvMaxCheckBox.isChecked(),
            "peak": self.batchSuvPeakCheckBox.isChecked(),
            "tlg": self.batchTlgCheckBox.isChecked(),
            "volume": self.batchVolumeCheckBox.isChecked(),
        }

    def _collectBatchRadiomicsOptions(self):
        selectedFeatureKeys = []
        featureCheckBoxes = (
            ("p10", self.batchRadP10CheckBox),
            ("p90", self.batchRadP90CheckBox),
            ("entropy", self.batchRadEntropyCheckBox),
            ("skewness", self.batchRadSkewnessCheckBox),
            ("contrast", self.batchRadContrastCheckBox),
            ("sahgle", self.batchRadSahgleCheckBox),
            ("lalgle", self.batchRadLalgleCheckBox),
            ("zone_entropy", self.batchRadZoneEntropyCheckBox),
        )
        for featureKey, checkBox in featureCheckBoxes:
            if checkBox.isChecked():
                selectedFeatureKeys.append(featureKey)

        return {
            "selected_feature_keys": selectedFeatureKeys,
            "derived": False,
            "bin_width": float(self.batchRadBinWidthSpinBox.value),
            "resample_isotropic": self.batchRadResampleCheckBox.isChecked(),
            "resampled_spacing_mm": float(
                self.batchRadResampleSpacingSpinBox.value
            ),
            # Full-class extraction is intentionally disabled in the current UI.
            "firstorder": False,
            "shape": False,
            "glcm": False,
            "glrlm": False,
            "glszm": False,
            "gldm": False,
            "ngtdm": False,
        }

    def onBatchCancelClicked(self, checked=False):
        self._batchCancel = True
        self.batchStatusLabel.text = "Cancelling after the current subject…"

    def onBatchRunClicked(self, checked=False):
        root = self.batchRootEdit.currentPath
        segmentSelections = self._collectBatchSegmentSelections()
        metricsOptions = self._collectBatchMetricOptions()
        radiomicsOptions = self._collectBatchRadiomicsOptions()
        outputFile = self.batchOutputEdit.currentPath

        if not segmentSelections:
            self.batchStatusLabel.text = "Select at least one segment file (section 2)."
            return
        if not any(metricsOptions.values()):
            self.batchStatusLabel.text = "Select at least one metric (section 3)."
            return
        if not outputFile:
            self.batchStatusLabel.text = "Set an output Excel path (section 4)."
            return

        self._batchCancel = False
        self.batchRunButton.setEnabled(False)
        self.batchCancelButton.setEnabled(True)
        self.batchProgressBar.setValue(0)
        slicer.app.processEvents()

        def progressCb(done, total, subject):
            self.batchProgressBar.setMaximum(total)
            self.batchProgressBar.setValue(done)
            self.batchStatusLabel.text = f"[{done}/{total}] {subject}"
            slicer.app.processEvents()

        def shouldCancel():
            slicer.app.processEvents()
            return self._batchCancel

        try:
            summary = self.logic.runBatchQuantification(
                root=root,
                segmentSelections=segmentSelections,
                metricsOptions=metricsOptions,
                radiomicsOptions=radiomicsOptions,
                outputFile=outputFile,
                append=self.batchAppendCheckBox.isChecked(),
                skipDone=self.batchSkipDoneCheckBox.isChecked(),
                progressCb=progressCb,
                shouldCancel=shouldCancel,
            )
            msg = (
                f"Done. Subjects processed: {summary['processed']}, "
                f"skipped: {summary['skipped']}, errors: {summary['errors']}.\n"
                f"Saved to: {summary['savedPath']}"
            )
            self.batchStatusLabel.text = msg

            visualizationPath = summary.get("savedPath")
            if not visualizationPath and os.path.isfile(outputFile):
                visualizationPath = outputFile
            if (visualizationPath
                    and self.visualizationAutoShowCheckBox.isChecked()):
                preferredSegment = (
                    segmentSelections[0][1] if segmentSelections else None
                )
                self.showVisualizationFromExcel(
                    visualizationPath,
                    preferredSegment=preferredSegment,
                    preferredMetric="suv_max",
                )

            slicer.util.infoDisplay(msg)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.batchStatusLabel.text = f"ERROR — see Python console: {e}"
            slicer.util.errorDisplay(str(e))
        finally:
            self.batchRunButton.setEnabled(True)
            self.batchCancelButton.setEnabled(False)


# --------------------------------------------------
# Logic
# --------------------------------------------------
class PETBiomarkerStudioLogic(ScriptedLoadableModuleLogic):

    # Kept on the Logic class so the Widget can read the same keys as lib.
    SELECTED_RADIOMICS_FEATURES = _lib_radiomics.SELECTED_RADIOMICS_FEATURES
    SELECTED_RADIOMICS_FEATURE_ORDER = _lib_radiomics.SELECTED_RADIOMICS_FEATURE_ORDER
    RADIOMICS_CLASS_KEYS = _lib_radiomics.RADIOMICS_CLASS_KEYS

    def selectedRadiomicsFeatureKeys(self, radiomicsOptions):
        return _lib_radiomics.selected_radiomics_feature_keys(radiomicsOptions)

    def selectedRadiomicsFeaturesByClass(self, radiomicsOptions):
        return _lib_radiomics.selected_radiomics_features_by_class(radiomicsOptions)

    def isRadiomicsEnabled(self, radiomicsOptions):
        return _lib_radiomics.is_radiomics_enabled(radiomicsOptions)

    def radiomicsConfigSignature(self, radiomicsOptions):
        return _lib_radiomics.radiomics_config_signature(radiomicsOptions)

    def computationSignature(self, metricsOptions, radiomicsOptions):
        return _lib_computation_signature(metricsOptions, radiomicsOptions)

    def loadVolumeNodeFromFile(self, filePath):
        if not filePath:
            raise ValueError("No volume file selected")
        
        filePath = os.path.abspath(filePath)

        if not os.path.exists(filePath):
            raise ValueError(f"Volume file does not exist: {filePath}")

        # Load Volume
        volumeNode = slicer.util.loadVolume(filePath)

        if volumeNode is None:
            raise RuntimeError(f"Failed to load volume file: {filePath}")
        
        if not volumeNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"Loaded node is not a scalar volume node: {volumeNode.GetClassName()}"
            )
        
        return volumeNode

    def loadDicomDirAsVolumeNode(self, dicomDir, modalityPreference=None,
                                 preferSUV=False, roleName="Volume",
                                 subjectId=None):
        """Import a DICOM series *folder* into the Slicer DICOM database and load
        it as a scalar volume node.

        modalityPreference : tuple of modality codes to prefer, e.g. ("PT","NM")
                             for PET or ("CT",) for CT. None = accept any.
        preferSUV          : when True (PET), prefer a SUV-converted volume that
                             the DICOM PET-SUV plugin produces (name contains 'suv').
        """
        from DICOMLib import DICOMUtils

        if not dicomDir:
            raise ValueError("No DICOM folder selected.")

        dicomDir = os.path.abspath(dicomDir)
        if not os.path.isdir(dicomDir):
            raise ValueError(f"DICOM folder does not exist: {dicomDir}")

        db = slicer.dicomDatabase
        if db is None:
            raise RuntimeError(
                "Slicer DICOM database is not available. Open the 'DICOM' module "
                "once to initialise it, then retry."
            )

        def _all_series():
            uids = set()
            for pat in db.patients():
                for study in db.studiesForPatient(pat):
                    for series in db.seriesForStudy(study):
                        uids.add(series)
            return uids

        before_series = _all_series()
        before_node_ids = {
            n.GetID() for n in slicer.util.getNodesByClass("vtkMRMLScalarVolumeNode")
        }

        print(f"[PETBiomarkerStudio] Importing DICOM folder: {dicomDir}")
        DICOMUtils.importDicom(dicomDir)

        after_series = _all_series()
        new_series = after_series - before_series

        # Fallback: if the importer reports nothing new (already imported earlier),
        # match the series whose files live in this folder.
        if not new_series:
            dir_norm = os.path.normpath(dicomDir)
            for uid in after_series:
                files = db.filesForSeries(uid)
                if files and os.path.normpath(os.path.dirname(files[0])) == dir_norm:
                    new_series.add(uid)

        if not new_series:
            raise RuntimeError(f"No DICOM series found in folder: {dicomDir}")

        # Filter by modality (PT/NM for PET, CT for CT)
        selected = []
        for uid in new_series:
            files = db.filesForSeries(uid)
            mod = db.fileValue(files[0], "0008,0060") if files else ""
            desc = db.fileValue(files[0], "0008,103e") if files else ""
            mod = (mod or "").strip().upper()
            print(f"[PETBiomarkerStudio]   series modality={mod!r} desc={desc!r}")
            if modalityPreference is None or mod in modalityPreference:
                selected.append(uid)
        if not selected:
            selected = list(new_series)

        loaded_ids = []
        for uid in selected:
            try:
                node_ids = DICOMUtils.loadSeriesByUID([uid])
                loaded_ids.extend(node_ids)
            except Exception as e:
                print(f"[PETBiomarkerStudio]   load failed for series {uid}: {e}")

        # Collect scalar volumes that were loaded for this series, plus any new
        # volumes that appeared (the SUV plugin may add an extra SUVbw node).
        candidates = []
        seen = set()
        for n in slicer.util.getNodesByClass("vtkMRMLScalarVolumeNode"):
            nid = n.GetID()
            if nid in seen:
                continue
            if nid in loaded_ids or nid not in before_node_ids:
                candidates.append(n)
                seen.add(nid)

        if not candidates:
            raise RuntimeError(
                f"DICOM series found but no scalar volume was loaded: {dicomDir}"
            )

        chosen = None
        if preferSUV:
            chosen = next(
                (n for n in candidates if "suv" in n.GetName().lower()), None
            )
            if chosen is None:
                slicer.util.warningDisplay(
                    f"{roleName}: no SUV-converted volume was produced. The raw PET "
                    "series was loaded instead, but SUV metrics need an SUVbw volume. "
                    "Enable the DICOM PET-SUV plugin (DICOM module > settings) and reload."
                )
        if chosen is None:
            chosen = candidates[0]

        # Give PET a clear SUVbw name
        if preferSUV and "suv" in chosen.GetName().lower() and "SUVbw" not in chosen.GetName():
            tag = subjectId or os.path.basename(dicomDir)
            chosen.SetName(slicer.mrmlScene.GenerateUniqueName(f"SUVbw_{tag}"))

        print(f"[PETBiomarkerStudio] {roleName} DICOM volume: {chosen.GetName()}")
        return chosen

    def loadSegmentationNodeFromFile(self, filePath, nodeName=None):
        if not filePath:
            raise ValueError("No segmentation file selected.")
        
        filePath = os.path.abspath(filePath)

        if not os.path.exists(filePath):
            raise ValueError(f"Segmentation file does not exist: {filePath}")

        if nodeName is None:
            # Remove extension from file name
            baseName = os.path.basename(filePath)
            baseName = baseName.replace(".nrrd", "")
            baseName = baseName.replace(".seg.nrrd", "")
            baseName = baseName.replace(".nii", "")
            baseName = baseName.replace(".nii.gz", "")
            nodeName = slicer.mrmlScene.GenerateUniqueName(baseName)

        segNode = slicer.util.loadSegmentation(filePath)

        if segNode:
            segNode.SetName(nodeName)
            return segNode
        
        # If failed to load as segmentation, try to load as label volume
        labelNode = slicer.util.loadLabelVolume(filePath)

        if labelNode is None:
            raise RuntimeError(
                f"Failed to load file as segmentation or labelmap: {filePath}"
            )
        
        segNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode", nodeName)

        slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(labelNode, segNode)

        segNode.CreateClosedSurfaceRepresentation()
        slicer.mrmlScene.RemoveNode(labelNode)

        return segNode

    def resolveCTInputFromScene(self, ctNode):
        print("[PETBiomarkerExtractor] Resolving CT input from scene:")
        # 1. Validity Check
        # 2. Return Info

        # 1-1.Check existence
        if ctNode is None:
            raise ValueError("No CT volume selected from the scene")
        # 1-2. Check valid type
        if not ctNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"Selected node is not a scalar volume node: {ctNode.GetClassName()}"
            )

        # 2. Return Info
        return {
            "sourceMode": "scene",
            "inputType": "volume_node",
            "node": ctNode,
            "description": ctNode.GetName()
        }

    # TODO: unused function
    def resolveCTInputFromPath(self, ctPath):
        if not ctPath:
            raise ValueError("No CT path provided")
        
        ctPath = os.path.abspath(ctPath)

        if not os.path.exists(ctPath):
            raise ValueError(f"CT path does not exist: {ctPath}")

        if os.path.isdir(ctPath):
            return {
                "sourceMode": "path",
                "inputType": "dicom_folder",
                "path": ctPath,
                "description": ctPath
            }

        if os.path.isfile(ctPath):
            lowerPath = ctPath.lower()

            if lowerPath.endswith(".nii") or lowerPath.endswith(".nii.gz"):

                return {
                    "sourceMode": "path",
                    "inputType": "nifti_file",
                    "path": ctPath,
                    "description": ctPath
                }
            
            raise ValueError(
                "Unsupported CT file type. "
                "Please select a NIfTI file (*.nii, *.nii.gz) or a DICOM folder." 
            )
        raise ValueError(f"invalid CT path: {ctPath}")

    # ----- TotalSegmentator related functions -----
    def prepareCTNiftiInput(self, ctInputInfo):
        inputType = ctInputInfo["inputType"]

        if inputType == "volume_node":
            ctNode = ctInputInfo["node"]

            safeName = re.sub(r"[^A-Za-z0-9_.-]", "_", ctNode.GetName())
            workDir = tempfile.mkdtemp(
                prefix=f"PETBio_CT_{safeName}_",
                dir=slicer.app.temporaryPath,
            )

            ctNiftiPath = os.path.join(workDir, f"{safeName}.nii.gz")

            print(f"[PETBiomarkerExtractor] Saving CT volume node to:")
            print(ctNiftiPath)

            success = slicer.util.saveNode(ctNode, ctNiftiPath)

            if not success:
                raise RuntimeError(f"Failed to save CT volume node to: {ctNiftiPath}")
            
            return {
                    "inputType": "nifti_file",
                    "niftiPath": ctNiftiPath,
                    "workDir": workDir,
                    "description": ctNiftiPath,
            }

        if inputType == "nifti_file":

            return {
                    "inputType": "nifti_file",
                    "niftiPath": ctInputInfo["path"],
                    "workDir": None,
                    "description": ctInputInfo["path"],
            }

        if inputType == "dicom_folder":
            raise NotImplementedError(
                "DICOM folder input is recognized, but DICOM-to-NIfTI conversion "
                "will be added in the next step."
            )
        
        raise ValueError(f"Invalid CT input type: {inputType}")
        
    def runTotalSegmentator(self, ctNiftiPath, workDir=None, selectedStems=None):
        if not ctNiftiPath:
            raise ValueError("No CT NIfTI path provided.")

        if not os.path.exists(ctNiftiPath):
            raise ValueError(f"CT NIfTI file does not exist: {ctNiftiPath}")

        if workDir is None:
            workDir = tempfile.mkdtemp(
                prefix="PETBio_TotalSeg_",
                dir=slicer.app.temporaryPath,
            )
        
        outputDir = os.path.join(workDir, "totalseg_output")
        os.makedirs(outputDir, exist_ok=True)

        command = [
            "TotalSegmentator",
            "-i", ctNiftiPath,
            "-o", outputDir,
            "--nr_thr_resamp", "1",
            "--nr_thr_saving", "1",
        ]

        if selectedStems:
            command += ["--roi_subset"] + selectedStems

        print("[PETBiomarkerExtractor] Running TotalSegmentator:")
        print(" ".join(command))

        # Run TotalSegmentator
        result = subprocess.run(
            command,
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )

        # Log stdout and stderr
        stdoutText = self.decodeProcessOutput(result.stdout)
        stderrText = self.decodeProcessOutput(result.stderr)

        if stdoutText.strip():
            print("[TotalSegmentator stdout]")
            print(stdoutText)
        
        if stderrText.strip():
            print("[TotalSegmentator stderr]")
            print(stderrText)
        
        if result.returncode != 0:
            raise RuntimeError(
                "TotalSegmentator failed. "
                f"Exit code: {result.returncode}"
                "Check the Python colsole output above for stdout/stderr"
            )

        # Get output files
        outputFiles = [
            f for f in os.listdir(outputDir)
            if f.endswith(".nii.gz") or f.endswith(".nii")
        ]

        return {
            "outputDir": outputDir,
            "outputFileCount": len(outputFiles),
            "outputFiles": outputFiles,
        }

    def decodeProcessOutput(self, data):
        if not data:
            return ""
        
        encodings = ["utf-8", "cp949", "mbcs"]

        for encoding in encodings:
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                pass
            except LookupError:
                pass
        
        return data.decode("utf-8", errors="replace")

    def loadResultNIfTIAsSegmentationNode(self, outputDir, stem):
        if not outputDir:
            raise ValueError("No TotalSegmentator output directory provided.")
        
        maskPath = os.path.join(outputDir, f"{stem}.nii.gz")

        if not os.path.exists(maskPath):
            raise ValueError(f"Requested mask does not exist: {maskPath}")
        
        nodeName = slicer.mrmlScene.GenerateUniqueName(f"Totalseg_{stem}")

        print(f"[PETBiomarkerExtractor] Loading TotalSegmentator mask:")
        print(maskPath)

        return self.loadSegmentationNodeFromFile(maskPath, nodeName)

    def loadSelectedMasksAsSegmentation(self, outputDir, selectedStems):
        if isinstance(selectedStems, str):
            selectedStems = [selectedStems]
        
        loadedSegNodes = []
        missingStems = []

        for stem in selectedStems:
            try:
                segNode = self.loadResultNIfTIAsSegmentationNode(outputDir, stem)
                loadedSegNodes.append(segNode)
            
            except ValueError as e:
                print(f"[PETBiomarkerExtractor] Missing or Failed mask: {stem}")
                print(e)
                missingStems.append(stem)
        
        if not loadedSegNodes:
            raise RuntimeError(
                f"No requested TotalSegmentator masks were loaded. "
                f"Requested: {selectedStems}, missing: {missingStems}"
            )
        
        print(f"[PETBiomarkerExtractor] Loaded segmentations:")
        for node in loadedSegNodes:
            print(f"  - {node.GetName()}")
        
        if missingStems:
            print(f"[PETBiomarkerExtractor] Missing stems: {missingStems}")

        return loadedSegNodes

    def loadResultNIfTIAsSingleSegmentationNode(self, outputDir, stems, nodeName):
        if not outputDir:
            raise ValueError("No TotalSegmentator output directory provided.")

        if isinstance(stems, str):
            stems = [stems]

        uniqueNodeName = slicer.mrmlScene.GenerateUniqueName(nodeName)

        combinedSegNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLSegmentationNode",
            uniqueNodeName,
        )

        loadedStems = []
        missingStems = []

        try:
            for stem in stems:
                maskPath = os.path.join(outputDir, f"{stem}.nii.gz")

                if not os.path.exists(maskPath):
                    print(f"[PETBiomarkerExtractor] Missing mask: {maskPath}")
                    missingStems.append(stem)
                    continue
                
                print("[PETBiomarkerExtractor] Importing vertebra mask:")
                print(maskPath)

                labelNode = slicer.util.loadLabelVolume(maskPath)

                if labelNode is None:
                    missingStems.append(stem)
                    continue

                labelNode.SetName(stem)

                beforeCount = combinedSegNode.GetSegmentation().GetNumberOfSegments()

                slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                    labelNode,
                    combinedSegNode,
                )

                afterCount = combinedSegNode.GetSegmentation().GetNumberOfSegments()

                if afterCount > beforeCount:
                    newSegment = combinedSegNode.GetSegmentation().GetNthSegment(afterCount - 1)
                    newSegment.SetName(stem)

                slicer.mrmlScene.RemoveNode(labelNode)
                loadedStems.append(stem)
            
            if not loadedStems:
                slicer.mrmlScene.RemoveNode(combinedSegNode)
                raise ValueError(
                    f"No vertebrae masks were loaded. Requested: {stems}, missing: {missingStems}"
                )
            
            combinedSegNode.CreateClosedSurfaceRepresentation()

            print("[PETBiomarkerExtractor] Combined segmentation created:")
            print(f"  Node: {combinedSegNode.GetName()}")
            print(f"  Loaded stems: {loadedStems}")
            print(f"  Missing stems: {missingStems}")

            return combinedSegNode

        except Exception:
            try:
                slicer.mrmlScene.RemoveNode(combinedSegNode)
            except Exception:
                pass
            raise

    # ----- Spine Level Slicing related functions -----
    def generateVertebraeFromCTNode(self, ctNode):
        print("[PETBiomarkerExtractor] Generating vertebrae from CT node:")
        print(f"  CT: {ctNode.GetName()}")

        if ctNode is None:
            slicer.util.errorDisplay("No CT volume selected.")
            raise ValueError("No CT volume selected.")
        
        if not ctNode.IsA("vtkMRMLScalarVolumeNode"):
            slicer.util.errorDisplay(
                f"CT node must be a scalar volume node, but got: {ctNode.GetClassName()}"
            )
            raise ValueError(
                f"CT node must be a scalar volume node, but got: {ctNode.GetClassName()}"
            )

        selectedStems = [
            "vertebrae_L1",
            "vertebrae_L2",
            "vertebrae_L3",
            "vertebrae_L4",
            "vertebrae_L5",
        ]

        ctInputInfo = self.resolveCTInputFromScene(ctNode)
        ctNiftiInfo = self.prepareCTNiftiInput(ctInputInfo)
      
        totalSegResult = self.runTotalSegmentator(ctNiftiInfo["niftiPath"], ctNiftiInfo["workDir"], selectedStems)

        vertebraeNode = self.loadResultNIfTIAsSingleSegmentationNode(
            totalSegResult["outputDir"], 
            selectedStems, 
            nodeName="TotalSeg_vertebrae_L1_L5"
        )

        generationInfo = {
            "selectedStems": selectedStems,
            "outputDir": totalSegResult["outputDir"],
            "outputFiles": totalSegResult["outputFiles"],
            "ctNiftiPath": ctNiftiInfo["niftiPath"],
        }

        return vertebraeNode, generationInfo         

    def findSpineSegmentNames(self, vertebraeNode, selectedLevels):
        if vertebraeNode is None:
            raise ValueError("No vertebrae segmentation selected.")

        if not vertebraeNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"Vertebrae node must be a segmentation node, "
                f"but got: {vertebraeNode.GetClassName()}"
            )

        segmentation = vertebraeNode.GetSegmentation()

        if segmentation is None or segmentation.GetNumberOfSegments() == 0:
            raise ValueError(
                f"Vertebrae segmentation has no segments: {vertebraeNode.GetName()}"
            )

        availableNames = [
            segmentation.GetNthSegment(i).GetName() for i in range(segmentation.GetNumberOfSegments())
        ]

        matchedNames = []

        for level in selectedLevels:
            # "L2" -> "2"
            levelNumber = level[1:]
            pattern = rf"(^|[^a-zA-Z0-9])l{levelNumber}([^0-9]|$)"

            matches = [
                name for name in availableNames 
                if re.search(pattern, name.lower())
            ]

            if not matches:
                raise ValueError(
                    f"Could not find segment for {level}.\n"
                    f"Available segments: {availableNames}"
                )
            
            matchedNames.append(matches[0])

        return matchedNames
            
    def computeSpineRangeZBounds(self, vertebraeNode, selectedLevels):
        import numpy as np
        
        segmentNames = self.findSpineSegmentNames(vertebraeNode, selectedLevels)

        segmentation = vertebraeNode.GetSegmentation()
        segmentIds = vtk.vtkStringArray()

        for name in segmentNames:
            segmentId = segmentation.GetSegmentIdBySegmentName(name)

            if not segmentId:
                raise ValueError(f"Segment ID not found for segment: {name}")

            segmentIds.InsertNextValue(segmentId)

        labelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_spine_range_labelmap",
        )

        try:
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                vertebraeNode,
                segmentIds,
                labelNode
            )

            arr = slicer.util.arrayFromVolume(labelNode)
            voxels = np.argwhere(arr > 0)

            if len(voxels) == 0:
                raise ValueError(
                    f"Selected spine segments contain no foreground voxels: {segmentNames}"
                )
            
            mat = vtk.vtkMatrix4x4()
            labelNode.GetIJKToRASMatrix(mat)
            affine = self.matrix4x4ToNumpy(mat)

            ijkHom = np.column_stack([
                voxels[:, 2].astype(float),  # I
                voxels[:, 1].astype(float),  # J
                voxels[:, 0].astype(float),  # K
                np.ones(len(voxels)),
            ])

            ras = (affine @ ijkHom.T).T
            rasZ = ras[:, 2]

            return {
                "selectedLevels": selectedLevels,
                "segmentNames": segmentNames,
                "zInferior": float(rasZ.min()),
                "zSuperior": float(rasZ.max()),
                "voxelCount": int(len(voxels)),
            }

        finally:
            slicer.mrmlScene.RemoveNode(labelNode)
        
    def matrix4x4ToNumpy(self, matrix):
        import numpy as np
        return np.array([
            [matrix.GetElement(r, c) for c in range(4)]
            for r in range(4)
        ])

    def applySpineRangeSlicing(self, roiNode, vertebraeNode, selectedLevels, referenceVolumeNode):
        import numpy as np

        if roiNode is None:
            raise ValueError("No ROI segmentation selected.")

        if not roiNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"ROI node must be a segmentation node, but got: {roiNode.GetClassName()}"
            )

        if referenceVolumeNode is None:
            raise ValueError("No reference volume selected. Please select a CT volume.")

        if not referenceVolumeNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"Reference node must be a scalar volume node, but got: {referenceVolumeNode.GetClassName()}"
            )

        boundsInfo = self.computeSpineRangeZBounds(vertebraeNode, selectedLevels)

        zInferior = boundsInfo['zInferior']
        zSuperior = boundsInfo['zSuperior']

        segmentation = roiNode.GetSegmentation()

        if segmentation is None or segmentation.GetNumberOfSegments() == 0:
            raise ValueError(f"ROI segmentation has no segments: {roiNode.GetName()}")
        
        # TODO: For now, process the first segment only.
        segmentId = segmentation.GetNthSegmentID(0)
        segmentName = segmentation.GetNthSegment(0).GetName()

        segmentIds = vtk.vtkStringArray()
        segmentIds.InsertNextValue(segmentId)

        labelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_spine_slicing_labelmap",
        )

        try:
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                roiNode,
                segmentIds,
                labelNode,
                referenceVolumeNode,
                slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY
            )

            arr = slicer.util.arrayFromVolume(labelNode)
            if int(np.count_nonzero(arr > 0)) == 0:
                raise ValueError(f"ROI segment contains no foreground voxels: {segmentName}")

            mat = vtk.vtkMatrix4x4()
            labelNode.GetIJKToRASMatrix(mat)
            affine = self.matrix4x4ToNumpy(mat)

            clipped, clipInfo = _lib_clip_binary_mask_by_ras_z(
                arr, affine, zInferior, zSuperior
            )
            arr[:] = np.where(clipped, arr, 0)
            slicer.util.arrayFromVolumeModified(labelNode)

            originalVoxelCount = clipInfo["originalVoxelCount"]
            remainingVoxelCount = clipInfo["remainingVoxelCount"]
            removedVoxelCount = clipInfo["removedVoxelCount"]

            outputName = slicer.mrmlScene.GenerateUniqueName(
                f"{roiNode.GetName()}_spine_{selectedLevels[0]}_{selectedLevels[-1]}"
            )

            processedNode = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLSegmentationNode",
                outputName,
            )
            
            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                labelNode,
                processedNode
            )
            
            processedSegmentation = processedNode.GetSegmentation()
            if processedSegmentation.GetNumberOfSegments() > 0:
                processedSegmentation.GetNthSegment(0).SetName(
                    f"{segmentName}_spine_{selectedLevels[0]}_{selectedLevels[-1]}"
                )

            processedNode.CreateClosedSurfaceRepresentation()

            slicingInfo = {
                **boundsInfo,
                "originalRoiNode": roiNode.GetName(),
                "processedRoiNode": processedNode.GetName(),
                "originalSegmentName": segmentName,
                "originalVoxelCount": originalVoxelCount,
                "remainingVoxelCount": remainingVoxelCount,
                "removedVoxelCount": removedVoxelCount,
            }

            return processedNode, slicingInfo
        finally:
            slicer.mrmlScene.RemoveNode(labelNode)

    # ----- Kidney exclusion related functions -----
    def generateKidneysFromCTNode(self, ctNode):
        if ctNode is None:
            raise ValueError("No CT volume selected.")

        if not ctNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"CT node must be a scalar volume node, but got: {ctNode.GetClassName()}"
            )

        print("[PETBiomarkerExtractor] Generating kidneys from CT node:")
        print(f"  CT: {ctNode.GetName()}")

        selectedStems = [
            "kidney_left",
            "kidney_right",
        ]

        ctInputInfo = self.resolveCTInputFromScene(ctNode)
        ctNiftiInfo = self.prepareCTNiftiInput(ctInputInfo)

        totalSegResult = self.runTotalSegmentator(
            ctNiftiPath=ctNiftiInfo["niftiPath"],
            workDir=ctNiftiInfo["workDir"],
            selectedStems=selectedStems,
        )

        kidneyNode = self.loadResultNIfTIAsSingleSegmentationNode(
            outputDir=totalSegResult["outputDir"],
            stems=selectedStems,
            nodeName="TotalSeg_kidneys",
        )

        generationInfo = {
            "selectedStems": selectedStems,
            "outputDir": totalSegResult["outputDir"],
            "outputFiles": totalSegResult["outputFiles"],
            "ctNiftiPath": ctNiftiInfo["niftiPath"],
        }

        return kidneyNode, generationInfo
    
    def findKidneySegmentNames(self, kidneyNode):
        if kidneyNode is None:
            raise ValueError("No kidney segmentation selected.")

        if not kidneyNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"Kidney node must be a segmentation node, "
                f"but got: {kidneyNode.GetClassName()}"
            )

        segmentation = kidneyNode.GetSegmentation()

        if segmentation is None or segmentation.GetNumberOfSegments() == 0:
            raise ValueError(
                f"Kidney segmentation has no segments: {kidneyNode.GetName()}"
            )

        availableNames = [
            segmentation.GetNthSegment(i).GetName()
            for i in range(segmentation.GetNumberOfSegments())
        ]

        matchedNames = []

        for name in availableNames:
            lowerName = name.lower()
            if "kidney" in lowerName:
                matchedNames.append(name)

        if matchedNames:
            return matchedNames

        print(
            "[PETBiomarkerExtractor] No segment name contains 'kidney'. "
            "Using all segments in the selected kidney node."
        )

        return availableNames

    def createPhysicalDilationStructure(self, spacing, radiusMm):
        return _lib_physical_dilation_structure(spacing, radiusMm)
    
    def applyKidneyExclusion(
        self,
        roiNode,
        kidneyNode,
        referenceVolumeNode,
        dilationRadiusMm,
    ):
        from scipy import ndimage

        if roiNode is None:
            raise ValueError("No ROI segmentation selected.")

        if not roiNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"ROI node must be a segmentation node, "
                f"but got: {roiNode.GetClassName()}"
            )

        if kidneyNode is None:
            raise ValueError("No kidney segmentation selected.")

        if not kidneyNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"Kidney node must be a segmentation node, "
                f"but got: {kidneyNode.GetClassName()}"
            )

        if referenceVolumeNode is None:
            raise ValueError("No reference volume selected. Please select a CT volume.")

        if not referenceVolumeNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"Reference node must be a scalar volume node, "
                f"but got: {referenceVolumeNode.GetClassName()}"
            )

        roiSegmentation = roiNode.GetSegmentation()

        if roiSegmentation is None or roiSegmentation.GetNumberOfSegments() == 0:
            raise ValueError(f"ROI segmentation has no segments: {roiNode.GetName()}")

        # 현재는 첫 번째 ROI segment만 처리
        roiSegmentId = roiSegmentation.GetNthSegmentID(0)
        roiSegmentName = roiSegmentation.GetNthSegment(0).GetName()

        roiSegmentIds = vtk.vtkStringArray()
        roiSegmentIds.InsertNextValue(roiSegmentId)

        kidneySegmentNames = self.findKidneySegmentNames(kidneyNode)
        kidneySegmentation = kidneyNode.GetSegmentation()

        kidneySegmentIds = vtk.vtkStringArray()

        for name in kidneySegmentNames:
            segmentId = kidneySegmentation.GetSegmentIdBySegmentName(name)

            if not segmentId:
                raise ValueError(f"Kidney segment ID not found for segment: {name}")

            kidneySegmentIds.InsertNextValue(segmentId)

        roiLabelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_roi_for_kidney_exclusion",
        )

        kidneyLabelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_kidney_for_exclusion",
        )

        try:
            # ROI를 CT geometry 기준 labelmap으로 export
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                roiNode,
                roiSegmentIds,
                roiLabelNode,
                referenceVolumeNode,
                slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
            )

            # Kidney도 같은 CT geometry 기준 labelmap으로 export
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                kidneyNode,
                kidneySegmentIds,
                kidneyLabelNode,
                referenceVolumeNode,
                slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
            )

            roiArr = slicer.util.arrayFromVolume(roiLabelNode)
            kidneyArr = slicer.util.arrayFromVolume(kidneyLabelNode)

            if int(np.count_nonzero(roiArr > 0)) == 0:
                raise ValueError(
                    f"ROI segment contains no foreground voxels: {roiSegmentName}"
                )
            if int(np.count_nonzero(kidneyArr > 0)) == 0:
                raise ValueError(
                    f"Kidney segmentation contains no foreground voxels: "
                    f"{kidneySegmentNames}"
                )

            spacing = referenceVolumeNode.GetSpacing()
            cleaned, exclInfo = _lib_exclude_dilated_structure(
                roiArr,
                kidneyArr,
                spacing=spacing,
                dilation_radius_mm=float(dilationRadiusMm),
            )
            roiArr[:] = cleaned
            slicer.util.arrayFromVolumeModified(roiLabelNode)

            originalVoxelCount = exclInfo["originalVoxelCount"]
            remainingVoxelCount = exclInfo["remainingVoxelCount"]
            removedVoxelCount = exclInfo["removedVoxelCount"]

            outputName = slicer.mrmlScene.GenerateUniqueName(
                f"{roiNode.GetName()}_kidney_excluded_{dilationRadiusMm:.1f}mm"
            )

            processedNode = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLSegmentationNode",
                outputName,
            )

            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                roiLabelNode,
                processedNode,
            )

            processedSegmentation = processedNode.GetSegmentation()

            if processedSegmentation.GetNumberOfSegments() > 0:
                processedSegmentation.GetNthSegment(0).SetName(
                    f"{roiSegmentName}_kidney_excluded_{dilationRadiusMm:.1f}mm"
                )

            processedNode.CreateClosedSurfaceRepresentation()

            exclusionInfo = {
                "originalRoiNode": roiNode.GetName(),
                "processedRoiNode": processedNode.GetName(),
                "originalSegmentName": roiSegmentName,
                "kidneyNode": kidneyNode.GetName(),
                "kidneySegmentNames": kidneySegmentNames,
                "dilationRadiusMm": float(dilationRadiusMm),
                "originalVoxelCount": originalVoxelCount,
                "remainingVoxelCount": remainingVoxelCount,
                "removedVoxelCount": removedVoxelCount,
            }

            return processedNode, exclusionInfo

        finally:
            slicer.mrmlScene.RemoveNode(roiLabelNode)
            slicer.mrmlScene.RemoveNode(kidneyLabelNode)
   
    # ----- Urinary activity cleanup related functions -----
    def generatePETDerivedUreterMask(
        self,
        petNode,
        vertebraeNode,
        selectedLevels,
        suvThreshold,
        dilationRadiusMm,
        superiorExtensionMm=50.0,
        inferiorExtensionMm=200.0,
    ):
        if petNode is None:
            raise ValueError("No PET volume selected.")

        if not petNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"PET node must be a scalar volume node, but got: {petNode.GetClassName()}"
            )

        if petNode.GetImageData() is None:
            raise ValueError(f"PET node has no image data: {petNode.GetName()}")

        if vertebraeNode is None:
            raise ValueError("No vertebrae segmentation selected.")

        boundsInfo = self.computeSpineRangeZBounds(
            vertebraeNode=vertebraeNode,
            selectedLevels=selectedLevels,
        )

        # Ureter / urinary activity may extend below the selected spine range.
        zInferiorUsed = boundsInfo["zInferior"] - float(inferiorExtensionMm)
        zSuperiorUsed = boundsInfo["zSuperior"] + float(superiorExtensionMm)

        petArr = slicer.util.arrayFromVolume(petNode)

        mat = vtk.vtkMatrix4x4()
        petNode.GetIJKToRASMatrix(mat)
        affine = self.matrix4x4ToNumpy(mat)
        spacing = petNode.GetSpacing()

        maskArr, maskCoreInfo = _lib_build_pet_urinary_mask(
            petArr,
            affine,
            spacing,
            z_inferior=zInferiorUsed,
            z_superior=zSuperiorUsed,
            suv_threshold=float(suvThreshold),
            dilation_radius_mm=float(dilationRadiusMm),
        )
        initialHotVoxelCount = maskCoreInfo["initialHotVoxelCount"]
        zClippedHotVoxelCount = maskCoreInfo["zClippedHotVoxelCount"]
        afterBladderRemovalVoxelCount = maskCoreInfo["afterBladderRemovalVoxelCount"]
        finalMaskVoxelCount = maskCoreInfo["finalMaskVoxelCount"]

        labelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_pet_derived_urinary_activity_mask",
        )

        try:

            slicer.util.updateVolumeFromArray(labelNode, maskArr)
            labelNode.CopyOrientation(petNode)

            outputName = slicer.mrmlScene.GenerateUniqueName(
                "PET_derived_urinary_activity_mask"
            )

            maskSegNode = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLSegmentationNode",
                outputName,
            )

            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                labelNode,
                maskSegNode,
            )

            segmentation = maskSegNode.GetSegmentation()
            if segmentation.GetNumberOfSegments() > 0:
                segmentation.GetNthSegment(0).SetName(
                    "pet_derived_urinary_activity"
                )

            maskSegNode.CreateClosedSurfaceRepresentation()

            maskInfo = {
                **boundsInfo,
                "zInferiorUsed": float(zInferiorUsed),
                "zSuperiorUsed": float(zSuperiorUsed),
                "suvThreshold": float(suvThreshold),
                "dilationRadiusMm": float(dilationRadiusMm),
                "initialHotVoxelCount": initialHotVoxelCount,
                "zClippedHotVoxelCount": zClippedHotVoxelCount,
                "afterBladderRemovalVoxelCount": afterBladderRemovalVoxelCount,
                "finalMaskVoxelCount": finalMaskVoxelCount,
            }

            return maskSegNode, maskInfo

        finally:
            slicer.mrmlScene.RemoveNode(labelNode)
   
    def getAllSegmentIds(self, segNode):
        if segNode is None:
            raise ValueError("No segmentation node selected.")

        if not segNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"Expected vtkMRMLSegmentationNode, got: {segNode.GetClassName()}"
            )

        segmentation = segNode.GetSegmentation()

        if segmentation is None or segmentation.GetNumberOfSegments() == 0:
            raise ValueError(f"Segmentation has no segments: {segNode.GetName()}")

        segmentIds = vtk.vtkStringArray()

        for i in range(segmentation.GetNumberOfSegments()):
            segmentIds.InsertNextValue(segmentation.GetNthSegmentID(i))

        return segmentIds

    def applyUrinaryActivityCleanup(
        self,
        roiNode,
        petNode,
        urinaryMaskNode,
        cleanSUVThreshold,
    ):
        import numpy as np

        if roiNode is None:
            raise ValueError("No ROI segmentation selected.")

        if not roiNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"ROI node must be a segmentation node, got: {roiNode.GetClassName()}"
            )

        if petNode is None:
            raise ValueError("No PET volume selected.")

        if not petNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"PET node must be a scalar volume node, got: {petNode.GetClassName()}"
            )

        if petNode.GetImageData() is None:
            raise ValueError(f"PET node has no image data: {petNode.GetName()}")

        if urinaryMaskNode is None:
            raise ValueError("No PET-derived urinary activity mask selected.")

        if not urinaryMaskNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"Urinary mask node must be a segmentation node, got: {urinaryMaskNode.GetClassName()}"
            )

        roiSegmentation = roiNode.GetSegmentation()

        if roiSegmentation is None or roiSegmentation.GetNumberOfSegments() == 0:
            raise ValueError(f"ROI segmentation has no segments: {roiNode.GetName()}")

        # 현재는 첫 번째 ROI segment만 처리
        roiSegmentId = roiSegmentation.GetNthSegmentID(0)
        roiSegmentName = roiSegmentation.GetNthSegment(0).GetName()

        roiSegmentIds = vtk.vtkStringArray()
        roiSegmentIds.InsertNextValue(roiSegmentId)

        urinarySegmentIds = self.getAllSegmentIds(urinaryMaskNode)

        roiLabelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_roi_for_urinary_cleanup",
        )

        urinaryLabelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_urinary_mask_for_cleanup",
        )

        try:
            # PET SUV threshold 조건을 써야 하므로 PET geometry 기준으로 맞춘다.
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                roiNode,
                roiSegmentIds,
                roiLabelNode,
                petNode,
                slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
            )

            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                urinaryMaskNode,
                urinarySegmentIds,
                urinaryLabelNode,
                petNode,
                slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
            )

            roiArr = slicer.util.arrayFromVolume(roiLabelNode)
            urinaryArr = slicer.util.arrayFromVolume(urinaryLabelNode)
            petArr = slicer.util.arrayFromVolume(petNode)

            try:
                cleaned, cleanupCore = _lib_apply_urinary_cleanup(
                    roiArr,
                    urinaryArr,
                    petArr,
                    clean_suv_threshold=float(cleanSUVThreshold),
                )
            except ValueError as e:
                msg = str(e)
                if "no foreground voxels" in msg.lower():
                    raise ValueError(
                        f"ROI segment contains no foreground voxels: {roiSegmentName}"
                    ) from e
                raise

            roiArr[:] = cleaned
            slicer.util.arrayFromVolumeModified(roiLabelNode)

            originalVoxelCount = cleanupCore["originalVoxelCount"]
            overlapVoxelCount = cleanupCore["overlapVoxelCount"]
            removedVoxelCount = cleanupCore["removedVoxelCount"]
            remainingVoxelCount = cleanupCore["remainingVoxelCount"]

            outputName = slicer.mrmlScene.GenerateUniqueName(
                f"{roiNode.GetName()}_urinary_cleaned_{cleanSUVThreshold:.1f}"
            )

            processedNode = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLSegmentationNode",
                outputName,
            )

            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                roiLabelNode,
                processedNode,
            )

            processedSegmentation = processedNode.GetSegmentation()

            if processedSegmentation.GetNumberOfSegments() > 0:
                processedSegmentation.GetNthSegment(0).SetName(
                    f"{roiSegmentName}_urinary_cleaned_{cleanSUVThreshold:.1f}"
                )

            processedNode.CreateClosedSurfaceRepresentation()

            cleanupInfo = {
                "originalRoiNode": roiNode.GetName(),
                "processedRoiNode": processedNode.GetName(),
                "originalSegmentName": roiSegmentName,
                "urinaryMaskNode": urinaryMaskNode.GetName(),
                "cleanSUVThreshold": float(cleanSUVThreshold),
                "originalVoxelCount": originalVoxelCount,
                "overlapVoxelCount": overlapVoxelCount,
                "removedVoxelCount": removedVoxelCount,
                "remainingVoxelCount": remainingVoxelCount,
            }

            return processedNode, cleanupInfo

        finally:
            slicer.mrmlScene.RemoveNode(roiLabelNode)
            slicer.mrmlScene.RemoveNode(urinaryLabelNode)
    
    # ----- PET-IndiC calculation related functions -----
    def validatePetMetricsInputs(self, petNode, roiNode, metricsOptions):
        if petNode is None:
            raise ValueError("No PET volume selected.")
        
        if not petNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"PET node must be a scalar volume node, but got: {petNode.GetClassName()}"
            )
        
        if petNode.GetImageData() is None:
            raise ValueError(f"PET node has no image data: {petNode.GetName()}")
        
        if roiNode is None:
            raise ValueError("No ROI segmentation selected.")
        
        if not roiNode.IsA("vtkMRMLSegmentationNode"):
            raise ValueError(
                f"ROI node must be a segmentation node, but got: {roiNode.GetClassName()}"
            )
        
        segmentation = roiNode.GetSegmentation()

        if segmentation is None or segmentation.GetNumberOfSegments() == 0:
            raise ValueError(f"ROI segmentation has no segments: {roiNode.GetName()}")

        if not any(metricsOptions.values()):
            raise ValueError("Select at least one metric to calculate")
        
        return True
    
    def runPetIndic(self, petNode, roiNode, metricsOptions):
        self.validatePetMetricsInputs(
            petNode=petNode,
            roiNode=roiNode,
            metricsOptions=metricsOptions,
        )

        segmentation = roiNode.GetSegmentation()
        segmentId = segmentation.GetNthSegmentID(0)

        print("[PETBiomarkerExtractor] Running PET metrics")
        print(f"  PET: {petNode.GetName()}")
        print(f"  ROI: {roiNode.GetName()}")
        print(f"  Segment ID: {segmentId}")

        labelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_pet_metric_label"
        )

        segmentIds = vtk.vtkStringArray()
        segmentIds.InsertNextValue(segmentId)

        slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
            roiNode,
            segmentIds, # segment IDs to export
            labelNode, # output label map node
            petNode, # reference volume node
            slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY
        )

        try:
            # QuantitativeIndicesCLI module
            qiModule = slicer.modules.quantitativeindicescli

            parameters = {
                "Grayscale_Image": petNode.GetID(),
                "Label_Image": labelNode.GetID(), # takes label map instead of segmentation node
                "Label_Value": "1"
            }

            parameters["Mean"] = metricsOptions["mean"]
            parameters["Max"] = metricsOptions["max"]
            parameters["Peak"] = metricsOptions["peak"]
            parameters["TLG"] = metricsOptions["tlg"]
            parameters["Volume"] = metricsOptions["volume"]

            print("[PETBiomarkerExtractor] QuantitativeIndicesCLI parameters:")
            print(parameters)

            # Run module in Slicer CLI
            cliNode = slicer.cli.run(
                qiModule,
                None,
                parameters,
                wait_for_completion=True,
            )

            try:
                status = cliNode.GetStatusString() 
                print(f"[PETBiomarkerExtractor] CLI status: {status}")

                if status.lower() not in ("completed", "complete"):
                    raise RuntimeError(f"QuantitativeIndicesCLI failed with status: {status}")
                
                results = self.parseQuantitativeIndicesCliResults(cliNode)

                if not results:
                    raise RuntimeError(
                        "QuantitativeIndicesCLI completed, but no numeric metric was parsed."
                    )

                return results

            finally:
                slicer.mrmlScene.RemoveNode(cliNode)
        finally:
            slicer.mrmlScene.RemoveNode(labelNode)

    def parseQuantitativeIndicesCliResults(self, cliNode):
        pairs = []
        nParams = cliNode.GetNumberOfParametersInGroup(3)
        for i in range(nParams):
            pairs.append(
                (cliNode.GetParameterName(3, i), cliNode.GetParameterDefault(3, i))
            )
        results = _lib_parse_qi_results(pairs)
        print("[PETBiomarkerExtractor] Parsed PET metric results:")
        print(results)
        return results

   # ----- Radiomics calculation related functions -----
    def ensureRadiomics(self):
        try:
            return _lib_radiomics.ensure_radiomics_featureextractor()
        except ImportError:
            print("[PETBiomarkerExtractor] pyradiomics is not installed. Installing...")
            slicer.util.pip_install("pyradiomics")
            return _lib_radiomics.ensure_radiomics_featureextractor()

    def makeRadiomicsExtractor(self, radiomicsOptions):
        return _lib_radiomics.make_radiomics_extractor(radiomicsOptions)

    def _calculateWithinRoiRadiomicsDerived(self, features):
        return _lib_radiomics.derived_within_roi_features(features)

    def runRadiomics(self, petNode, roiNode, radiomicsOptions):
        if not self.isRadiomicsEnabled(radiomicsOptions):
            return {}

        self.validatePetMetricsInputs(
            petNode=petNode,
            roiNode=roiNode,
            metricsOptions={"dummy": True},
        )

        segmentation = roiNode.GetSegmentation()
        if segmentation.GetNumberOfSegments() == 0:
            raise ValueError("ROI segmentation has no segments.")

        segmentId = segmentation.GetNthSegmentID(0)

        labelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode",
            "temp_radiomics_label",
        )

        segmentIds = vtk.vtkStringArray()
        segmentIds.InsertNextValue(segmentId)

        slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
            roiNode,
            segmentIds,
            labelNode,
            petNode,
            slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY
        )

        try:
            with tempfile.TemporaryDirectory() as tmpDir:
                imagePath = os.path.join(tmpDir, "pet_suv.nrrd")
                maskPath = os.path.join(tmpDir, "mask.nrrd")

                if not slicer.util.saveNode(petNode, imagePath):
                    raise RuntimeError(f"Failed to save PET node to: {imagePath}")

                if not slicer.util.saveNode(labelNode, maskPath):
                    raise RuntimeError(f"Failed to save mask node to: {maskPath}")

                print("[PETBiomarkerExtractor] Running PyRadiomics via lib...")
                print(f"  Image: {imagePath}")
                print(f"  Mask: {maskPath}")
                selectedFeatureKeys = self.selectedRadiomicsFeatureKeys(
                    radiomicsOptions
                )
                print(
                    "  Selected features: "
                    + (", ".join(selectedFeatureKeys) if selectedFeatureKeys else "none")
                )
                print(f"  Bin width: {radiomicsOptions.get('bin_width', 0.25)}")

                features = _lib_radiomics.extract_radiomics_from_paths(
                    imagePath, maskPath, radiomicsOptions, label=1
                )

            radCount = len([key for key in features if key.startswith("rad_")])
            print(f"[PETBiomarkerExtractor] Radiomics extracted: {radCount} features")
            return features

        finally:
            slicer.mrmlScene.RemoveNode(labelNode)

    def loadMetricRowsFromExcel(self, excelPath):
        if not excelPath:
            raise ValueError("No quantification Excel file selected.")
        excelPath = os.path.abspath(excelPath)
        if not os.path.isfile(excelPath):
            raise FileNotFoundError(excelPath)

        try:
            import openpyxl
        except ModuleNotFoundError:
            slicer.util.pip_install("openpyxl")
            import importlib
            openpyxl = importlib.import_module("openpyxl")

        workbook = openpyxl.load_workbook(
            excelPath, read_only=True, data_only=True
        )
        try:
            if "Data" not in workbook.sheetnames:
                raise ValueError(
                    f"The Excel file has no 'Data' sheet: {excelPath}"
                )
            worksheet = workbook["Data"]
            header = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
            ]
            records = []
            for values in worksheet.iter_rows(min_row=2, values_only=True):
                record = dict(zip(header, values))
                if not any(value not in (None, "") for value in values):
                    continue
                status = str(record.get("status", "done")).strip().lower()
                if status != "done":
                    continue
                records.append(record)
            return records
        finally:
            workbook.close()

    def visualizationSegments(self, records):
        return sorted({
            str(record.get("segment", "")).strip()
            for record in records
            if str(record.get("segment", "")).strip()
        })

    def _finiteFloat(self, value):
        if value is None or value == "" or isinstance(value, bool):
            return None
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        return number if np.isfinite(number) else None

    def visualizationMetricKeys(self, records, segment):
        baseMetrics = [
            "suv_mean", "suv_max", "suv_peak", "tlg", "volume_mL"
        ]
        filtered = [
            record for record in records
            if str(record.get("segment", "")).strip() == str(segment).strip()
        ]
        if not filtered:
            return []

        allKeys = []
        for record in filtered:
            for key in record.keys():
                if key not in allKeys:
                    allKeys.append(key)

        candidates = [
            key for key in allKeys
            if key in baseMetrics
            or str(key).startswith("rad_")
            or str(key).startswith("derived_")
        ]

        available = []
        for key in candidates:
            if any(self._finiteFloat(record.get(key)) is not None
                   for record in filtered):
                available.append(key)

        ordered = [key for key in baseMetrics if key in available]
        ordered.extend(sorted(
            key for key in available if key not in baseMetrics
        ))
        return ordered

    def metricDisplayName(self, metricKey):
        displayNames = {
            "suv_mean": "SUVmean (Mean Standardized Uptake Value)",
            "suv_max": "SUVmax (Maximum Standardized Uptake Value)",
            "suv_peak": "SUVpeak (Peak Standardized Uptake Value)",
            "tlg": "TLG (Total Lesion Glycolysis)",
            "volume_mL": "Volume (mL)",
            "rad_firstorder_10Percentile": "P10 (10th Percentile)",
            "rad_firstorder_90Percentile": "P90 (90th Percentile)",
            "rad_firstorder_Entropy": "First-Order Entropy",
            "rad_firstorder_Skewness": "First-Order Skewness",
            "rad_glcm_Contrast": (
                "GLCM Contrast (Gray Level Co-occurrence Matrix Contrast)"
            ),
            # Retain the legacy display label so previously exported workbooks
            # containing IDMN can still be visualized correctly.
            "rad_glcm_Idmn": (
                "IDMN (Gray Level Co-occurrence Matrix Inverse "
                "Difference Moment Normalized)"
            ),
            "rad_glszm_SmallAreaHighGrayLevelEmphasis": (
                "SAHGLE (Gray Level Size Zone Matrix Small Area "
                "High Gray Level Emphasis)"
            ),
            "rad_glszm_LargeAreaLowGrayLevelEmphasis": (
                "LALGLE (Gray Level Size Zone Matrix Large Area "
                "Low Gray Level Emphasis)"
            ),
            "rad_glszm_ZoneEntropy": (
                "Zone Entropy (Gray Level Size Zone Matrix)"
            ),
        }
        return displayNames.get(metricKey, str(metricKey))

    def metricObservations(self, records, segment, metricKey):
        observations = []
        for record in records:
            if str(record.get("segment", "")).strip() != str(segment).strip():
                continue
            value = self._finiteFloat(record.get(metricKey))
            if value is None:
                continue
            observations.append({
                "subject_id": str(record.get("subject_id", "") or ""),
                "patient_id": str(record.get("patient_id", "") or ""),
                "scan_date": str(record.get("scan_date", "") or ""),
                "segment": str(segment),
                "metric": str(metricKey),
                "value": value,
            })
        return observations

    def computeMetricStatistics(self, observations):
        if not observations:
            raise ValueError("No valid observations were provided.")

        values = np.asarray(
            [observation["value"] for observation in observations], dtype=float
        )
        q1, median, q3 = np.percentile(values, [25, 50, 75])
        iqr = float(q3 - q1)
        lowerFence = float(q1 - 1.5 * iqr)
        upperFence = float(q3 + 1.5 * iqr)
        outlierMask = (values < lowerFence) | (values > upperFence)

        top3 = sorted(
            observations, key=lambda item: item["value"], reverse=True
        )[:3]

        return {
            "n": int(values.size),
            "mean": float(np.mean(values)),
            "std": (
                float(np.std(values, ddof=1)) if values.size > 1 else None
            ),
            "median": float(median),
            "q1": float(q1),
            "q3": float(q3),
            "iqr": iqr,
            "lower_fence": lowerFence,
            "upper_fence": upperFence,
            "outlier_count": int(np.count_nonzero(outlierMask)),
            "top3": top3,
        }

   # ----- Excel output related functions -----
    def saveMetricRowsToExcel(self, rows, outputFile, append=True):
        if not rows:
            raise ValueError("No metric rows to save.")
        if not outputFile:
            raise ValueError("No output Excel file selected.")
        
        outputFile = os.path.abspath(outputFile)

        if not outputFile.lower().endswith(".xlsx"):
            outputFile += ".xlsx"

        try:
            import openpyxl
        except ModuleNotFoundError:
            slicer.util.pip_install("openpyxl")
            import importlib
            openpyxl = importlib.import_module("openpyxl")

        if append and os.path.exists(outputFile):
            wb = openpyxl.load_workbook(outputFile)
        else:
            wb = openpyxl.Workbook()
        
        if "Data" in wb.sheetnames:
            ws = wb["Data"]
            existingHeader = [
                ws.cell(row=1, column=c).value
                for c in range(1, ws.max_column + 1)
            ]

            existingHeader = [
                str(h).strip()
                for h in existingHeader
                if h is not None and str(h).strip()
            ]
        else:
            if wb.active and wb.active.title == "Sheet":
                ws = wb.active
                ws.title = "Data"
            else:
                ws = wb.create_sheet("Data", 0)
            existingHeader = []
        
        newCols = []
        for row in rows:
            if not isinstance(row, dict):
                raise TypeError(
                    f"Each metric row must be a dict, but got {type(row).__name__}: {row}"
                )

            for key in row.keys():
                if key not in newCols:
                    newCols.append(key)
        
        columns = list(dict.fromkeys(existingHeader + newCols))

        for ci, col in enumerate(columns, start=1):
            ws.cell(row=1, column=ci, value=col)
        
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
            
        wb.save(outputFile)
        return outputFile

    # ----- QC / Hotspot related functions -----
    def computeQcRows(self, petNode, qcRows):
        """For each QC row, export its segment to a PET-geometry labelmap, then
        compute SUVmean/max/peak and the RAS location of the hottest voxel.

        SUVpeak here is a lightweight 1 mL-style local average around the max voxel
        (mean over the 26-neighbourhood). It is meant for QC ranking, not as a
        replacement for the PET-IndiC SUVpeak in section 3."""
        if petNode is None:
            raise ValueError("No PET volume selected.")
        if not petNode.IsA("vtkMRMLScalarVolumeNode"):
            raise ValueError(
                f"PET node must be a scalar volume node, but got: {petNode.GetClassName()}"
            )
        if petNode.GetImageData() is None:
            raise ValueError(f"PET node has no image data: {petNode.GetName()}")

        petArr = slicer.util.arrayFromVolume(petNode)

        ijkToRas = vtk.vtkMatrix4x4()
        petNode.GetIJKToRASMatrix(ijkToRas)

        for row in qcRows:
            roiNode = slicer.mrmlScene.GetNodeByID(row["roiNodeId"])
            if roiNode is None:
                print(f"[PETBiomarkerStudio] QC: ROI node gone: {row['roiNodeName']}")
                continue

            labelNode = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLLabelMapVolumeNode",
                "temp_qc_segment_mask",
            )
            try:
                segmentIds = vtk.vtkStringArray()
                if row.get("wholeRoi"):
                    seg = roiNode.GetSegmentation()
                    for i in range(seg.GetNumberOfSegments()):
                        segmentIds.InsertNextValue(seg.GetNthSegmentID(i))
                else:
                    segmentIds.InsertNextValue(row["segmentId"])

                slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                    roiNode,
                    segmentIds,
                    labelNode,
                    petNode,
                    slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
                )

                maskArr = slicer.util.arrayFromVolume(labelNode)
                if maskArr.shape != petArr.shape:
                    print(
                        f"[PETBiomarkerStudio] QC shape mismatch for "
                        f"{row['segmentName']}: mask {maskArr.shape} vs PET {petArr.shape}"
                    )
                    continue

                stats = _lib_qc_suv_stats(petArr, maskArr)
                if stats["max_ijk"] is None:
                    print(f"[PETBiomarkerStudio] QC: empty mask for {row['segmentName']}")
                    row["suv_mean"] = row["suv_max"] = row["suv_peak"] = None
                    row["ratio"] = None
                    row["rasHotspot"] = None
                    continue

                kz, ky, kx = stats["max_ijk"]
                ijk = [int(kx), int(ky), int(kz), 1.0]
                ras = ijkToRas.MultiplyPoint(ijk)

                row["suv_mean"] = stats["suv_mean"]
                row["suv_max"] = stats["suv_max"]
                row["suv_peak"] = stats["suv_peak"]
                row["ratio"] = stats["ratio"]
                row["rasHotspot"] = [float(ras[0]), float(ras[1]), float(ras[2])]

            finally:
                slicer.mrmlScene.RemoveNode(labelNode)

        return qcRows

    def flagQcOutliers(self, qcRows, madK=3.5, ratioThresh=4.0):
        """Flag a row when its SUVmax is a robust outlier across the analysed rows
        (median + k*MAD) OR when SUVmax/SUVmean exceeds ratioThresh."""
        return _lib_flag_qc_outliers(qcRows, mad_k=madK, ratio_thresh=ratioThresh)

    def jumpToHotspot(self, ras, label="hotspot", placeFiducial=True,
                      isOutlier=False, jump=True):
        """Centre slice views on a RAS point and (optionally) drop a colour-coded
        fiducial. Red marker = flagged outlier, green = normal."""
        r, a, s = float(ras[0]), float(ras[1]), float(ras[2])

        if jump:
            try:
                slicer.util.jumpSlices(
                    r, a, s,
                    sliceWidget=None,
                    jumpSliceMode=slicer.util.JumpSlice.Centered,
                    crosshairJumpMode=slicer.util.CrosshairJumpMode.Centered,
                )
            except AttributeError:
                try:
                    slicer.modules.markups.logic().JumpSlicesToLocation(r, a, s, True)
                except Exception:
                    slicer.util.warningDisplay(
                        "Could not jump slices automatically; use the reported RAS."
                    )

        if not placeFiducial:
            return

        name = slicer.mrmlScene.GenerateUniqueName(
            "QC_outlier" if isOutlier else "QC_hotspot"
        )
        fid = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLMarkupsFiducialNode", name)
        fid.CreateDefaultDisplayNodes()
        disp = fid.GetDisplayNode()
        if disp:
            if isOutlier:
                disp.SetSelectedColor(1, 0, 0)  # red
            else:
                disp.SetSelectedColor(0, 1, 0)  # green
            disp.SetGlyphScale(5.0)
            disp.SetTextScale(4.0)

        try:
            fid.AddControlPoint(vtk.vtkVector3d(r, a, s))
        except AttributeError:
            fid.AddControlPoint(r, a, s)

        try:
            n = fid.GetNumberOfControlPoints()
            if n > 0:
                fid.SetNthControlPointLabel(n - 1, label)
        except AttributeError:
            pass

    def saveQcRowsToExcel(self, qcRows, outputFile, baselineSuvMax=None):
        """Write the QC table to a 'QC' sheet (append/overwrite that one sheet only)."""
        return _lib_save_qc_rows_to_excel(
            qcRows, outputFile, baseline_suv_max=baselineSuvMax
        )

    # ----- Batch (cohort quantification) related functions -----
    AUXILIARY_STEM_TOKENS = _LIB_AUXILIARY_STEM_TOKENS

    def isAuxiliarySegmentStem(self, stem):
        """Files that are usually masks/aux rather than target organs -> unticked
        by default in the batch table (the user can still enable them)."""
        return _lib_is_auxiliary_segment_stem(stem)

    def defaultExcelLabel(self, stem):
        return _lib_default_excel_label(stem)

    def _parseBatchBaseName(self, base):
        """'MSP0001_2025-07-09' -> ('MSP0001', '2025-07-09').
        Falls back to (base, '') when no date is present."""
        return _lib_parse_batch_base_name(base)

    def scanBatchDataset(self, root):
        """Discover subjects (Segments/<base>_Seg) and the candidate segment files
        in the first subject. Returns a summary dict for the UI."""
        return _lib_scan_batch_dataset(root)

    def _findBatchSegmentFile(self, segDir, stem):
        return _lib_find_batch_segment_file(segDir, stem)

    def _loadBatchPet(self, root, base):
        """Load PET for a subject. Prefer a DICOM series folder, fall back to a
        NIfTI file. Returns a scalar volume node or None."""
        petRoot = os.path.join(root, "PET")
        dicomDir = os.path.join(petRoot, base + "_PET")
        if os.path.isdir(dicomDir):
            return self.loadDicomDirAsVolumeNode(
                dicomDir=dicomDir,
                modalityPreference=("PT", "NM"),
                preferSUV=True,
                roleName="PET",
                subjectId=base,
            )
        for ext in (".nii.gz", ".nii"):
            niftiPath = os.path.join(petRoot, base + "_PET" + ext)
            if os.path.exists(niftiPath):
                return self.loadVolumeNodeFromFile(niftiPath)
        return None

    def _existingBatchKeys(self, outputFile, requiredSegments=None,
                           requiredSignature=None):
        """Return subjects that are complete for this exact computation."""
        return _lib_existing_batch_keys(
            outputFile,
            required_segments=requiredSegments,
            required_signature=requiredSignature,
        )

    def runBatchQuantification(self, root, segmentSelections, metricsOptions,
                               radiomicsOptions, outputFile, append=True,
                               skipDone=True, progressCb=None, shouldCancel=None):
        """Iterate subjects, quantify the selected segment files, write Data +
        Summary sheets. Computation reuses runPetIndic / runRadiomics so single
        and batch results are identical.

        segmentSelections : list of (stem, excelLabel)
        """
        import datetime

        scan = self.scanBatchDataset(root)
        subjectDirs = scan["subjectDirs"]
        segRoot = os.path.join(root, "Segments")

        if not outputFile.lower().endswith(".xlsx"):
            outputFile += ".xlsx"

        computationSignature = self.computationSignature(
            metricsOptions, radiomicsOptions
        )
        requestedLabels = [label for _, label in segmentSelections]
        doneKeys = (
            self._existingBatchKeys(
                outputFile,
                requiredSegments=requestedLabels,
                requiredSignature=computationSignature,
            )
            if (append and skipDone)
            else set()
        )

        runRadiomics = self.isRadiomicsEnabled(radiomicsOptions)
        allRows = []
        processed = skipped = errors = 0
        total = len(subjectDirs)

        for idx, segFolder in enumerate(subjectDirs):
            if shouldCancel and shouldCancel():
                print("[PETBiomarkerStudio] Batch cancelled by user.")
                break

            base = segFolder[: -len("_Seg")]
            subjectId, scanDate = self._parseBatchBaseName(base)

            if progressCb:
                progressCb(idx, total, subjectId)

            if skipDone and (subjectId, scanDate) in doneKeys:
                print(f"[PETBiomarkerStudio] [{idx+1}/{total}] {subjectId}: skip (done).")
                skipped += 1
                continue

            print(f"\n{'='*60}\n[{idx+1}/{total}] {subjectId} ({scanDate})\n{'='*60}")
            slicer.mrmlScene.Clear(0)

            segDir = os.path.join(segRoot, segFolder)
            try:
                petNode = self._loadBatchPet(root, base)
                if petNode is None:
                    raise RuntimeError(f"No PET (DICOM folder or NIfTI) for {base}")

                for stem, label in segmentSelections:
                    segPath = self._findBatchSegmentFile(segDir, stem)
                    if segPath is None:
                        print(f"  [SEG] '{stem}' missing — error row.")
                        allRows.append(self._batchErrorRow(
                            subjectId, scanDate, label, "missing_file", stem))
                        continue

                    try:
                        roiNode = self.loadSegmentationNodeFromFile(segPath)
                    except Exception as e:
                        print(f"  [SEG] load failed '{stem}': {e}")
                        allRows.append(self._batchErrorRow(
                            subjectId, scanDate, label, f"load_error:{str(e)[:80]}", stem))
                        continue

                    try:
                        results = self.runPetIndic(
                            petNode=petNode, roiNode=roiNode,
                            metricsOptions=metricsOptions)
                        radStatus = "not_run"
                        if runRadiomics:
                            try:
                                radResults = self.runRadiomics(
                                    petNode=petNode, roiNode=roiNode,
                                    radiomicsOptions=radiomicsOptions)
                                results.update(radResults)
                                radStatus = "done"
                            except Exception as re_:
                                radStatus = f"radiomics_error:{str(re_)[:80]}"
                                print(f"  [RAD] failed: {re_}")

                        row = {
                            "subject_id": subjectId,
                            "patient_id": "",
                            "scan_date": scanDate,
                            "segment": label,
                            "source_file": os.path.basename(segPath),
                            "radiomics_status": radStatus,
                            "computation_signature": computationSignature,
                            "status": "done",
                            "computed_at":
                                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        row.update(results)
                        allRows.append(row)
                        print(f"  [OK] {label}: SUVmax="
                              f"{results.get('suv_max', 'NA')}")
                    except Exception as e:
                        print(f"  [METRIC] failed '{stem}': {e}")
                        allRows.append(self._batchErrorRow(
                            subjectId, scanDate, label, f"metric_error:{str(e)[:80]}", stem))
                    finally:
                        slicer.mrmlScene.RemoveNode(roiNode)

                processed += 1

            except Exception as e:
                import traceback
                print(traceback.format_exc())
                errors += 1
                allRows.append(self._batchErrorRow(
                    subjectId, scanDate, "", f"subject_error:{str(e)[:80]}", ""))

        if progressCb:
            progressCb(total, total, "done")

        savedPath = None
        if allRows:
            savedPath = self.saveBatchRowsToExcel(allRows, outputFile, append=append)

        return {
            "processed": processed,
            "skipped": skipped,
            "errors": errors,
            "rowCount": len(allRows),
            "savedPath": savedPath,
        }

    def _batchErrorRow(self, subjectId, scanDate, segment, status, stem):
        return _lib_batch_error_row(subjectId, scanDate, segment, status, stem)

    def _normaliseSegmentLabel(self, label):
        return re.sub(r"[^a-z0-9]+", "", str(label).lower())

    def _segmentRole(self, label):
        return _lib_segment_role(label)

    def _safeAsymmetry(self, left, right):
        return _lib_safe_asymmetry(left, right)

    def _calculateCrossRoiDerivedBySubject(self, allData):
        return _lib_cross_roi_derived_by_subject(allData)

    def saveBatchRowsToExcel(self, rows, outputFile, append=True):
        """Write Data + Summary sheets (radiomics-aware schema merge)."""
        return _lib_save_batch_rows_to_excel(rows, outputFile, append=append)

    # ----- Batch PREPARE (bulk mask ablation -> *_processed.nii.gz) -----
    def _loadBatchCt(self, root, base):
        """Load CT for a subject (DICOM folder preferred, NIfTI fallback)."""
        ctRoot = os.path.join(root, "CT")
        dicomDir = os.path.join(ctRoot, base + "_CT")
        if os.path.isdir(dicomDir):
            return self.loadDicomDirAsVolumeNode(
                dicomDir=dicomDir,
                modalityPreference=("CT",),
                preferSUV=False,
                roleName="CT",
                subjectId=base,
            )
        for ext in (".nii.gz", ".nii"):
            niftiPath = os.path.join(ctRoot, base + "_CT" + ext)
            if os.path.exists(niftiPath):
                return self.loadVolumeNodeFromFile(niftiPath)
        return None

    def _findAuxFile(self, segDir, filename):
        """Resolve an auxiliary mask filename inside a subject's _Seg folder.
        Accepts a name with or without extension."""
        if not filename:
            return None
        direct = os.path.join(segDir, filename)
        if os.path.exists(direct):
            return direct
        for ext in (".seg.nrrd", ".nii.gz", ".nii", ".nrrd"):
            cand = os.path.join(segDir, filename + ext)
            if os.path.exists(cand):
                return cand
        return None

    def _resolveAuxSeg(self, root, base, segDir, filename, generate, kind):
        """Return a segmentation node for an auxiliary structure (vertebrae or
        kidney): from a file in the _Seg folder, or generated from CT via
        TotalSegmentator when missing and 'generate' is enabled."""
        path = self._findAuxFile(segDir, filename)
        if path is not None:
            return self.loadSegmentationNodeFromFile(path)

        if generate:
            ctNode = self._loadBatchCt(root, base)
            if ctNode is None:
                raise RuntimeError(f"CT not found to generate {kind} for {base}")
            if kind == "vertebrae":
                node, _ = self.generateVertebraeFromCTNode(ctNode)
            else:
                node, _ = self.generateKidneysFromCTNode(ctNode)
            return node

        raise RuntimeError(
            f"{kind} mask not found for {base} "
            f"(file '{filename or '<blank>'}' missing in _Seg and generation disabled)"
        )

    def _exportSegmentationNodeToNifti(self, segNode, outPath):
        """Export all segments of a segmentation node to a labelmap and save it
        as a NIfTI file."""
        os.makedirs(os.path.dirname(outPath), exist_ok=True)
        labelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "temp_prep_export"
        )
        try:
            ok = slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(
                segNode, labelNode
            )
            if not ok:
                raise RuntimeError("ExportAllSegmentsToLabelmapNode returned False.")
            if not slicer.util.saveNode(labelNode, outPath):
                raise RuntimeError(f"saveNode failed for: {outPath}")
        finally:
            slicer.mrmlScene.RemoveNode(labelNode)

    def _prepareOutputsExist(self, segDir, organStems, suffix):
        """True only if every selected organ already has its <stem><suffix>.nii.gz."""
        for stem in organStems:
            out = os.path.join(segDir, stem + suffix + ".nii.gz")
            if not os.path.exists(out):
                return False
        return True

    def runBatchPrepare(self, root, organStems, steps, auxConfig, params,
                        outputSuffix="_processed", skipDone=True,
                        progressCb=None, shouldCancel=None):
        """Bulk ROI ablation. For each subject, apply the selected ablation steps
        to each organ mask and write <stem><suffix>.nii.gz into the subject's
        _Seg folder. Reuses the Interactive tab's ablation methods verbatim, so
        single / batch preprocessing cannot diverge.

        steps      : dict(spine, kidney, urinary) -> bool
        auxConfig  : dict(vertebraeFile, kidneyFile, generateIfMissing)
        params     : dict(spineLevels, kidneyDilationMm, ureterSuvThresh,
                          ureterDilationMm, cleanSuvThresh)
        PET geometry is used as the reference volume for every step, so the
        processed masks live on the PET grid the Quantify stage computes on.
        """
        scan = self.scanBatchDataset(root)
        subjectDirs = scan["subjectDirs"]
        segRoot = os.path.join(root, "Segments")

        needPet = any(steps.values())                       # PET = reference + SUV source
        needVertebrae = steps.get("spine") or steps.get("urinary")
        needKidney = steps.get("kidney")

        processed = skipped = errors = 0
        total = len(subjectDirs)

        for idx, segFolder in enumerate(subjectDirs):
            if shouldCancel and shouldCancel():
                print("[PETBiomarkerStudio] Prepare cancelled by user.")
                break

            base = segFolder[: -len("_Seg")]
            subjectId, scanDate = self._parseBatchBaseName(base)
            if progressCb:
                progressCb(idx, total, subjectId)

            segDir = os.path.join(segRoot, segFolder)

            if skipDone and self._prepareOutputsExist(segDir, organStems, outputSuffix):
                print(f"[PETBiomarkerStudio] [{idx+1}/{total}] {subjectId}: skip (outputs exist).")
                skipped += 1
                continue

            print(f"\n{'='*60}\n[PREP {idx+1}/{total}] {subjectId} ({scanDate})\n{'='*60}")
            slicer.mrmlScene.Clear(0)

            try:
                petNode = None
                if needPet:
                    petNode = self._loadBatchPet(root, base)
                    if petNode is None:
                        raise RuntimeError(f"No PET for {base} (required for ablation)")
                referenceNode = petNode

                vertebraeNode = None
                if needVertebrae:
                    vertebraeNode = self._resolveAuxSeg(
                        root, base, segDir, auxConfig.get("vertebraeFile"),
                        generate=auxConfig.get("generateIfMissing"), kind="vertebrae")

                kidneyNode = None
                if needKidney:
                    kidneyNode = self._resolveAuxSeg(
                        root, base, segDir, auxConfig.get("kidneyFile"),
                        generate=auxConfig.get("generateIfMissing"), kind="kidney")

                urinaryMaskNode = None
                if steps.get("urinary"):
                    urinaryMaskNode, _ = self.generatePETDerivedUreterMask(
                        petNode=petNode,
                        vertebraeNode=vertebraeNode,
                        selectedLevels=params["spineLevels"],
                        suvThreshold=params["ureterSuvThresh"],
                        dilationRadiusMm=params["ureterDilationMm"],
                    )

                anyOrgan = False
                for stem in organStems:
                    organPath = self._findBatchSegmentFile(segDir, stem)
                    if organPath is None:
                        print(f"  [PREP] '{stem}' missing — skip.")
                        continue

                    chain = []
                    try:
                        node = self.loadSegmentationNodeFromFile(organPath)
                        chain.append(node)
                        cur = node

                        if steps.get("spine"):
                            cur, _ = self.applySpineRangeSlicing(
                                cur, vertebraeNode, params["spineLevels"], referenceNode)
                            chain.append(cur)
                        if steps.get("kidney"):
                            cur, _ = self.applyKidneyExclusion(
                                roiNode=cur, kidneyNode=kidneyNode,
                                referenceVolumeNode=referenceNode,
                                dilationRadiusMm=params["kidneyDilationMm"])
                            chain.append(cur)
                        if steps.get("urinary"):
                            cur, _ = self.applyUrinaryActivityCleanup(
                                roiNode=cur, petNode=petNode,
                                urinaryMaskNode=urinaryMaskNode,
                                cleanSUVThreshold=params["cleanSuvThresh"])
                            chain.append(cur)

                        outPath = os.path.join(segDir, stem + outputSuffix + ".nii.gz")
                        self._exportSegmentationNodeToNifti(cur, outPath)
                        print(f"  [PREP] {stem} -> {os.path.basename(outPath)}")
                        anyOrgan = True
                    except Exception as e:
                        print(f"  [PREP] failed '{stem}': {e}")
                    finally:
                        # free the per-organ chain (aux/PET nodes persist for the subject)
                        for n in chain:
                            try:
                                slicer.mrmlScene.RemoveNode(n)
                            except Exception:
                                pass

                if anyOrgan:
                    processed += 1

            except Exception as e:
                import traceback
                print(traceback.format_exc())
                errors += 1

        if progressCb:
            progressCb(total, total, "done")

        return {"processed": processed, "skipped": skipped, "errors": errors}


#####################################################################
# Test codes
if __name__ == "__main__":
    pass
