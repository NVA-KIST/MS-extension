"""
Unit tests for lib.quantification.biomarker_batch (no Slicer / pyradiomics).

Run:
    python tests/test_biomarker_batch.py
"""
from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from lib.quantification.biomarker_batch import (
    AUXILIARY_STEM_TOKENS,
    computation_signature,
    default_excel_label,
    is_auxiliary_segment_stem,
    parse_batch_base_name,
    parse_quantitative_indices_results,
    safe_asymmetry,
    segment_role,
)


def test_parse_batch_base_name():
    assert parse_batch_base_name("MSP0001_2025-07-09") == ("MSP0001", "2025-07-09")
    assert parse_batch_base_name("no_date_here") == ("no_date_here", "")
    print("  PASS  parse_batch_base_name")


def test_segment_role():
    assert segment_role("aorta") == "blood_pool"
    assert segment_role("psoas_left") == "psoas_left"
    assert segment_role("psoas_right") == "psoas_right"
    assert segment_role("liver") == "other"
    print("  PASS  segment_role")


def test_safe_asymmetry():
    assert abs(safe_asymmetry(2.0, 4.0) - (2.0 / 3.0)) < 1e-9
    assert safe_asymmetry(None, 1.0) is None
    assert safe_asymmetry(0.0, 0.0) is None
    print("  PASS  safe_asymmetry")


def test_auxiliary_and_label():
    assert "kidney" in AUXILIARY_STEM_TOKENS
    assert is_auxiliary_segment_stem("kidney_left")
    assert not is_auxiliary_segment_stem("psoas_major")
    assert default_excel_label("liver.seg") == "liver"
    assert default_excel_label("visceral_fat_processed") == "visceral_fat"
    print("  PASS  auxiliary / default_excel_label")


def test_parse_qi_and_signature():
    parsed = parse_quantitative_indices_results([
        ("Mean", "1.5"),
        ("Max_s", "3.0"),
        ("Peak", "--"),
        ("Volume", "10"),
        ("Unknown", "9"),
    ])
    assert parsed["suv_mean"] == 1.5
    assert parsed["suv_max"] == 3.0
    assert "suv_peak" not in parsed
    assert parsed["volume_mL"] == 10.0

    sig = computation_signature(
        {"mean": True, "max": True, "peak": False, "tlg": False, "volume": True},
        {},
    )
    assert sig.startswith("metrics=mean,max,volume|")
    assert "radiomics=off" in sig
    print("  PASS  parse_quantitative_indices_results / computation_signature")


def test_excel_quant_radiomics_sheets():
    import tempfile
    from openpyxl import load_workbook
    from lib.quantification.biomarker_batch import (
        QUANTIFICATION_SHEET,
        RADIOMICS_SHEET,
        save_batch_rows_to_excel,
    )

    rows = [
        {
            "subject_id": "MSP0001",
            "patient_id": "",
            "scan_date": "2025-01-01",
            "segment": "spleen",
            "source_file": "spleen.nii.gz",
            "volume_mL": 100.0,
            "suv_mean": 1.2,
            "suv_max": 3.4,
            "suv_peak": 2.1,
            "tlg": 120.0,
            "rad_firstorder_Entropy": 0.5,
            "rad_glcm_Contrast": 0.1,
            "radiomics_status": "done",
            "computation_signature": "test",
            "status": "done",
        }
    ]
    with tempfile.TemporaryDirectory() as td:
        out = str(Path(td) / "metrics.xlsx")
        save_batch_rows_to_excel(rows, out, append=False)
        wb = load_workbook(out)
        assert QUANTIFICATION_SHEET in wb.sheetnames
        assert RADIOMICS_SHEET in wb.sheetnames
        assert "Summary" in wb.sheetnames
        q_header = [c.value for c in wb[QUANTIFICATION_SHEET][1]]
        assert "suv_max" in q_header
        assert "rad_firstorder_Entropy" not in q_header
        r_header = [c.value for c in wb[RADIOMICS_SHEET][1]]
        assert "rad_firstorder_Entropy" in r_header
        assert "suv_max" not in r_header
    print("  PASS  excel Quantification + Radiomics sheets")


def main():
    print("test_biomarker_batch")
    test_parse_batch_base_name()
    test_segment_role()
    test_safe_asymmetry()
    test_auxiliary_and_label()
    test_parse_qi_and_signature()
    test_excel_quant_radiomics_sheets()
    print("ALL PASSED")


if __name__ == "__main__":
    main()
