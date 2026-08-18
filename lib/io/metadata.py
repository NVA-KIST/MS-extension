"""
Extract patient / scan metadata from DICOM headers.

Works on a single file, a series folder, or a full organized dataset
(``ROOT/CT`` + ``ROOT/PET``).
"""
from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any, Optional, Sequence

from lib.io.dicom_utils import (
    ds_get,
    first_dicom_header,
    fmt_study_date,
    parse_age_years,
    radiopharma_fields,
    read_dicom_header,
    to_float,
)
from lib.io.paths import discover_patients


def extract_dicom_metadata(path: str | Path) -> dict[str, Any]:
    """
    Read demographics + scan tags from one DICOM file or series folder.

    Prefers the first readable header. PET series usually carry weight / dose.

    Returns a flat dict (empty-ish if unreadable). Always includes ``source_path``.
    """
    path = Path(path)
    if path.is_file():
        ds = read_dicom_header(path)
        src = path
    else:
        ds, src = first_dicom_header(path)
    if ds is None:
        return {"source_path": str(path), "readable": False}

    weight = to_float(ds_get(ds, "PatientWeight", None))
    height = to_float(ds_get(ds, "PatientSize", None))
    # PatientSize is metres in DICOM; some sites store cm
    if height is not None and height > 3.0:
        height = height / 100.0
    bmi = round(weight / (height ** 2), 1) if (weight and height and height > 0) else None

    pixel_spacing = ds_get(ds, "PixelSpacing", None)
    if pixel_spacing is not None:
        try:
            pixel_spacing = "\\".join(str(x) for x in list(pixel_spacing))
        except TypeError:
            pixel_spacing = str(pixel_spacing)

    birth = str(ds_get(ds, "PatientBirthDate"))
    study_date_raw = str(ds_get(ds, "StudyDate") or ds_get(ds, "SeriesDate") or "")

    row: dict[str, Any] = {
        "readable": True,
        "source_path": str(src),
        "patient_id": str(ds_get(ds, "PatientID")),
        "patient_name": str(ds_get(ds, "PatientName")),
        "patient_sex": str(ds_get(ds, "PatientSex")),
        "patient_age_y": parse_age_years(ds_get(ds, "PatientAge", None)),
        "patient_birth_date": fmt_study_date(birth) if birth else "",
        "weight_kg": weight,
        "height_m": height,
        "bmi": bmi,
        "study_date": fmt_study_date(study_date_raw),
        "study_date_raw": study_date_raw,
        "study_time": str(ds_get(ds, "StudyTime")).split(".")[0],
        "acquisition_date": str(ds_get(ds, "AcquisitionDate")),
        "acquisition_time": str(ds_get(ds, "AcquisitionTime")).split(".")[0],
        "accession_number": str(ds_get(ds, "AccessionNumber")),
        "study_uid": str(ds_get(ds, "StudyInstanceUID")),
        "series_uid": str(ds_get(ds, "SeriesInstanceUID")),
        "study_description": str(ds_get(ds, "StudyDescription")),
        "series_description": str(ds_get(ds, "SeriesDescription")),
        "modality": str(ds_get(ds, "Modality")),
        "manufacturer": str(ds_get(ds, "Manufacturer")),
        "model": str(ds_get(ds, "ManufacturerModelName")),
        "station": str(ds_get(ds, "StationName")),
        "institution": str(ds_get(ds, "InstitutionName")),
        "rows": to_float(ds_get(ds, "Rows", None)),
        "columns": to_float(ds_get(ds, "Columns", None)),
        "slice_thickness_mm": to_float(ds_get(ds, "SliceThickness", None)),
        "pixel_spacing": pixel_spacing or "",
        "kvp": to_float(ds_get(ds, "KVP", None)),
        "image_comments": str(ds_get(ds, "ImageComments")),
    }
    row.update(radiopharma_fields(ds))
    return row


def extract_dataset_metadata(
    root: str | Path,
    *,
    prefer_pet: bool = True,
    log=None,
) -> list[dict[str, Any]]:
    """
    Extract one metadata row per scan under an organized dataset root.

    Uses ``discover_patients`` (``ROOT/CT`` + ``ROOT/PET``). PET headers are
    preferred because they carry weight, height, and radiopharmaceutical tags.
    """
    root = Path(root)
    patients = discover_patients(str(root))
    rows: list[dict[str, Any]] = []

    def _info(msg: str):
        print(msg)
        if log and hasattr(log, "info"):
            log.info(msg)

    for pat in patients:
        pet = pat.get("pet_path")
        ct = pat.get("ct_path")
        source = None
        if prefer_pet and pet and os.path.isdir(pet):
            source = pet
            src_kind = "PET"
        elif ct and os.path.isdir(ct):
            source = ct
            src_kind = "CT"
        elif pet and os.path.isdir(pet):
            source = pet
            src_kind = "PET"
        else:
            _info(f"  [skip] no DICOM folder for {pat['subject_id']} {pat['scan_date']}")
            continue

        meta = extract_dicom_metadata(source)
        meta["subject_id"] = pat["subject_id"]
        meta["scan_date"] = pat["scan_date"] or meta.get("study_date")
        meta["meta_source"] = src_kind
        meta["has_CT"] = bool(ct and os.path.isdir(ct))
        meta["has_PET"] = bool(pet and os.path.isdir(pet))
        meta["ct_path"] = ct
        meta["pet_path"] = pet
        rows.append(meta)
        _info(
            f"  {meta['subject_id']} {meta['scan_date']}  "
            f"sex={meta.get('patient_sex')} age={meta.get('patient_age_y')}  "
            f"wt={meta.get('weight_kg')} kg  ht={meta.get('height_m')} m"
        )

    # Longitudinal weight change vs previous scan of same subject
    by_subj: dict[str, list[dict]] = {}
    for r in rows:
        by_subj.setdefault(str(r.get("subject_id")), []).append(r)
    for subj, group in by_subj.items():
        group.sort(key=lambda r: str(r.get("scan_date") or ""))
        baseline_w = group[0].get("weight_kg")
        prev_date = None
        prev_w = None
        for i, r in enumerate(group):
            r["n_scans_for_patient"] = len(group)
            r["scan_number_for_subject"] = i + 1
            w = r.get("weight_kg")
            if i == 0:
                r["weight_delta_prev_kg"] = 0.0
                r["days_since_prev"] = 0.0
                r["weight_delta_baseline_kg"] = 0.0
                r["weight_pct_change_baseline"] = 0.0
            else:
                r["weight_delta_prev_kg"] = (
                    round(w - prev_w, 1) if (w is not None and prev_w is not None) else None
                )
                r["days_since_prev"] = _days_between(prev_date, r.get("scan_date"))
                r["weight_delta_baseline_kg"] = (
                    round(w - baseline_w, 1)
                    if (w is not None and baseline_w is not None)
                    else None
                )
                if w is not None and baseline_w:
                    r["weight_pct_change_baseline"] = round(
                        100.0 * (w - baseline_w) / baseline_w, 1
                    )
                else:
                    r["weight_pct_change_baseline"] = None
            prev_w = w
            prev_date = r.get("scan_date")

    return rows


def _days_between(a, b) -> Optional[float]:
    from datetime import datetime

    try:
        da = datetime.strptime(str(a)[:10], "%Y-%m-%d")
        db = datetime.strptime(str(b)[:10], "%Y-%m-%d")
        return float((db - da).days)
    except (TypeError, ValueError):
        return None


_PREFERRED_COLS = [
    "subject_id",
    "scan_date",
    "meta_source",
    "has_CT",
    "has_PET",
    "patient_id",
    "patient_name",
    "patient_sex",
    "patient_age_y",
    "patient_birth_date",
    "weight_kg",
    "height_m",
    "bmi",
    "study_date",
    "study_time",
    "acquisition_date",
    "acquisition_time",
    "accession_number",
    "study_uid",
    "study_description",
    "series_description",
    "modality",
    "manufacturer",
    "model",
    "station",
    "institution",
    "rows",
    "columns",
    "slice_thickness_mm",
    "pixel_spacing",
    "kvp",
    "radiopharmaceutical",
    "injected_dose_MBq",
    "radionuclide_half_life_s",
    "radiopharm_start_time",
    "pet_units",
    "decay_correction",
    "scan_number_for_subject",
    "n_scans_for_patient",
    "weight_delta_prev_kg",
    "days_since_prev",
    "weight_delta_baseline_kg",
    "weight_pct_change_baseline",
    "ct_path",
    "pet_path",
    "source_path",
]


def save_metadata(
    rows: Sequence[dict],
    *,
    csv_path: str | Path | None = None,
    xlsx_path: str | Path | None = None,
    sheet_name: str = "Metadata",
) -> dict[str, str]:
    """Write metadata rows to CSV and/or Excel. Returns written paths."""
    written: dict[str, str] = {}
    if not rows:
        return written
    cols = [c for c in _PREFERRED_COLS if any(c in r for r in rows)]
    extra = []
    for r in rows:
        for k in r:
            if k not in cols and k not in extra:
                extra.append(k)
    cols.extend(extra)

    if csv_path:
        csv_path = Path(csv_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({k: r.get(k, "") for k in cols})
        written["csv"] = str(csv_path)

    if xlsx_path:
        from openpyxl import Workbook, load_workbook

        xlsx_path = Path(xlsx_path)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        if xlsx_path.is_file():
            wb = load_workbook(xlsx_path)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
        else:
            wb = Workbook()
            # drop default empty sheet if we are creating fresh
            if wb.sheetnames == ["Sheet"]:
                del wb["Sheet"]
        ws = wb.create_sheet(sheet_name, 0)
        ws.append(cols)
        for r in rows:
            ws.append([_excel_cell(r.get(k)) for k in cols])
        wb.save(xlsx_path)
        written["xlsx"] = str(xlsx_path)

    return written


def _excel_cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    return v
