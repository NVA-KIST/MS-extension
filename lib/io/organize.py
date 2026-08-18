"""
Organize raw PET/CT exports into the pipeline layout::

    OUT/
      CT/   {SubjectID}_{YYYY-MM-DD}_CT/    *.dcm
      PET/  {SubjectID}_{YYYY-MM-DD}_PET/   *.dcm

Accepted *input* layouts (auto-detected):

1. **MIM / PACS ``*__Studies`` export** (usual inbound)::
       2026-03__Studies/
         {uid}_CT_{YYYY-MM-DD}_{HHMMSS}_…/
         {uid}_PT_{YYYY-MM-DD}_{HHMMSS}_…
     or a parent folder that contains several ``*__Studies`` batches.

2. **Already organized pipeline** ``CT/`` + ``PET/`` with
   ``{ID}_{YYYY-MM-DD}_{CT|PET}`` folders — copied through (or no-op if dest=src).

3. **Generic DICOM tree** — any nested folders; series are grouped by
   ``StudyInstanceUID`` + modality from the DICOM header (folder names ignored).
"""
from __future__ import annotations

import csv
import os
import re
import shutil
from pathlib import Path
from typing import Any, Optional

from lib.io.dicom_utils import (
    count_files,
    ds_get,
    first_dicom_header,
    folder_has_dicom,
    fmt_study_date,
    normalize_modality,
)
from lib.io.metadata import extract_dataset_metadata, save_metadata
from lib.io.paths import parse_folder_name

SERIES_RE = re.compile(
    r"^.+_(CT|PT)_(\d{4}-\d{2}-\d{2})_\d{6}_.+$",
    re.IGNORECASE,
)
SUBJECT_PREFIX = "MSP"
SUBJECT_PAD = 4


# ── Layout detection ──────────────────────────────────────────────────────────

def detect_input_layout(src: str | Path) -> str:
    """
    Return ``\"studies\"`` | ``\"studies_parent\"`` | ``\"pipeline\"`` | ``\"dicom_tree\"``.
    """
    src = Path(src)
    if not src.is_dir():
        raise FileNotFoundError(f"Input folder not found: {src}")

    if (src / "CT").is_dir() or (src / "PET").is_dir():
        # pipeline if any child matches {id}_{date}_CT|PET
        for sub in ("CT", "PET"):
            d = src / sub
            if not d.is_dir():
                continue
            for child in d.iterdir():
                if child.is_dir() and parse_folder_name(child.name):
                    return "pipeline"

    if src.name.endswith("__Studies"):
        return "studies"

    studies_kids = [p for p in src.iterdir() if p.is_dir() and p.name.endswith("__Studies")]
    if studies_kids:
        return "studies_parent"

    # series folders directly in src (MIM without the parent suffix)
    n_series = 0
    for child in src.iterdir():
        if child.is_dir() and SERIES_RE.match(child.name):
            n_series += 1
            if n_series >= 2:
                return "studies"

    return "dicom_tree"


def _iter_studies_roots(src: Path, layout: str) -> list[Path]:
    if layout == "studies":
        return [src]
    if layout == "studies_parent":
        return sorted(p for p in src.iterdir() if p.is_dir() and p.name.endswith("__Studies"))
    return [src]


# ── Series discovery ──────────────────────────────────────────────────────────

def _series_from_folder(folder: Path, batch: str = "") -> Optional[dict[str, Any]]:
    ds, fp = first_dicom_header(folder)
    if ds is None:
        return None
    hint = None
    m = SERIES_RE.match(folder.name)
    if m:
        hint = m.group(1)
    modality = normalize_modality(str(ds_get(ds, "Modality")) or hint or "")
    if modality is None:
        return None
    study_uid = str(ds_get(ds, "StudyInstanceUID"))
    study_date = fmt_study_date(str(ds_get(ds, "StudyDate") or ds_get(ds, "SeriesDate") or ""))
    if m and not study_date:
        study_date = m.group(2)
    return {
        "path": folder,
        "modality": modality,
        "study_uid": study_uid,
        "study_date": study_date,
        "patient_id": str(ds_get(ds, "PatientID")),
        "patient_name": str(ds_get(ds, "PatientName")),
        "patient_sex": str(ds_get(ds, "PatientSex")),
        "accession_number": str(ds_get(ds, "AccessionNumber")),
        "study_description": str(ds_get(ds, "StudyDescription")),
        "n_files": count_files(folder),
        "batch": batch,
        "header_file": str(fp) if fp else "",
    }


def discover_raw_studies(src: str | Path, *, layout: Optional[str] = None) -> tuple[list[dict], list[str]]:
    """
    Discover CT/PET series and group them into studies.

    Returns ``(studies, warnings)``. Each study dict has ``series`` =
    ``{\"CT\": {...}, \"PET\": {...}}`` (either side may be missing).
    """
    src = Path(src)
    layout = layout or detect_input_layout(src)
    warnings: list[str] = []
    studies: dict[str, dict] = {}

    def _add(series: dict):
        uid = series["study_uid"] or f"NOUID_{series['path']}"
        rec = studies.setdefault(
            uid,
            {
                "study_uid": series["study_uid"],
                "patient_id": series["patient_id"],
                "patient_name": series["patient_name"],
                "patient_sex": series["patient_sex"],
                "study_date": series["study_date"],
                "accession_number": series["accession_number"],
                "study_description": series["study_description"],
                "batch": series.get("batch", ""),
                "series": {},
            },
        )
        mod = series["modality"]
        if mod in rec["series"]:
            warnings.append(
                f"[warn] study {uid} has extra {mod} series "
                f"({rec['series'][mod]['path']} vs {series['path']}) — keeping first"
            )
            return
        rec["series"][mod] = series
        # fill blanks from this series if first was empty
        for k in ("patient_id", "patient_name", "patient_sex", "study_date",
                  "accession_number", "study_description"):
            if not rec.get(k) and series.get(k):
                rec[k] = series[k]

    if layout in ("studies", "studies_parent"):
        for root in _iter_studies_roots(src, layout):
            batch = root.name
            for child in sorted(p for p in root.iterdir() if p.is_dir()):
                series = _series_from_folder(child, batch=batch)
                if series is None:
                    if folder_has_dicom(child):
                        warnings.append(f"[skip] unrecognized modality: {child}")
                    continue
                _add(series)
    elif layout == "pipeline":
        for kind, sub in (("CT", "CT"), ("PET", "PET")):
            d = src / sub
            if not d.is_dir():
                continue
            for child in sorted(p for p in d.iterdir() if p.is_dir()):
                series = _series_from_folder(child, batch=src.name)
                if series is None:
                    continue
                series["modality"] = kind
                parsed = parse_folder_name(child.name)
                if parsed and not series.get("study_date"):
                    series["study_date"] = parsed[1]
                _add(series)
    else:
        # Generic DICOM tree: every directory that itself contains DICOM files
        # (not only nested) is a series candidate.
        seen_dirs: set[Path] = set()
        for dirpath, dirnames, filenames in os.walk(src):
            folder = Path(dirpath)
            # skip organized dest-like names to avoid double-walking
            if folder.name in ("CT_NIfTI", "PET_NIfTI", "Segments", "pipeline_logs"):
                dirnames.clear()
                continue
            if not any(Path(dirpath, f).is_file() for f in filenames):
                continue
            if folder in seen_dirs:
                continue
            series = _series_from_folder(folder, batch=src.name)
            if series is None:
                continue
            seen_dirs.add(folder)
            _add(series)

    return list(studies.values()), warnings


# ── Subject ID registry ───────────────────────────────────────────────────────

def _load_subject_map(
    dest: Path,
    existing_map: Optional[Path] = None,
    prefix: str = SUBJECT_PREFIX,
) -> tuple[dict[str, str], int]:
    pid_to_subj: dict[str, str] = {}
    candidates = []
    if existing_map:
        candidates.append(Path(existing_map))
    candidates.append(dest / "patient_id_mapping.csv")
    for csv_path in candidates:
        if not csv_path.is_file():
            continue
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                pid = str(row.get("patient_id") or "").strip()
                subj = str(row.get("subject_id") or "").strip()
                if pid and subj:
                    pid_to_subj[pid] = subj
    max_n = 0
    pat = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.I)
    for subj in pid_to_subj.values():
        m = pat.match(str(subj))
        if m:
            max_n = max(max_n, int(m.group(1)))
    return pid_to_subj, max_n


def _next_subject_id(counter: list[int], prefix: str = SUBJECT_PREFIX, pad: int = SUBJECT_PAD) -> str:
    counter[0] += 1
    return f"{prefix}{counter[0]:0{pad}d}"


def _save_subject_map(dest: Path, pid_to_subj: dict[str, str]) -> Path:
    out = dest / "patient_id_mapping.csv"
    dest.mkdir(parents=True, exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["patient_id", "subject_id"])
        w.writeheader()
        for pid, subj in sorted(pid_to_subj.items(), key=lambda kv: kv[1]):
            w.writerow({"patient_id": pid, "subject_id": subj})
    return out


# ── Copy ──────────────────────────────────────────────────────────────────────

def _copy_series(src_folder: Path, dst_folder: Path, *, skip_existing: bool = True) -> str:
    dst_folder.mkdir(parents=True, exist_ok=True)
    src_files = [p for p in src_folder.iterdir() if p.is_file()]
    if skip_existing:
        existing = [p for p in dst_folder.iterdir() if p.is_file()]
        if existing and len(existing) >= len(src_files) > 0:
            return f"skipped ({len(existing)} files already present)"
    n = 0
    for p in src_files:
        shutil.copy2(p, dst_folder / p.name)
        n += 1
    return f"copied {n} files"


# ── Public API ────────────────────────────────────────────────────────────────

def organize_dataset(
    src: str | Path,
    dest: str | Path,
    *,
    existing_map: str | Path | None = None,
    subject_prefix: str = SUBJECT_PREFIX,
    skip_existing: bool = True,
    write_metadata: bool = True,
    log=None,
) -> dict[str, Any]:
    """
    Copy raw inbound scans into ``dest/CT`` and ``dest/PET``.

    Parameters
    ----------
    src:
        ``2026-03__Studies``, a parent of several ``*__Studies`` batches,
        an already-organized pipeline root, or any DICOM tree.
    dest:
        Pipeline dataset root (created if missing).
    existing_map:
        Optional CSV with columns ``patient_id,subject_id`` to reuse codes
        (also reads ``dest/patient_id_mapping.csv`` if present).
    write_metadata:
        After copy, extract DICOM demographics → ``dest/scan_metadata.csv``
        and ``dest/scan_mapping.xlsx``.

    Returns
    -------
    dict with ``layout``, ``n_studies``, ``mapping_rows``, ``warnings``,
    ``dest``, ``map_csv``, ``metadata_csv``, ``mapping_xlsx``.
    """
    src = Path(src)
    dest = Path(dest)
    dest.mkdir(parents=True, exist_ok=True)
    (dest / "CT").mkdir(exist_ok=True)
    (dest / "PET").mkdir(exist_ok=True)

    def _info(msg: str):
        print(msg)
        if log and hasattr(log, "info"):
            log.info(msg)

    layout = detect_input_layout(src)
    _info(f"[organize] layout={layout}  src={src}")
    studies, warnings = discover_raw_studies(src, layout=layout)
    _info(f"[organize] found {len(studies)} study/studies")

    pid_to_subj, max_n = _load_subject_map(dest, Path(existing_map) if existing_map else None, subject_prefix)
    known = set(pid_to_subj.keys())
    counter = [max_n]
    mapping_rows: list[dict] = []

    for rec in sorted(studies, key=lambda r: (str(r.get("study_date") or ""), str(r.get("patient_id") or ""))):
        pid = rec.get("patient_id") or "UNKNOWN"
        date = rec.get("study_date") or "unknown-date"
        is_new = pid not in known
        if pid not in pid_to_subj:
            pid_to_subj[pid] = _next_subject_id(counter, subject_prefix)
            known.add(pid)
        subj = pid_to_subj[pid]

        ct = rec["series"].get("CT")
        pet = rec["series"].get("PET")
        if not ct:
            warnings.append(f"[warn] {subj} {date}: missing CT")
        if not pet:
            warnings.append(f"[warn] {subj} {date}: missing PET")

        ct_status = pet_status = "n/a"
        ct_dst = pet_dst = ""
        if ct:
            ct_dst = dest / "CT" / f"{subj}_{date}_CT"
            ct_status = _copy_series(Path(ct["path"]), ct_dst, skip_existing=skip_existing)
            _info(f"  CT  {ct_dst.name}: {ct_status}")
        if pet:
            pet_dst = dest / "PET" / f"{subj}_{date}_PET"
            pet_status = _copy_series(Path(pet["path"]), pet_dst, skip_existing=skip_existing)
            _info(f"  PET {pet_dst.name}: {pet_status}")

        mapping_rows.append(
            {
                "subject_id": subj,
                "patient_id": pid,
                "patient_name": rec.get("patient_name"),
                "patient_sex": rec.get("patient_sex"),
                "is_new_patient": is_new,
                "scan_date": date,
                "accession_number": rec.get("accession_number"),
                "study_uid": rec.get("study_uid"),
                "study_description": rec.get("study_description"),
                "batch": rec.get("batch"),
                "has_CT": bool(ct),
                "has_PET": bool(pet),
                "ct_source_folder": str(ct["path"]) if ct else "",
                "pet_source_folder": str(pet["path"]) if pet else "",
                "ct_dest_folder": str(ct_dst) if ct else "",
                "pet_dest_folder": str(pet_dst) if pet else "",
                "ct_copy_status": ct_status,
                "pet_copy_status": pet_status,
            }
        )

    map_csv = _save_subject_map(dest, pid_to_subj)

    result: dict[str, Any] = {
        "layout": layout,
        "n_studies": len(mapping_rows),
        "mapping_rows": mapping_rows,
        "warnings": warnings,
        "dest": str(dest),
        "map_csv": str(map_csv),
        "metadata_csv": "",
        "mapping_xlsx": "",
    }

    xlsx = dest / "scan_mapping.xlsx"
    if write_metadata:
        meta_rows = extract_dataset_metadata(dest)
        csv_path = dest / "scan_metadata.csv"
        save_metadata(meta_rows, csv_path=csv_path, xlsx_path=xlsx, sheet_name="Metadata")
        result["metadata_csv"] = str(csv_path)
        result["mapping_xlsx"] = str(xlsx)
        # Mapping sheet alongside metadata
        _write_mapping_sheet(xlsx, mapping_rows)

    for w in warnings:
        _info(w)
    _info(f"[organize] done  studies={len(mapping_rows)}  dest={dest}")
    return result


def _write_mapping_sheet(xlsx_path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    name = "Mapping"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    wb.save(xlsx_path)
