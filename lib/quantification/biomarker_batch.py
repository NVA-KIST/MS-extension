"""
Pure batch Excel / derived-metric helpers from PETBiomarkerStudioLogic.

No Slicer imports. openpyxl is required for Excel I/O (ImportError if missing).
"""
from __future__ import annotations

import datetime
import math
import os
import re
from typing import Any, Dict, Iterable, Optional, Sequence, Set, Tuple

import numpy as np

from lib.quantification.radiomics import radiomics_config_signature

AUXILIARY_STEM_TOKENS = (
    "combined_seg", "totalseg", "total_seg", "body_trunc", "body_extremities",
    "body", "skin", "ct", "vertebra", "vertebrae", "kidney", "ureter",
    "urinary", "bladder", "spine",
)


def is_auxiliary_segment_stem(stem: str) -> bool:
    """Files that are usually masks/aux rather than target organs."""
    nl = stem.lower()
    return any(tok in nl for tok in AUXILIARY_STEM_TOKENS)


def default_excel_label(stem: str) -> str:
    label = stem
    for suffix in (".seg",):
        if label.lower().endswith(suffix):
            label = label[: -len(suffix)]
    # Keep Excel segment names stable whether raw or *_processed was used
    if label.lower().endswith("_processed"):
        label = label[: -len("_processed")]
    return label


def parse_batch_base_name(base: str) -> Tuple[str, str]:
    """'MSP0001_2025-07-09' -> ('MSP0001', '2025-07-09'). Falls back to (base, '')."""
    m = re.match(r"(?P<sid>.+?)_(?P<date>\d{4}-\d{2}-\d{2})", base)
    if m:
        return m.group("sid"), m.group("date")
    return base, ""


def scan_batch_dataset(root: str) -> dict:
    """Discover subjects (Segments/<base>_Seg) and candidate segment stems."""
    if not root or not os.path.isdir(root):
        raise ValueError(f"Root folder not found: {root}")

    seg_root = os.path.join(root, "Segments")
    if not os.path.isdir(seg_root):
        raise ValueError(
            f"Expected a 'Segments' subfolder under: {root}\n"
            "Layout: <root>/PET, <root>/CT, <root>/Segments/<base>_Seg/"
        )

    subject_dirs = sorted(
        d for d in os.listdir(seg_root)
        if d.endswith("_Seg") and os.path.isdir(os.path.join(seg_root, d))
    )
    if not subject_dirs:
        raise ValueError(f"No '*_Seg' subject folders found in: {seg_root}")

    first_dir = os.path.join(seg_root, subject_dirs[0])
    segment_stems = []
    for f in sorted(os.listdir(first_dir)):
        stem = None
        if f.endswith(".seg.nrrd"):
            stem = f[: -len(".seg.nrrd")]
        elif f.endswith(".nii.gz"):
            stem = f[: -len(".nii.gz")]
        elif f.endswith(".nii"):
            stem = f[: -len(".nii")]
        elif f.endswith(".nrrd"):
            stem = f[: -len(".nrrd")]
        if stem is not None:
            segment_stems.append(stem)

    return {
        "subjectCount": len(subject_dirs),
        "subjectDirs": subject_dirs,
        "segmentStems": segment_stems,
    }


def find_batch_segment_file(seg_dir: str, stem: str) -> Optional[str]:
    for ext in (".seg.nrrd", ".nii.gz", ".nii", ".nrrd"):
        candidate = os.path.join(seg_dir, stem + ext)
        if os.path.exists(candidate):
            return candidate
    return None


def resolve_batch_segment_file(
    seg_dir: str,
    stem: str,
    *,
    prefer_processed: bool = True,
    processed_suffix: str = "_processed",
) -> Optional[str]:
    """
    Resolve a mask path for quantification.

    If ``prefer_processed`` and ``{stem}_processed`` exists, use that;
    otherwise fall back to ``stem``. Stems that already end with the
    processed suffix are used as-is.
    """
    stem = str(stem).replace(".nii.gz", "").replace(".nii", "").strip()
    if not stem:
        return None
    candidates: list[str] = []
    already_processed = stem.lower().endswith(processed_suffix.lower())
    if prefer_processed and not already_processed:
        candidates.append(stem + processed_suffix)
    candidates.append(stem)
    for cand in candidates:
        hit = find_batch_segment_file(seg_dir, cand)
        if hit:
            return hit
    return None


QUANTIFICATION_SHEET = "Quantification"
RADIOMICS_SHEET = "Radiomics"
# Legacy sheet name used before Quantification/Radiomics split
LEGACY_DATA_SHEET = "Data"

QUANTIFICATION_BASE_COLS = [
    "subject_id", "patient_id", "scan_date", "segment", "source_file",
    "volume_mL", "suv_mean", "suv_max", "suv_peak", "tlg",
    "radiomics_status", "computation_signature", "status", "computed_at",
]

RADIOMICS_META_COLS = [
    "subject_id", "patient_id", "scan_date", "segment", "source_file",
    "radiomics_status", "computation_signature", "status", "computed_at",
]


def is_radiomics_column(name: Any) -> bool:
    return str(name).startswith("rad_")


def quantification_sheet_name(sheetnames: Sequence[str]) -> Optional[str]:
    """Prefer Quantification; fall back to legacy Data."""
    names = list(sheetnames)
    if QUANTIFICATION_SHEET in names:
        return QUANTIFICATION_SHEET
    if LEGACY_DATA_SHEET in names:
        return LEGACY_DATA_SHEET
    return None


def existing_batch_keys(
    output_file: str,
    required_segments: Optional[Sequence[str]] = None,
    required_signature: Optional[str] = None,
) -> Set[Tuple[str, str]]:
    """Return subjects that are complete for this exact computation."""
    keys: Set[Tuple[str, str]] = set()
    if not output_file or not os.path.isfile(output_file):
        return keys
    try:
        import openpyxl
    except ModuleNotFoundError:
        return keys
    try:
        wb = openpyxl.load_workbook(output_file, read_only=True)
    except Exception:
        return keys
    sheet = quantification_sheet_name(wb.sheetnames)
    if sheet is None:
        return keys

    requested = {
        str(segment).strip() for segment in (required_segments or [])
        if str(segment).strip()
    }
    completed: Dict[Tuple[str, str], Set[str]] = {}
    ws = wb[sheet]
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        rec = dict(zip(header, row))
        if str(rec.get("status", "")).strip().lower() != "done":
            continue
        if required_signature is not None:
            found_signature = str(
                rec.get("computation_signature", "")
            ).strip()
            if found_signature != required_signature:
                continue
        key = (
            str(rec.get("subject_id", "")).strip(),
            str(rec.get("scan_date", "")).strip(),
        )
        segment = str(rec.get("segment", "")).strip()
        completed.setdefault(key, set()).add(segment)

    for key, segments in completed.items():
        if not requested or requested.issubset(segments):
            keys.add(key)
    return keys


def split_quant_and_radiomics_row(row: dict) -> Tuple[dict, dict]:
    """Split one combined result dict into quantification + radiomics dicts."""
    quant = {
        k: v for k, v in row.items()
        if not is_radiomics_column(k)
    }
    rad = {
        k: row.get(k, "")
        for k in RADIOMICS_META_COLS
    }
    for k, v in row.items():
        if is_radiomics_column(k):
            rad[k] = v
    return quant, rad


def row_has_radiomics(row: dict) -> bool:
    if any(is_radiomics_column(k) for k in row.keys()):
        return True
    return str(row.get("radiomics_status", "")).strip().lower() == "done"

def _normalise_segment_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(label).lower())


def segment_role(label: str) -> str:
    """Infer only the roles needed for optional cross-ROI derived metrics."""
    raw = str(label).lower()
    compact = _normalise_segment_label(raw)
    tokens = set(re.findall(r"[a-z0-9]+", raw))

    if any(token in compact for token in (
        "bloodpool", "blood", "femoralvessel", "femoralvein",
        "femoralartery", "aorta", "vessel"
    )):
        return "blood_pool"

    if "psoas" in compact:
        is_left = (
            bool(tokens.intersection({"left", "lt", "l"}))
            or compact.endswith("left")
            or compact.endswith("lt")
            or compact.endswith("l")
        )
        is_right = (
            bool(tokens.intersection({"right", "rt", "r"}))
            or compact.endswith("right")
            or compact.endswith("rt")
            or compact.endswith("r")
        )
        if is_left and not is_right:
            return "psoas_left"
        if is_right and not is_left:
            return "psoas_right"

    return "other"


def safe_asymmetry(left: Any, right: Any) -> Optional[float]:
    if left in (None, "") or right in (None, ""):
        return None
    try:
        left_f = float(left)
        right_f = float(right)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(left_f) or not math.isfinite(right_f):
        return None
    denominator = (abs(left_f) + abs(right_f)) / 2.0
    if denominator <= 1e-12:
        return None
    return abs(left_f - right_f) / denominator


def cross_roi_derived_by_subject(all_data: Sequence[dict]) -> dict:
    """Return Summary-only features that require two or more ROI rows."""
    grouped: Dict[tuple, list] = {}
    for rec in all_data:
        if str(rec.get("status", "")).strip().lower() != "done":
            continue
        key = (
            rec.get("subject_id", ""),
            rec.get("patient_id", ""),
            rec.get("scan_date", ""),
        )
        grouped.setdefault(key, []).append(rec)

    output = {}
    for key, records in grouped.items():
        derived = {}

        blood_medians = []
        for rec in records:
            if segment_role(rec.get("segment", "")) != "blood_pool":
                continue
            value = rec.get("rad_firstorder_Median")
            try:
                value = float(value)
            except (TypeError, ValueError):
                continue
            if math.isfinite(value) and abs(value) > 1e-12:
                blood_medians.append(value)

        if blood_medians:
            blood_median = float(np.mean(blood_medians))
            for rec in records:
                if segment_role(rec.get("segment", "")) == "blood_pool":
                    continue
                p90 = rec.get("rad_firstorder_90Percentile")
                try:
                    p90 = float(p90)
                except (TypeError, ValueError):
                    continue
                if not math.isfinite(p90):
                    continue
                segment = str(rec.get("segment", "")).strip()
                if segment:
                    derived[
                        f"{segment}_derived_P90ToBloodPoolMedian"
                    ] = p90 / blood_median

        psoas = {}
        for rec in records:
            role = segment_role(rec.get("segment", ""))
            if role in ("psoas_left", "psoas_right"):
                psoas[role] = rec

        left = psoas.get("psoas_left")
        right = psoas.get("psoas_right")
        if left is not None and right is not None:
            feature_map = {
                "P90": "rad_firstorder_90Percentile",
                "Entropy": "rad_firstorder_Entropy",
                "SAHGLE": "rad_glszm_SmallAreaHighGrayLevelEmphasis",
            }
            for short_name, feature_name in feature_map.items():
                value = safe_asymmetry(
                    left.get(feature_name), right.get(feature_name)
                )
                if value is not None:
                    derived[
                        f"derived_Psoas_{short_name}_Asymmetry"
                    ] = value

        output[key] = derived

    return output


def batch_error_row(
    subject_id: str,
    scan_date: str,
    segment: str,
    status: str,
    stem: str,
) -> dict:
    return {
        "subject_id": subject_id,
        "patient_id": "",
        "scan_date": scan_date,
        "segment": segment,
        "source_file": stem,
        "status": status,
    }


def computation_signature(
    metrics_options: Optional[dict],
    radiomics_options: Optional[dict],
) -> str:
    opts = metrics_options or {}
    metrics = ",".join(
        key for key in ("mean", "max", "peak", "tlg", "volume")
        if opts.get(key, False)
    ) or "none"
    return f"metrics={metrics}|{radiomics_config_signature(radiomics_options)}"


def parse_quantitative_indices_results(
    name_value_pairs: Iterable[Tuple[Any, Any]],
) -> dict:
    """Parse QuantitativeIndicesCLI-style (rawName, rawValue) pairs."""
    name_map = {
        "Mean": "suv_mean",
        "Max": "suv_max",
        "Peak": "suv_peak",
        "TLG": "tlg",
        "Volume": "volume_mL",
    }

    results = {}
    for raw_name, raw_value in name_value_pairs:
        if raw_value in (None, "", "--"):
            continue

        clean_name = str(raw_name).replace("_s", "").replace("_", " ").strip()
        if clean_name not in name_map:
            continue

        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue

        if math.isnan(value):
            continue

        results[name_map[clean_name]] = value

    return results


def _append_rows_to_sheet(wb, sheet_name: str, rows: Sequence[dict], base_cols: list[str]) -> None:
    """Create/append a sheet with schema merge; never keep rad_* on Quantification."""
    extra_cols: list[str] = []
    for row in rows:
        for key in row.keys():
            if key in base_cols or key in extra_cols:
                continue
            if sheet_name == QUANTIFICATION_SHEET and is_radiomics_column(key):
                continue
            if sheet_name == RADIOMICS_SHEET and key not in RADIOMICS_META_COLS and not is_radiomics_column(key):
                continue
            extra_cols.append(key)

    if sheet_name == RADIOMICS_SHEET:
        current_cols = list(dict.fromkeys(
            RADIOMICS_META_COLS[:-2]
            + sorted(c for c in extra_cols if is_radiomics_column(c))
            + RADIOMICS_META_COLS[-2:]
        ))
    else:
        # status / computed_at last
        current_cols = list(dict.fromkeys(
            [c for c in base_cols if c not in ("status", "computed_at")]
            + sorted(extra_cols)
            + ["status", "computed_at"]
        ))

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        existing_header = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        existing_header = [
            str(h).strip() for h in existing_header
            if h is not None and str(h).strip() != ""
        ]
        if sheet_name == QUANTIFICATION_SHEET:
            existing_header = [h for h in existing_header if not is_radiomics_column(h)]
        data_cols = list(dict.fromkeys(existing_header + current_cols))
    else:
        if (
            sheet_name == QUANTIFICATION_SHEET
            and wb.active
            and wb.active.title in ("Sheet", "Results", LEGACY_DATA_SHEET)
            and LEGACY_DATA_SHEET not in wb.sheetnames
            and QUANTIFICATION_SHEET not in wb.sheetnames
        ):
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        data_cols = list(dict.fromkeys(current_cols))
        for ci, col in enumerate(data_cols, 1):
            ws.cell(row=1, column=ci, value=col)

    for ci, col in enumerate(data_cols, 1):
        ws.cell(row=1, column=ci, value=col)
    for row in rows:
        ws.append([row.get(c, "") for c in data_cols])


def save_batch_rows_to_excel(
    rows: Sequence[dict],
    output_file: str,
    append: bool = True,
) -> str:
    """
    Write Quantification (+ optional Radiomics) sheets, then rebuild Summary.

    SUV / volume / TLG go to ``Quantification``.
    ``rad_*`` features go to ``Radiomics`` (created only when radiomics ran).
    ``Summary`` is a subject×segment pivot of Quantification metrics only.
    """
    if not rows:
        raise ValueError("No batch rows to save.")
    output_file = os.path.abspath(output_file)
    if not output_file.lower().endswith(".xlsx"):
        output_file += ".xlsx"

    try:
        import openpyxl
    except ModuleNotFoundError as e:
        raise ImportError(
            "openpyxl is required to save batch Excel files. "
            "Install with: pip install openpyxl"
        ) from e

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    quant_rows: list[dict] = []
    rad_rows: list[dict] = []
    for row in rows:
        row = dict(row)
        row.setdefault("computed_at", now)
        quant, rad = split_quant_and_radiomics_row(row)
        quant_rows.append(quant)
        if row_has_radiomics(row):
            rad_rows.append(rad)

    if append and os.path.isfile(output_file):
        wb = openpyxl.load_workbook(output_file)
        for stale in ("Sheet", "Sheet1", "Results"):
            if stale in wb.sheetnames:
                del wb[stale]
        # Migrate legacy "Data" → "Quantification"
        if (
            LEGACY_DATA_SHEET in wb.sheetnames
            and QUANTIFICATION_SHEET not in wb.sheetnames
        ):
            wb[LEGACY_DATA_SHEET].title = QUANTIFICATION_SHEET
    else:
        wb = openpyxl.Workbook()

    _append_rows_to_sheet(
        wb, QUANTIFICATION_SHEET, quant_rows, QUANTIFICATION_BASE_COLS
    )
    if rad_rows:
        _append_rows_to_sheet(
            wb, RADIOMICS_SHEET, rad_rows, RADIOMICS_META_COLS
        )

    # Rebuild Summary from Quantification only
    ws = wb[QUANTIFICATION_SHEET]
    data_cols = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    data_cols = [c for c in data_cols if c is not None]
    header = data_cols
    all_data = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(data_cols) + 1)]
        if any(v is not None and v != "" for v in vals):
            all_data.append(dict(zip(header, vals)))

    def _is_done(rec):
        return str(rec.get("status", "")).strip().lower() == "done"

    metric_cols = [
        c for c in data_cols
        if c in ("volume_mL", "suv_mean", "suv_max", "suv_peak", "tlg")
    ]
    seen_segs = list(dict.fromkeys(
        str(r.get("segment", "")).strip()
        for r in all_data if _is_done(r) and str(r.get("segment", "")).strip()
    ))

    pivot = {}
    for r in all_data:
        if not _is_done(r):
            continue
        key = (r.get("subject_id", ""), r.get("patient_id", ""), r.get("scan_date", ""))
        pivot.setdefault(key, {})
        seg = str(r.get("segment", "")).strip()
        for m in metric_cols:
            val = r.get(m)
            if val not in (None, ""):
                pivot[key][f"{seg}_{m}"] = val

    cross_roi_derived = cross_roi_derived_by_subject(all_data)
    for key, values in cross_roi_derived.items():
        pivot.setdefault(key, {}).update(values)

    id_cols = ["subject_id", "patient_id", "scan_date"]
    metric_order = [f"{seg}_{m}" for seg in seen_segs for m in metric_cols]
    derived_order = sorted({
        col
        for values in pivot.values()
        for col in values.keys()
        if col.startswith("derived_") or "_derived_" in col
    })
    metric_order.extend(
        col for col in derived_order if col not in metric_order
    )
    sum_cols = id_cols + metric_order

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_s = wb.create_sheet("Summary")
    for ci, col in enumerate(sum_cols, 1):
        ws_s.cell(row=1, column=ci, value=col)
    for ri, (subj, pid, date) in enumerate(sorted(pivot), start=2):
        ws_s.cell(ri, 1, subj)
        ws_s.cell(ri, 2, pid)
        ws_s.cell(ri, 3, date)
        data = pivot[(subj, pid, date)]
        for ci, col in enumerate(metric_order, start=4):
            ws_s.cell(ri, ci, data.get(col, ""))

    wb.save(output_file)
    return output_file


def save_qc_rows_to_excel(
    qc_rows: Sequence[dict],
    output_file: str,
    baseline_suv_max: Optional[float] = None,
) -> str:
    """Write the QC table to a 'QC' sheet (append/overwrite that one sheet only)."""
    if not qc_rows:
        raise ValueError("No QC rows to save.")
    if not output_file:
        raise ValueError("No output Excel file selected.")

    output_file = os.path.abspath(output_file)
    if not output_file.lower().endswith(".xlsx"):
        output_file += ".xlsx"

    try:
        import openpyxl
    except ModuleNotFoundError as e:
        raise ImportError(
            "openpyxl is required to save QC Excel files. "
            "Install with: pip install openpyxl"
        ) from e

    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
        default = wb.active
        if default is not None and default.title == "Sheet" and default.max_row == 1:
            wb.remove(default)

    if "QC" in wb.sheetnames:
        del wb["QC"]
    ws = wb.create_sheet("QC")

    columns = [
        "roi_node", "segment", "suv_mean", "suv_max", "suv_peak",
        "max_over_mean", "baseline_suv_max", "delta_suv_max_pct",
        "ras_r", "ras_a", "ras_s", "flag", "computed_at",
    ]
    for ci, col in enumerate(columns, start=1):
        ws.cell(row=1, column=ci, value=col)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for ri, row in enumerate(qc_rows, start=2):
        ras = row.get("rasHotspot") or [None, None, None]
        ws.cell(ri, 1, row.get("roiNodeName", ""))
        ws.cell(ri, 2, row.get("segmentName", ""))
        ws.cell(ri, 3, row.get("suv_mean"))
        ws.cell(ri, 4, row.get("suv_max"))
        ws.cell(ri, 5, row.get("suv_peak"))
        ws.cell(ri, 6, row.get("ratio"))
        ws.cell(ri, 7, baseline_suv_max)
        ws.cell(ri, 8, row.get("deltaMaxPct"))
        ws.cell(ri, 9, ras[0])
        ws.cell(ri, 10, ras[1])
        ws.cell(ri, 11, ras[2])
        ws.cell(ri, 12, row.get("flag", ""))
        ws.cell(ri, 13, now)

    wb.save(output_file)
    return output_file
