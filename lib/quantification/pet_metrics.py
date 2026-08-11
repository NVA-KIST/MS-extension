"""
PET quantification helpers that do NOT need Slicer.

Slicer-only pieces (DICOM load, registration, QuantitativeIndicesCLI) stay in Logic.
"""
from __future__ import annotations

import math
import os
from datetime import datetime
from typing import Any, Optional


def error_row(
    subject_id: str,
    scan_date: str,
    segment: str,
    status: str,
    patient_id: str = "",
) -> dict[str, Any]:
    return {
        "subject_id": subject_id,
        "patient_id": patient_id,
        "scan_date": scan_date,
        "segment": segment,
        "status": status,
    }


def hhmmss_to_seconds(t: str) -> int:
    t = str(t).strip()[-6:]
    try:
        return int(t[0:2]) * 3600 + int(t[2:4]) * 60 + int(t[4:6])
    except (ValueError, IndexError):
        return 0


def compute_suvbw_factor(
    weight_kg: float,
    dose_bq: float,
    injection_time: str,
    acquisition_time: str,
    half_life_s: float,
    decay_correction: str = "START",
) -> float:
    """
    SUVbw factor so that  SUV = Bq/mL * factor.
    factor = weight_g / decay_corrected_dose_Bq
    """
    weight_g = float(weight_kg) * 1000.0
    decay_corr = str(decay_correction).strip().upper()
    if decay_corr == "ADMIN":
        decay_corrected_dose_bq = float(dose_bq)
    else:
        inj_s = hhmmss_to_seconds(injection_time)
        acq_s = hhmmss_to_seconds(acquisition_time)
        if acq_s < inj_s:
            acq_s += 86400
        decay_time_s = acq_s - inj_s
        decay_corrected_dose_bq = float(dose_bq) * (2.0 ** (-decay_time_s / float(half_life_s)))

    if decay_corrected_dose_bq <= 0:
        raise ValueError(f"Invalid decay_corrected_dose={decay_corrected_dose_bq}")
    return weight_g / decay_corrected_dose_bq


def suvbw_factor_from_dicom_folder(dicom_folder: str) -> tuple[float, dict[str, Any]]:
    """
    Read PET DICOM headers and return (SUVbw factor, meta).

    If Units is already ``SUV``, factor is ``1.0``.
    Prefers files that contain RadiopharmaceuticalInformationSequence.
    """
    import pydicom

    candidates: list[str] = []
    for root, _, files in os.walk(dicom_folder):
        for f in sorted(files):
            fp = os.path.join(root, f)
            try:
                ds = pydicom.dcmread(fp, stop_before_pixels=True)
            except Exception:
                continue
            candidates.append(fp)
            # Prefer a true PET image slice with radio / dose metadata
            if (0x0054, 0x0016) in ds and getattr(ds, "PatientWeight", None) is not None:
                candidates.insert(0, fp)
                break

    if not candidates:
        raise ValueError(f"No readable DICOM file found in {dicom_folder}")

    ds = None
    first_dcm = None
    for fp in candidates:
        try:
            cand = pydicom.dcmread(fp, stop_before_pixels=True)
        except Exception:
            continue
        if (0x0054, 0x0016) in cand:
            ds = cand
            first_dcm = fp
            break
        if ds is None:
            ds = cand
            first_dcm = fp

    if ds is None or first_dcm is None:
        raise ValueError(f"No readable DICOM file found in {dicom_folder}")

    units = str(getattr(ds, "Units", "BQML")).strip().upper()
    meta: dict[str, Any] = {"units": units, "dicom_file": first_dcm}
    if units == "SUV":
        meta["skipped"] = True
        return 1.0, meta

    try:
        weight_kg = float(str(getattr(ds, "PatientWeight", None)))
        if math.isnan(weight_kg):
            raise ValueError("PatientWeight is NaN")
    except (TypeError, ValueError) as e:
        raise ValueError(f"PatientWeight missing or unreadable in DICOM header: {e}") from e

    try:
        radio_seq = ds[0x0054, 0x0016][0]
    except (KeyError, IndexError) as e:
        raise ValueError(
            "RadiopharmaceuticalInformationSequence (0054,0016) missing - cannot compute SUV"
        ) from e

    try:
        dose_bq = float(str(radio_seq[0x0018, 0x1074].value))
    except (KeyError, ValueError) as e:
        raise ValueError(
            "RadionuclideTotalDose (0018,1074) missing from RadiopharmaceuticalInformationSequence"
        ) from e

    inj_dt = getattr(radio_seq, "RadiopharmaceuticalStartDateTime", None)
    inj_t = getattr(radio_seq, "RadiopharmaceuticalStartTime", None)
    if inj_dt:
        inj_str = str(inj_dt).split(".")[0]
    elif inj_t:
        inj_str = str(inj_t).split(".")[0]
    else:
        raise ValueError(
            "No injection time found in RadiopharmaceuticalInformationSequence"
        )

    try:
        half_life_s = float(str(radio_seq[0x0018, 0x1075].value))
    except (KeyError, ValueError):
        try:
            half_life_s = float(str(ds[0x0009, 0x103F].value))
        except (KeyError, ValueError):
            half_life_s = 6586.2  # F-18 default

    acq_t_raw = str(
        getattr(ds, "AcquisitionTime", "") or getattr(ds, "SeriesTime", "")
    ).split(".")[0].strip()
    decay_corr = str(getattr(ds, "DecayCorrection", "START")).strip().upper()

    factor = compute_suvbw_factor(
        weight_kg=weight_kg,
        dose_bq=dose_bq,
        injection_time=inj_str,
        acquisition_time=acq_t_raw,
        half_life_s=half_life_s,
        decay_correction=decay_corr,
    )
    meta.update(
        {
            "weight_kg": weight_kg,
            "dose_bq": dose_bq,
            "injection_time": inj_str,
            "acquisition_time": acq_t_raw,
            "half_life_s": half_life_s,
            "decay_correction": decay_corr,
            "suv_factor": factor,
        }
    )
    return factor, meta


def save_excel(rows: list[dict], output_file: str, append: bool = True, log=None) -> None:
    """
    Write Quantification + Summary sheets via openpyxl.
    Port of PETCTQuantAnalysis_v2Logic._saveExcel.
    """
    import openpyxl

    from lib.quantification.biomarker_batch import (
        LEGACY_DATA_SHEET,
        QUANTIFICATION_SHEET,
        quantification_sheet_name,
    )

    def _log(msg: str):
        if log:
            log("info", msg)

    DATA_COLS = [
        "subject_id", "patient_id", "scan_date",
        "segment", "source_file",
        "volume_mL", "suv_mean", "suv_max", "suv_peak", "tlg",
        "status", "computed_at",
    ]
    METRIC_COLS = ["volume_mL", "suv_mean", "suv_max", "suv_peak", "tlg"]

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row in rows:
        row["computed_at"] = now_str

    if append and os.path.isfile(output_file):
        wb = openpyxl.load_workbook(output_file)
        for stale in ("Sheet", "Sheet1", "Results"):
            if stale in wb.sheetnames:
                del wb[stale]
        if (
            LEGACY_DATA_SHEET in wb.sheetnames
            and QUANTIFICATION_SHEET not in wb.sheetnames
        ):
            wb[LEGACY_DATA_SHEET].title = QUANTIFICATION_SHEET
    else:
        wb = openpyxl.Workbook()

    sheet = quantification_sheet_name(wb.sheetnames)
    if sheet is not None:
        ws = wb[sheet]
        if sheet != QUANTIFICATION_SHEET:
            ws.title = QUANTIFICATION_SHEET
    else:
        if wb.active and wb.active.title in ("Sheet", "Results"):
            ws = wb.active
            ws.title = QUANTIFICATION_SHEET
        else:
            ws = wb.create_sheet(QUANTIFICATION_SHEET, 0)
        for ci, col in enumerate(DATA_COLS, 1):
            ws.cell(row=1, column=ci, value=col)

    for row in rows:
        ws.append([row.get(c, "") for c in DATA_COLS])

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    all_data = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in vals):
            all_data.append(dict(zip(header, vals)))

    seen_segs = list(
        dict.fromkeys(r.get("segment", "") for r in all_data if r.get("status") == "done")
    )

    pivot: dict[tuple, dict] = {}
    for r in all_data:
        if r.get("status") != "done":
            continue
        key = (r.get("subject_id", ""), r.get("patient_id", ""), r.get("scan_date", ""))
        pivot.setdefault(key, {})
        seg = r.get("segment", "")
        for m in METRIC_COLS:
            col_name = f"{seg}_{m}"
            val = r.get(m)
            if val not in (None, ""):
                pivot[key][col_name] = val

    id_cols = ["subject_id", "patient_id", "scan_date"]
    metric_order = [f"{seg}_{m}" for seg in seen_segs for m in METRIC_COLS]
    sum_cols = id_cols + metric_order

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_s = wb.create_sheet("Summary")
    for ci, col in enumerate(sum_cols, 1):
        ws_s.cell(row=1, column=ci, value=col)
    for ri, (subj, pid, date) in enumerate(sorted(pivot), start=2):
        ws_s.cell(ri, 1, subj)
        ws_s.cell(ri, 2, pid)
        ws_s.cell(ri, 3, date)
        data = pivot[(subj, pid, date)]
        for ci, col in enumerate(metric_order, start=4):
            ws_s.cell(ri, ci, data.get(col, ""))

    wb.save(output_file)
    _log(f"save_excel: saved {len(rows)} new row(s) → {output_file}")


def compute_segment_metrics(
    pet_arr,
    mask_arr,
    spacing_mm,
    *,
    metrics: Optional[dict] = None,
) -> dict[str, Any]:
    """
    Pure-numpy SUV-style metrics (no Slicer QuantitativeIndicesCLI).

    ``pet_arr`` / ``mask_arr`` must share the same shape (ZYX or XYZ — just match).
    ``spacing_mm`` is (sz, sy, sx) or (sx, sy, sz) consistent with that array order
    for volume — product of the three spacings is used.

    Peak ≈ mean of 3×3×3 neighbourhood around the max voxel (QC-style).
    TLG ≈ suv_mean * volume_mL.
    """
    import numpy as np

    metrics = metrics or {
        "mean": True, "max": True, "peak": True, "tlg": True, "volume": True
    }
    pet = np.asarray(pet_arr, dtype=np.float64)
    mask = np.asarray(mask_arr) > 0
    if pet.shape != mask.shape:
        raise ValueError(f"PET shape {pet.shape} != mask shape {mask.shape}")
    if not np.any(mask):
        raise ValueError("Mask is empty")

    vals = pet[mask]
    out: dict[str, Any] = {}
    sp = [float(s) for s in spacing_mm]
    vox_ml = (sp[0] * sp[1] * sp[2]) / 1000.0
    vol_ml = float(mask.sum()) * vox_ml

    if metrics.get("volume", True):
        out["volume_mL"] = vol_ml
    if metrics.get("mean", True):
        out["suv_mean"] = float(np.mean(vals))
    if metrics.get("max", True):
        out["suv_max"] = float(np.max(vals))

    if metrics.get("peak", True):
        masked = np.where(mask, pet, -np.inf)
        max_idx = np.unravel_index(int(np.argmax(masked)), masked.shape)
        slices = []
        for ax, ix in enumerate(max_idx):
            lo = max(0, int(ix) - 1)
            hi = min(pet.shape[ax], int(ix) + 2)
            slices.append(slice(lo, hi))
        out["suv_peak"] = float(np.mean(pet[tuple(slices)]))

    if metrics.get("tlg", True):
        mean = out.get("suv_mean", float(np.mean(vals)))
        out["tlg"] = float(mean * vol_ml)

    return out


def run_batch_quantification(
    root: str,
    segment_stems: list[str],
    output_file: str,
    *,
    metrics: Optional[dict] = None,
    radiomics_options: Optional[dict] = None,
    append: bool = True,
    skip_done: bool = True,
    prefer_processed: bool = True,
    limit: int = 0,
    log=None,
) -> dict[str, Any]:
    """
    File-based batch quantification (Python IDE / CLI — no Slicer).

    For each Segments/<base>_Seg subject:
      - load PET NIfTI (PET_NIfTI/ or convert DICOM)
      - for each stem, load mask NIfTI (prefer ``*_processed.nii.gz``)
      - compute numpy SUV metrics (+ optional PyRadiomics)
      - write Excel via ``save_batch_rows_to_excel``
      (Quantification sheet + Radiomics sheet when enabled)
    """
    from datetime import datetime as _dt
    from pathlib import Path

    import numpy as np

    from lib.processing.postprocess import load_pet_array
    from lib.quantification.biomarker_batch import (
        batch_error_row,
        computation_signature,
        default_excel_label,
        existing_batch_keys,
        parse_batch_base_name,
        resolve_batch_segment_file,
        save_batch_rows_to_excel,
        scan_batch_dataset,
    )
    from lib.quantification.radiomics import (
        extract_radiomics_from_paths,
        is_radiomics_enabled,
    )
    from lib.processing.dilate import resample_to_target

    def _log(msg: str):
        print(msg)
        if log and hasattr(log, "info"):
            log.info(msg)

    root_p = Path(root)
    scan = scan_batch_dataset(str(root_p))
    subjects = scan["subjectDirs"]
    if limit > 0:
        subjects = subjects[:limit]

    metrics = metrics or {
        "mean": True, "max": True, "peak": True, "tlg": True, "volume": True
    }
    radiomics_options = radiomics_options or {}
    do_rad = is_radiomics_enabled(radiomics_options)
    sig = computation_signature(metrics, radiomics_options)
    labels = [default_excel_label(s) for s in segment_stems]
    done = (
        existing_batch_keys(output_file, required_segments=labels, required_signature=sig)
        if append and skip_done
        else set()
    )

    rows: list[dict] = []
    processed = skipped = errors = 0
    seg_root = root_p / "Segments"

    for idx, seg_folder in enumerate(subjects, 1):
        base = seg_folder[: -len("_Seg")]
        subject_id, scan_date = parse_batch_base_name(base)
        if skip_done and (subject_id, scan_date) in done:
            _log(f"[{idx}/{len(subjects)}] {base}: skip (done)")
            skipped += 1
            continue

        _log(f"\n[{idx}/{len(subjects)}] {base}")
        seg_dir = seg_root / seg_folder
        try:
            nii = root_p / "PET_NIfTI" / f"{base}_PET.nii.gz"
            dcm = root_p / "PET" / f"{base}_PET"
            pet_arr, pet_aff, _ = load_pet_array(
                pet_path=nii if nii.is_file() else None,
                pet_dicom_dir=dcm if dcm.is_dir() else None,
                pet_nii_out=nii,
            )
            # Convert Bq/mL -> SUVbw using PET DICOM headers (required for
            # meaningful metrics and for PyRadiomics binWidth).
            if dcm.is_dir():
                try:
                    suv_factor, suv_meta = suvbw_factor_from_dicom_folder(str(dcm))
                    if suv_meta.get("skipped"):
                        _log("  [SUV] DICOM already in SUV units")
                    else:
                        import numpy as np

                        pet_arr = (
                            np.asarray(pet_arr, dtype=np.float32) * float(suv_factor)
                        ).astype(np.float32)
                        _log(
                            f"  [SUV] factor={suv_factor:.6g}  "
                            f"weight={suv_meta.get('weight_kg')} kg  "
                            f"PET max after={float(np.max(pet_arr)):.4g}"
                        )
                except Exception as e:
                    _log(f"  [SUV] WARN conversion failed ({e}); metrics may be Bq/mL")
            else:
                _log("  [SUV] WARN no PET DICOM folder - assuming NIfTI is already SUV")
            # spacing from affine diagonal (approx mm)
            sp = tuple(float(abs(pet_aff[i, i])) for i in range(3))
            # pet_arr is ZYX; affine diag is often XYZ → reorder
            spacing_zyx = (sp[2], sp[1], sp[0])

            for stem in segment_stems:
                label = default_excel_label(stem)
                seg_path = resolve_batch_segment_file(
                    str(seg_dir),
                    stem,
                    prefer_processed=prefer_processed,
                )
                if seg_path is None:
                    rows.append(
                        batch_error_row(subject_id, scan_date, label, "missing_file", stem)
                    )
                    continue
                _log(f"  [mask] {label} ← {os.path.basename(seg_path)}")

                try:
                    import nibabel as nib

                    m_img = nib.load(seg_path)
                    m_xyz = np.asarray(m_img.dataobj)
                    m_zyx = np.transpose(m_xyz, (2, 1, 0)) if m_xyz.ndim == 3 else m_xyz
                    if m_zyx.shape != pet_arr.shape:
                        m_zyx = resample_to_target(
                            m_zyx, m_img.affine, pet_arr.shape, pet_aff
                        )

                    results = compute_segment_metrics(
                        pet_arr, m_zyx, spacing_zyx, metrics=metrics
                    )
                except Exception as e:
                    err = f"{type(e).__name__}: {e}" if str(e) else type(e).__name__
                    rows.append(
                        batch_error_row(
                            subject_id,
                            scan_date,
                            label,
                            f"metric_error:{err[:80]}",
                            stem,
                        )
                    )
                    _log(f"  [ERR] {stem}: {err}")
                    continue

                rad_status = "not_run"
                if do_rad:
                    try:
                        import tempfile

                        with tempfile.TemporaryDirectory() as td:
                            pet_p = os.path.join(td, "pet.nii.gz")
                            msk_p = os.path.join(td, "mask.nii.gz")
                            pet_xyz = np.transpose(pet_arr, (2, 1, 0)).astype(np.float32)
                            msk_xyz = np.transpose(
                                (m_zyx > 0).astype(np.uint8), (2, 1, 0)
                            )
                            nib.save(nib.Nifti1Image(pet_xyz, pet_aff), pet_p)
                            nib.save(nib.Nifti1Image(msk_xyz, pet_aff), msk_p)
                            rad = extract_radiomics_from_paths(
                                pet_p, msk_p, radiomics_options, label=1
                            )
                            results.update(rad)
                            rad_status = "done"
                    except Exception as e:
                        err = f"{type(e).__name__}: {e}" if str(e) else type(e).__name__
                        rad_status = f"radiomics_error:{err[:80]}"
                        _log(f"  [RAD] {label}: {err}")

                row = {
                    "subject_id": subject_id,
                    "patient_id": "",
                    "scan_date": scan_date,
                    "segment": label,
                    "source_file": os.path.basename(seg_path),
                    "radiomics_status": rad_status,
                    "computation_signature": sig,
                    "status": "done",
                    "computed_at": _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                row.update(results)
                rows.append(row)
                _log(f"  [OK] {label} SUVmax={results.get('suv_max', 'NA')}")

            processed += 1
        except Exception as e:
            errors += 1
            rows.append(
                batch_error_row(subject_id, scan_date, "", f"subject_error:{str(e)[:80]}", "")
            )
            _log(f"  [ERR] subject: {e}")

    saved = None
    if rows:
        saved = save_batch_rows_to_excel(rows, output_file, append=append)
    return {
        "processed": processed,
        "skipped": skipped,
        "errors": errors,
        "rowCount": len(rows),
        "savedPath": saved,
    }
